"""Build the researcher-facing review workbook for common-pronunciation r2.

This workbook is an approval interface, not an approval action.  All 27 rows
start as ``pending``.  Suggested phones, alternatives, source provenance, and
staged WAV links are visible, while the final selected phone is formula-driven
from the researcher's explicit decision.

The user previously requested an openpyxl fallback for review workbooks, so
this standalone artifact intentionally uses openpyxl rather than a live Excel
session.
"""
from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, __version__ as openpyxl_version, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from pipeline_common import (  # noqa: E402
    atomic_write_json,
    file_fingerprint,
    now_iso,
    runtime_snapshot,
)


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SCHEMA_VERSION = "common_pron_researcher_review_workbook.v1"
STANDARD_RULE_URL = (
    "https://korean.go.kr/kornorms/regltn/"
    "regltnView.do?regltn_code=0002"
)
OEGOLSI_QNA_URL = (
    "https://www.korean.go.kr/front/onlineQna/"
    "onlineQnaView.do?mn_id=216&pageIndex=1&qna_seq=324494"
)
SIPYUK_QNA_URL = (
    "https://www.korean.go.kr/front/onlineQna/"
    "onlineQnaView.do?mn_id=216&pageIndex=1&qna_seq=278421"
)
CHILSIP_DICT_URL = (
    "https://krdict.korean.go.kr/eng/dicMarinerSearch/"
    "search?mainSearchWord=%EC%B9%A0&nation=eng"
)
OEGOLSU_QNA_URL = (
    "https://www.korean.go.kr/front/onlineQna/"
    "onlineQnaView.do?mn_id=216&pageIndex=1&qna_seq=329993"
)

NAVY = "183B56"
TEAL = "0F766E"
BLUE = "2563EB"
PALE_BLUE = "E8F1FB"
PALE_TEAL = "DDF3EE"
PALE_YELLOW = "FFF4CC"
PALE_RED = "FDE8E7"
PALE_GREEN = "E4F3E8"
PALE_GRAY = "F2F4F7"
WHITE = "FFFFFF"
TEXT = "17212B"
GRID = "D7DEE6"
THIN_GRAY = Side(style="thin", color=GRID)


def clean(value: object) -> str:
    return str(value or "").strip()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        return [
            {key: clean(value) for key, value in row.items()}
            for row in csv.DictReader(stream)
        ]


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def read_dictionary(path: Path) -> list[dict[str, str]]:
    rows = []
    with path.open("r", encoding="utf-8-sig") as stream:
        for line in stream:
            stripped = line.strip()
            if not stripped:
                continue
            parts = stripped.split("\t", 1)
            if len(parts) != 2:
                raise RuntimeError(f"probe dictionary row invalid: {line!r}")
            rows.append(
                {
                    "input": parts[0].strip(),
                    "phones": parts[1].strip(),
                }
            )
    return rows


def add_table(sheet, ref: str, name: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def set_widths(sheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def style_title(sheet, cell_range: str, text: str) -> None:
    sheet.merge_cells(cell_range)
    cell = sheet[cell_range.split(":")[0]]
    cell.value = text
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(
        name="맑은 고딕", size=16, bold=True, color=WHITE
    )
    cell.alignment = Alignment(
        horizontal="left", vertical="center"
    )
    sheet.row_dimensions[cell.row].height = 32


def style_header(sheet, row: int, start: int, end: int) -> None:
    for col in range(start, end + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(
            name="맑은 고딕", size=10, bold=True, color=WHITE
        )
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = Border(bottom=THIN_GRAY)
    sheet.row_dimensions[row].height = 34


def style_body(sheet, min_row: int, max_row: int, max_col: int) -> None:
    for row in sheet.iter_rows(
        min_row=min_row, max_row=max_row, min_col=1, max_col=max_col
    ):
        for cell in row:
            cell.font = Font(name="맑은 고딕", size=9, color=TEXT)
            cell.alignment = Alignment(
                vertical="top", wrap_text=True
            )
            cell.border = Border(bottom=THIN_GRAY)


def style_phone_columns(
    sheet,
    *,
    columns: tuple[str, ...],
    min_row: int,
    max_row: int,
) -> None:
    """Use an IPA-capable font without changing any phone value."""
    for column in columns:
        for row in range(min_row, max_row + 1):
            cell = sheet[f"{column}{row}"]
            cell.font = Font(name="Noto Sans", size=9, color=TEXT)


def make_hyperlink(cell, target: str, display: str | None = None) -> None:
    if not target:
        return
    cell.value = display or target
    cell.hyperlink = (
        target
        if target.startswith(("http://", "https://"))
        else Path(target).resolve().as_uri()
    )
    cell.style = "Hyperlink"
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def recommendation_for_no_path(row: dict[str, str]) -> dict[str, str]:
    surface = row["surface"]
    rule_id = row["rule_id"]
    if surface == "읊고":
        return {
            "recommended_hangul": "읍꼬",
            "recommended_phone": "ɨ p̚ k͈ o",
            "alternative_hangul": "",
            "alternative_phone": "",
            "recommendation_action": "manual_phone_override",
            "confidence": "high",
            "reason": (
                "표준 발음법 제11항이 읊고[읍꼬]를 직접 제시한다. "
                "동결 G2P의 재철자 읍꼬 후보는 ɨː m k͈ o로 잘못되어, "
                "같은 acoustic inventory의 ɨ p̚ k͈ o를 별도 승인해야 한다."
            ),
            "source_url": STANDARD_RULE_URL,
            "source_handling": "phonological_exception",
            "implementation_note": (
                "후보는 보존하고 approved phone만 다르게 기록; "
                "기존 모델 출력 대체 없음"
            ),
        }
    if "rule18" in rule_id:
        reason = (
            "표준 발음법 제18항은 읊는[음는]을 직접 제시한다. "
            "동일 활용 환경의 모델 후보를 권고한다."
        )
    else:
        reason = (
            "표준 발음법 제14항의 읊어[을퍼] 연음과 동일한 활용 "
            "환경이며, 동결 모델 후보를 그대로 권고한다."
        )
    return {
        "recommended_hangul": row["respelled"],
        "recommended_phone": row["pron_phones_mfa"],
        "alternative_hangul": "",
        "alternative_phone": "",
        "recommendation_action": "accept_model_candidate",
        "confidence": "high",
        "reason": reason,
        "source_url": STANDARD_RULE_URL,
        "source_handling": "phonological_exception",
        "implementation_note": (
            "연구자 승인 후 same-model fallback으로 누락 표층키만 추가"
        ),
    }


def recommendation_for_jamo(token: str) -> dict[str, str]:
    records = {
        "외곬을": {
            "recommended_hangul": "외골쓸",
            "recommended_phone": "w eː ɡ o ɭ s͈ ɨ ɭ",
            "alternative_hangul": "",
            "alternative_phone": "",
            "recommendation_action": "manual_phone_override",
            "confidence": "high",
            "reason": (
                "표준 발음법 제14항은 ㄽ의 뒤 ㅅ을 모음 앞에서 "
                "된소리로 연음하며 곬이[골씨]를 직접 예시한다."
            ),
            "source_url": STANDARD_RULE_URL,
            "source_handling": "legitimate_surface_phonology",
            "implementation_note": (
                "원 표층키 외곬을을 유지하고 승인 phone만 적용"
            ),
        },
        "외곬의": {
            "recommended_hangul": "외골씌",
            "recommended_phone": "w eː ɡ o ɭ ɕ͈ i",
            "alternative_hangul": "외골쎄",
            "alternative_phone": "w eː ɡ o ɭ s͈ e",
            "recommendation_action": "researcher_audio_choice",
            "confidence": "medium",
            "reason": (
                "국립국어원은 [외골씌]를 원칙, [외골쎄]를 허용한다. "
                "form은 외곬의, original은 외곬에이므로 WAV 청취로 "
                "두 허용 후보 중 실제 발화에 맞는 값을 선택해야 한다."
            ),
            "source_url": OEGOLSI_QNA_URL,
            "source_handling": "legitimate_surface_pronunciation_choice",
            "implementation_note": (
                "원칙/허용 발음 중 연구자 선택을 명시적으로 기록"
            ),
        },
        "외곬수적인": {
            "recommended_hangul": "외골쑤저긴",
            "recommended_phone": "w eː ɡ o ɭ s͈ u dʑ ʌ ɟ i n",
            "alternative_hangul": "외골수저긴",
            "alternative_phone": "w eː ɡ o ɭ s u dʑ ʌ ɟ i n",
            "recommendation_action": "source_correction_and_audio_choice",
            "confidence": "medium",
            "reason": (
                "원 JSON도 외곬수적인이지만 문맥상 표준어 "
                "외골수적인의 원전사 표기 오류 가능성이 높다. "
                "NIKL 어휘부의 외골수 제1의 발음은 외골쑤/웨골쑤다. "
                "phone 승인과 함께 source correction registry에 "
                "원형 보존·교정형을 모두 기록해야 한다."
            ),
            "source_url": OEGOLSU_QNA_URL,
            "source_handling": "source_spelling_correction_required",
            "implementation_note": (
                "raw는 보존; MFA alias와 검색용 correction overlay를 함께 생성"
            ),
        },
        "천구백칤비육": {
            "recommended_hangul": "천구백칠씸뉵",
            "recommended_phone": (
                "tɕʰ ʌ ŋ ɡ u b ɛː k̚ tɕʰ i ɭ "
                "ɕ͈ i m ɲ u k̚"
            ),
            "alternative_hangul": "",
            "alternative_phone": "",
            "recommendation_action": "numeric_placeholder_correction",
            "confidence": "high",
            "reason": (
                "form의 1976년을 original placeholder가 비정상 문자열로 "
                "복원한 사례다. 공식 자료의 칠십[칠씹], 십육[심뉵]을 "
                "결합한 규범 후보이며 동결 1-best probe에서도 재현됐다."
            ),
            "source_url": SIPYUK_QNA_URL,
            "source_handling": "numeric_placeholder_correction_required",
            "implementation_note": (
                "raw placeholder는 보존; 1976→천구백칠십육 correction "
                "overlay와 승인 phone을 함께 기록"
            ),
        },
    }
    if token not in records:
        raise RuntimeError(f"unexpected Jamo ㄽ token: {token}")
    return records[token]


def find_lexicon_rows(path: Path) -> list[dict[str, str]]:
    wanted = {"외골수", "외곬", "읊다", "칠십", "십육"}
    matches = []
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        reader = csv.DictReader(stream)
        required = {
            "word",
            "pron_1",
            "pron_2",
            "pron_g2p",
            "pos_full",
            "sense_no",
            "origin",
            "urimal_id",
            "in_stdict",
        }
        missing = required - set(reader.fieldnames or ())
        if missing:
            raise RuntimeError(
                f"NIKL lexicon required columns missing: {sorted(missing)}"
            )
        for row in reader:
            if clean(row.get("word")) in wanted:
                matches.append(
                    {key: clean(row.get(key)) for key in required}
                )
    if not any(
        row["word"] == "외골수" and row["pron_1"] == "외골쑤"
        for row in matches
    ):
        raise RuntimeError("NIKL lexicon 외골수[외골쑤] evidence missing")
    return sorted(
        matches, key=lambda row: (row["word"], row["sense_no"])
    )


def build_review_rows(
    *,
    no_path_review: list[dict[str, str]],
    jamo_review: list[dict[str, str]],
    occurrences: list[dict[str, str]],
) -> list[dict[str, object]]:
    occurrences_by_token: defaultdict[str, list[dict[str, str]]] = (
        defaultdict(list)
    )
    for row in occurrences:
        occurrences_by_token[row["target_token"]].append(row)

    review_rows: list[dict[str, object]] = []
    pending_no_path = [
        row for row in no_path_review if row["decision"] == "pending"
    ]
    if len(pending_no_path) != 23:
        raise RuntimeError(
            f"no-path pending count must be 23: {len(pending_no_path)}"
        )
    for row in pending_no_path:
        token = row["surface"]
        evidence = occurrences_by_token[token]
        if not evidence:
            raise RuntimeError(f"no occurrence evidence for {token}")
        rec = recommendation_for_no_path(row)
        review_rows.append(
            {
                "category": "no_path",
                "token": token,
                "occurrences": len(evidence),
                "years": ",".join(
                    sorted({item["year"] for item in evidence})
                ),
                "model_input": row["respelled"],
                "model_candidate_phone": row["pron_phones_mfa"],
                **rec,
                "example_utt_id": evidence[0]["utt_id"],
                "review_wav": evidence[0]["review_wav"],
            }
        )

    if len(jamo_review) != 4:
        raise RuntimeError(f"Jamo review count must be 4: {len(jamo_review)}")
    for row in jamo_review:
        token = row["token"]
        if row["decision"] != "pending":
            raise RuntimeError(
                f"unexpected pre-existing Jamo decision: {token}"
            )
        evidence = occurrences_by_token[token]
        if len(evidence) != 1:
            raise RuntimeError(
                f"Jamo occurrence count must be 1: {token} {len(evidence)}"
            )
        review_rows.append(
            {
                "category": "jamo_ls",
                "token": token,
                "occurrences": 1,
                "years": evidence[0]["year"],
                "model_input": row["model_input"],
                "model_candidate_phone": row["pron_phones_mfa"],
                **recommendation_for_jamo(token),
                "example_utt_id": evidence[0]["utt_id"],
                "review_wav": evidence[0]["review_wav"],
            }
        )
    if len(review_rows) != 27:
        raise RuntimeError(f"review row count must be 27: {len(review_rows)}")
    review_rows.sort(
        key=lambda row: (
            0 if row["category"] == "no_path" else 1,
            row["token"],
        )
    )
    return review_rows


def populate_summary(wb: Workbook, review_sheet_name: str) -> None:
    sheet = wb.create_sheet("검토안내", 0)
    sheet.sheet_view.showGridLines = False
    style_title(
        sheet,
        "A1:H1",
        "공통발음사전 MFA r2 — 연구자 승인 검토표",
    )
    sheet["A3"] = "목적"
    sheet["B3"] = (
        "동결 acoustic v3.3.0 / Jamo G2P v3.2.0의 phone 체계를 "
        "바꾸지 않고, 자동 생성이 불가능하거나 잘못된 27개 표층형만 "
        "명시적으로 승인한다."
    )
    sheet.merge_cells("B3:H3")
    sheet["A4"] = "중요"
    sheet["B4"] = (
        "이 파일은 승인 인터페이스일 뿐이다. 노란 셀의 결정만 "
        "연구자가 입력하며, 파일을 저장해도 사전이나 shard는 자동으로 "
        "변경되지 않는다."
    )
    sheet.merge_cells("B4:H4")
    sheet["A6"] = "전체 검토"
    sheet["B6"] = f"=COUNTA('{review_sheet_name}'!$C$2:$C$28)"
    sheet["C6"] = "기록 가능"
    sheet["D6"] = (
        f'=COUNTIF(\'{review_sheet_name}\'!$V$2:$V$28,'
        '"ready_to_record")'
    )
    sheet["E6"] = "대기"
    sheet["F6"] = (
        f'=COUNTIF(\'{review_sheet_name}\'!$V$2:$V$28,"pending")'
    )
    sheet["G6"] = "보류/기각"
    sheet["H6"] = (
        f'=COUNTIF(\'{review_sheet_name}\'!$V$2:$V$28,"hold")+'
        f'COUNTIF(\'{review_sheet_name}\'!$V$2:$V$28,"reject")'
    )
    for cell in ("A6", "C6", "E6", "G6"):
        sheet[cell].fill = PatternFill("solid", fgColor=NAVY)
        sheet[cell].font = Font(
            name="맑은 고딕", bold=True, color=WHITE
        )
        sheet[cell].alignment = Alignment(horizontal="center")
    for cell in ("B6", "D6", "F6", "H6"):
        sheet[cell].fill = PatternFill("solid", fgColor=PALE_BLUE)
        sheet[cell].font = Font(
            name="맑은 고딕", size=14, bold=True, color=NAVY
        )
        sheet[cell].alignment = Alignment(horizontal="center")

    instructions = [
        (
            9,
            "1",
            "발음검토 시트에서 파란 WAV 링크를 눌러 필요한 항목을 듣는다.",
        ),
        (
            10,
            "2",
            "R열 decision을 선택한다. 권고 승인/대안 승인/직접 승인/"
            "보류/기각 중 하나다.",
        ),
        (
            11,
            "3",
            "직접 승인일 때만 S열에 정확한 MFA phone 열을 입력한다.",
        ),
        (
            12,
            "4",
            "U열 notes에 청취·판단 근거를 적는다. T열과 V열은 수식이므로 "
            "수정하지 않는다.",
        ),
        (
            13,
            "5",
            "외곬수적인·천구백칤비육은 발음 승인과 별도로 원표기 보존형 "
            "correction registry를 생성해야 한다.",
        ),
    ]
    for row, number, text in instructions:
        sheet[f"A{row}"] = number
        sheet[f"A{row}"].fill = PatternFill("solid", fgColor=TEAL)
        sheet[f"A{row}"].font = Font(bold=True, color=WHITE)
        sheet[f"A{row}"].alignment = Alignment(
            horizontal="center", vertical="center"
        )
        sheet[f"B{row}"] = text
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        sheet[f"B{row}"].alignment = Alignment(
            vertical="center", wrap_text=True
        )
        sheet.row_dimensions[row].height = 32
    for row in (3, 4):
        sheet[f"A{row}"].font = Font(bold=True, color=NAVY)
        sheet[f"B{row}"].alignment = Alignment(wrap_text=True)
        sheet.row_dimensions[row].height = 38
    set_widths(
        sheet,
        {
            "A": 11,
            "B": 18,
            "C": 11,
            "D": 13,
            "E": 11,
            "F": 13,
            "G": 11,
            "H": 13,
        },
    )
    sheet.freeze_panes = "A6"


def populate_review_sheet(
    wb: Workbook, rows: list[dict[str, object]]
) -> None:
    sheet = wb.create_sheet("발음검토")
    sheet.sheet_view.showGridLines = False
    headers = [
        "review_order",
        "category",
        "token",
        "occurrences",
        "years",
        "model_input",
        "model_candidate_phone",
        "recommended_hangul",
        "recommended_phone",
        "alternative_hangul",
        "alternative_phone",
        "recommendation_action",
        "confidence",
        "reason",
        "source_url",
        "example_utt_id",
        "review_wav",
        "researcher_decision",
        "researcher_custom_phone",
        "selected_approved_phone",
        "researcher_notes",
        "validation_status",
        "source_handling",
        "implementation_note",
    ]
    sheet.append(headers)
    for order, row in enumerate(rows, start=1):
        sheet.append(
            [
                order,
                row["category"],
                row["token"],
                int(row["occurrences"]),
                row["years"],
                row["model_input"],
                row["model_candidate_phone"],
                row["recommended_hangul"],
                row["recommended_phone"],
                row["alternative_hangul"],
                row["alternative_phone"],
                row["recommendation_action"],
                row["confidence"],
                row["reason"],
                row["source_url"],
                row["example_utt_id"],
                row["review_wav"],
                "pending",
                "",
                None,
                "",
                None,
                row["source_handling"],
                row["implementation_note"],
            ]
        )
        excel_row = order + 1
        sheet[f"T{excel_row}"] = (
            f'=IF(R{excel_row}="approve_recommended",I{excel_row},'
            f'IF(R{excel_row}="approve_alternative",K{excel_row},'
            f'IF(R{excel_row}="approve_custom",S{excel_row},"")))'
        )
        sheet[f"V{excel_row}"] = (
            f'=IF(OR(R{excel_row}="hold",R{excel_row}="reject"),'
            f'R{excel_row},IF(R{excel_row}="pending","pending",'
            f'IF(T{excel_row}="","phone_required","ready_to_record")))'
        )
        make_hyperlink(
            sheet[f"O{excel_row}"],
            clean(row["source_url"]),
            "공식/사전 근거 열기",
        )
        make_hyperlink(
            sheet[f"Q{excel_row}"],
            clean(row["review_wav"]),
            "WAV 듣기",
        )

    style_header(sheet, 1, 1, len(headers))
    style_body(sheet, 2, len(rows) + 1, len(headers))
    style_phone_columns(
        sheet,
        columns=("G", "I", "K", "S", "T"),
        min_row=2,
        max_row=len(rows) + 1,
    )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:X{len(rows)+1}"
    add_table(sheet, f"A1:X{len(rows)+1}", "PronReviewTable")
    decision_validation = DataValidation(
        type="list",
        formula1=(
            '"pending,approve_recommended,approve_alternative,'
            'approve_custom,hold,reject"'
        ),
        allow_blank=False,
    )
    decision_validation.error = "목록에서 결정값을 선택하세요."
    decision_validation.errorTitle = "허용되지 않은 결정"
    decision_validation.prompt = (
        "권고/대안/직접 승인 또는 hold/reject를 선택하세요."
    )
    decision_validation.promptTitle = "연구자 결정"
    sheet.add_data_validation(decision_validation)
    decision_validation.add(f"R2:R{len(rows)+1}")
    for row in range(2, len(rows) + 2):
        for column in ("R", "S", "U"):
            sheet[f"{column}{row}"].fill = PatternFill(
                "solid", fgColor=PALE_YELLOW
            )
        for column in ("T", "V"):
            sheet[f"{column}{row}"].fill = PatternFill(
                "solid", fgColor=PALE_GRAY
            )

    decision_range = f"R2:R{len(rows)+1}"
    sheet.conditional_formatting.add(
        decision_range,
        FormulaRule(
            formula=['R2="approve_recommended"'],
            fill=PatternFill("solid", fgColor=PALE_GREEN),
        ),
    )
    sheet.conditional_formatting.add(
        decision_range,
        FormulaRule(
            formula=['OR(R2="hold",R2="reject")'],
            fill=PatternFill("solid", fgColor=PALE_RED),
        ),
    )
    status_range = f"V2:V{len(rows)+1}"
    sheet.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=['V2="ready_to_record"'],
            fill=PatternFill("solid", fgColor=PALE_GREEN),
        ),
    )
    sheet.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=['V2="phone_required"'],
            fill=PatternFill("solid", fgColor=PALE_RED),
        ),
    )
    set_widths(
        sheet,
        {
            "A": 9,
            "B": 12,
            "C": 15,
            "D": 11,
            "E": 12,
            "F": 18,
            "G": 34,
            "H": 20,
            "I": 34,
            "J": 20,
            "K": 34,
            "L": 28,
            "M": 11,
            "N": 52,
            "O": 18,
            "P": 28,
            "Q": 15,
            "R": 25,
            "S": 34,
            "T": 34,
            "U": 36,
            "V": 19,
            "W": 36,
            "X": 48,
        },
    )
    for row in range(2, len(rows) + 2):
        sheet.row_dimensions[row].height = 58


def populate_occurrences(
    wb: Workbook, rows: list[dict[str, str]]
) -> None:
    sheet = wb.create_sheet("발화근거")
    sheet.sheet_view.showGridLines = False
    headers = [
        "review_order",
        "source_kind",
        "target_token",
        "year",
        "utt_id",
        "session_id",
        "speaker_id",
        "form",
        "original_form",
        "pron_reference_hangul",
        "pron_reference_source",
        "pron_reference_status",
        "raw_json_match_status",
        "review_wav",
        "original_wav",
        "wav_bytes",
        "wav_sha256",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
        excel_row = sheet.max_row
        make_hyperlink(
            sheet[f"N{excel_row}"], row["review_wav"], "검토 WAV"
        )
        make_hyperlink(
            sheet[f"O{excel_row}"], row["original_wav"], "원본 WAV"
        )
    style_header(sheet, 1, 1, len(headers))
    style_body(sheet, 2, len(rows) + 1, len(headers))
    sheet.freeze_panes = "A2"
    add_table(sheet, f"A1:Q{len(rows)+1}", "OccurrenceEvidenceTable")
    set_widths(
        sheet,
        {
            "A": 10,
            "B": 12,
            "C": 15,
            "D": 9,
            "E": 29,
            "F": 20,
            "G": 14,
            "H": 52,
            "I": 52,
            "J": 52,
            "K": 27,
            "L": 22,
            "M": 18,
            "N": 14,
            "O": 14,
            "P": 13,
            "Q": 68,
        },
    )
    for row in range(2, len(rows) + 2):
        sheet.row_dimensions[row].height = 54


def populate_raw_candidates(
    wb: Workbook,
    no_path_review: list[dict[str, str]],
    jamo_review: list[dict[str, str]],
) -> None:
    sheet = wb.create_sheet("모델후보원자료")
    sheet.sheet_view.showGridLines = False
    headers = [
        "source_kind",
        "token",
        "model_input",
        "model_candidate_phone",
        "existing_approved_phone",
        "decision",
        "rule_or_evidence",
        "notes",
    ]
    sheet.append(headers)
    for row in no_path_review:
        sheet.append(
            [
                "no_path",
                row["surface"],
                row["respelled"],
                row["pron_phones_mfa"],
                row.get("approved_pron_phones_mfa", ""),
                row["decision"],
                f"{row['rule_id']} | {row['evidence_source']}",
                row["notes"],
            ]
        )
    for row in jamo_review:
        sheet.append(
            [
                "jamo_ls",
                row["token"],
                row["model_input"],
                row["pron_phones_mfa"],
                row.get("approved_pron_phones_mfa", ""),
                row["decision"],
                row.get("evidence_source", ""),
                row["notes"],
            ]
        )
    style_header(sheet, 1, 1, len(headers))
    style_body(sheet, 2, sheet.max_row, len(headers))
    style_phone_columns(
        sheet,
        columns=("D", "E"),
        min_row=2,
        max_row=sheet.max_row,
    )
    sheet.freeze_panes = "A2"
    add_table(sheet, f"A1:H{sheet.max_row}", "RawCandidateTable")
    set_widths(
        sheet,
        {
            "A": 14,
            "B": 17,
            "C": 22,
            "D": 43,
            "E": 43,
            "F": 13,
            "G": 55,
            "H": 38,
        },
    )


def populate_sources(
    wb: Workbook,
    lexicon_rows: list[dict[str, str]],
    lexicon_path: Path,
) -> None:
    sheet = wb.create_sheet("근거자료")
    sheet.sheet_view.showGridLines = False
    headers = ["evidence_id", "scope", "finding", "source"]
    sheet.append(headers)
    evidence = [
        (
            "RULE11",
            "읊고",
            "표준 발음법 제11항: 읊고[읍꼬]",
            STANDARD_RULE_URL,
        ),
        (
            "RULE14",
            "읊어·외곬을",
            "제14항: 읊어[을퍼], 곬이[골씨]; ㄽ의 ㅅ은 모음 앞 [ㅆ]",
            STANDARD_RULE_URL,
        ),
        (
            "RULE18",
            "읊는",
            "제18항: 읊는[음는]",
            STANDARD_RULE_URL,
        ),
        (
            "OEGOLSI",
            "외곬의",
            "원칙 [외골씌], 허용 [외골쎄]",
            OEGOLSI_QNA_URL,
        ),
        (
            "CHILSIP",
            "1976",
            "한국어기초사전: 칠십[칠씹]",
            CHILSIP_DICT_URL,
        ),
        (
            "SIPYUK",
            "1976",
            "국립국어원: 십육[심뉵]",
            SIPYUK_QNA_URL,
        ),
        (
            "OEGOLSU",
            "외곬수적인",
            "외곬과 외골수는 구조·뜻이 다른 어휘; 원전사 교정 검토 필요",
            OEGOLSU_QNA_URL,
        ),
        (
            "NIKL_LEXICON",
            "외골수",
            "로컬 NIKL 어휘부에 외골수 제1 발음 외골쑤/웨골쑤",
            str(lexicon_path.resolve()),
        ),
    ]
    for row in evidence:
        sheet.append(list(row))
        make_hyperlink(
            sheet[f"D{sheet.max_row}"], row[3], "근거 열기"
        )
    style_header(sheet, 1, 1, len(headers))
    style_body(sheet, 2, sheet.max_row, len(headers))
    add_table(sheet, f"A1:D{sheet.max_row}", "EvidenceSourceTable")
    set_widths(sheet, {"A": 18, "B": 20, "C": 70, "D": 18})

    start = sheet.max_row + 3
    lex_headers = [
        "word",
        "sense_no",
        "pos_full",
        "pron_1",
        "pron_2",
        "pron_g2p",
        "origin",
        "urimal_id",
        "in_stdict",
    ]
    for col, value in enumerate(lex_headers, start=1):
        sheet.cell(start, col, value)
    for row in lexicon_rows:
        sheet.append([row.get(header, "") for header in lex_headers])
    style_header(sheet, start, 1, len(lex_headers))
    style_body(sheet, start + 1, sheet.max_row, len(lex_headers))
    add_table(
        sheet,
        f"A{start}:I{sheet.max_row}",
        "NiklLexiconEvidenceTable",
    )
    for row in range(start + 1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = 28


def populate_probes(
    wb: Workbook,
    original: list[dict[str, str]],
    addendum: list[dict[str, str]],
) -> None:
    sheet = wb.create_sheet("동결모델프로브")
    sheet.sheet_view.showGridLines = False
    headers = ["probe_set", "input", "phones", "purpose"]
    sheet.append(headers)
    for row in original:
        sheet.append(
            [
                "original_20260729",
                row["input"],
                row["phones"],
                "Jamo ㄽ 및 no-path 후보 분해 확인",
            ]
        )
    for row in addendum:
        sheet.append(
            [
                "addendum_1best_20260729",
                row["input"],
                row["phones"],
                "외골수·1976 source correction 후보 확인",
            ]
        )
    style_header(sheet, 1, 1, len(headers))
    style_body(sheet, 2, sheet.max_row, len(headers))
    style_phone_columns(
        sheet,
        columns=("C",),
        min_row=2,
        max_row=sheet.max_row,
    )
    add_table(sheet, f"A1:D{sheet.max_row}", "FrozenProbeTable")
    set_widths(sheet, {"A": 28, "B": 25, "C": 72, "D": 45})


def populate_contract(
    wb: Workbook,
    *,
    bundle_manifest: dict,
    model_bundle: dict,
    inputs: dict[str, Path],
) -> None:
    sheet = wb.create_sheet("재현계약")
    sheet.sheet_view.showGridLines = False
    headers = ["key", "value"]
    sheet.append(headers)
    values = [
        ("workbook_schema", SCHEMA_VERSION),
        ("recorded_at", now_iso()),
        (
            "acoustic_version",
            model_bundle["contract"]["acoustic_version"],
        ),
        ("g2p_version", model_bundle["contract"]["g2p_version"]),
        (
            "acoustic_model_sha256",
            model_bundle["outputs"]["acoustic_model"]["sha256"],
        ),
        (
            "g2p_model_sha256",
            model_bundle["outputs"]["g2p_model"]["sha256"],
        ),
        (
            "phone_inventory_count",
            model_bundle["contract"]["phone_count"],
        ),
        (
            "phone_inventory_sha256",
            model_bundle["contract"]["phone_sorted_sha256"],
        ),
        (
            "review_bundle_unique_wavs",
            bundle_manifest["counts"]["unique_wavs"],
        ),
        (
            "review_bundle_hash_equal",
            bundle_manifest["gates"][
                "all_review_copies_hash_equal_source"
            ],
        ),
        ("source_corpus_modified", False),
        ("approval_applied_to_r2", False),
    ]
    for key, value in values:
        sheet.append([key, value])
    sheet.append(["", ""])
    for label, path in sorted(inputs.items()):
        record = file_fingerprint(path, with_sha256=True)
        sheet.append([f"{label}_path", record["path"]])
        sheet.append([f"{label}_bytes", record["bytes"]])
        sheet.append([f"{label}_sha256", record["sha256"]])
    style_header(sheet, 1, 1, len(headers))
    style_body(sheet, 2, sheet.max_row, len(headers))
    set_widths(sheet, {"A": 38, "B": 105})
    sheet.freeze_panes = "A2"


def build_workbook(
    *,
    no_path_review_path: Path,
    jamo_review_path: Path,
    staged_occurrences_path: Path,
    staged_manifest_path: Path,
    model_bundle_path: Path,
    lexicon_path: Path,
    original_probe_path: Path,
    addendum_probe_path: Path,
    output_path: Path,
) -> dict:
    manifest_path = output_path.with_suffix(".manifest.json")
    if output_path.exists() or manifest_path.exists():
        raise FileExistsError(f"workbook output already exists: {output_path}")

    no_path_review = read_csv(no_path_review_path)
    jamo_review = read_csv(jamo_review_path)
    occurrences = read_csv(staged_occurrences_path)
    staged_manifest = read_json(staged_manifest_path)
    model_bundle = read_json(model_bundle_path)
    original_probes = read_dictionary(original_probe_path)
    addendum_probes = read_dictionary(addendum_probe_path)
    lexicon_rows = find_lexicon_rows(lexicon_path)

    if (
        staged_manifest.get("status") != "success"
        or staged_manifest.get("counts", {}).get(
            "target_occurrences"
        )
        != len(occurrences)
        or staged_manifest.get("gates", {}).get(
            "all_review_copies_hash_equal_source"
        )
        is not True
        or staged_manifest.get("gates", {}).get("source_corpus_modified")
        is not False
    ):
        raise RuntimeError("staged review bundle contract mismatch")
    review_rows = build_review_rows(
        no_path_review=no_path_review,
        jamo_review=jamo_review,
        occurrences=occurrences,
    )

    wb = Workbook()
    wb.remove(wb.active)
    populate_summary(wb, "발음검토")
    populate_review_sheet(wb, review_rows)
    populate_occurrences(wb, occurrences)
    populate_raw_candidates(wb, no_path_review, jamo_review)
    populate_sources(wb, lexicon_rows, lexicon_path)
    populate_probes(wb, original_probes, addendum_probes)
    input_paths = {
        "no_path_review": no_path_review_path,
        "jamo_review": jamo_review_path,
        "staged_occurrences": staged_occurrences_path,
        "staged_manifest": staged_manifest_path,
        "model_bundle": model_bundle_path,
        "lexicon": lexicon_path,
        "original_probe": original_probe_path,
        "addendum_probe": addendum_probe_path,
    }
    populate_contract(
        wb,
        bundle_manifest=staged_manifest,
        model_bundle=model_bundle,
        inputs=input_paths,
    )
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    for sheet in wb.worksheets:
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_view.zoomScale = 90

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    # Structural verification after a real round-trip through openpyxl.
    check = load_workbook(output_path, data_only=False, read_only=False)
    expected_sheets = [
        "검토안내",
        "발음검토",
        "발화근거",
        "모델후보원자료",
        "근거자료",
        "동결모델프로브",
        "재현계약",
    ]
    if check.sheetnames != expected_sheets:
        raise RuntimeError(f"workbook sheet contract mismatch: {check.sheetnames}")
    review_sheet = check["발음검토"]
    if (
        review_sheet.max_row != 28
        or review_sheet.max_column != 24
        or review_sheet["T2"].value
        != '=IF(R2="approve_recommended",I2,IF(R2="approve_alternative",K2,IF(R2="approve_custom",S2,"")))'
        or review_sheet["V2"].value
        != '=IF(OR(R2="hold",R2="reject"),R2,IF(R2="pending","pending",IF(T2="","phone_required","ready_to_record")))'
        or check["발화근거"].max_row != len(occurrences) + 1
    ):
        raise RuntimeError("workbook structural verification failed")
    data_validation_present = bool(
        review_sheet.data_validations.count
    )
    check.close()

    category_counts = Counter(
        clean(row["category"]) for row in review_rows
    )
    action_counts = Counter(
        clean(row["recommendation_action"]) for row in review_rows
    )
    manifest = {
        "schema_version": SCHEMA_VERSION,
        "status": "success",
        "kind": "common_pron_r2_researcher_review_workbook",
        "recorded_at": now_iso(),
        "counts": {
            "review_rows": len(review_rows),
            "no_path_rows": category_counts["no_path"],
            "jamo_ls_rows": category_counts["jamo_ls"],
            "occurrence_rows": len(occurrences),
            "unique_review_wavs": staged_manifest["counts"][
                "unique_wavs"
            ],
            "initial_researcher_decisions": 0,
            "recommendation_actions": dict(sorted(action_counts.items())),
        },
        "policy": {
            "all_decisions_initially_pending": True,
            "workbook_does_not_apply_decisions": True,
            "source_corpus_read_only": True,
            "candidate_and_approved_phone_separate": True,
            "same_acoustic_phone_inventory_required": True,
            "source_spelling_and_numeric_corrections_separate": True,
        },
        "inputs": {
            label: file_fingerprint(path, with_sha256=True)
            for label, path in input_paths.items()
        },
        "implementation": {
            "builder_script": file_fingerprint(
                Path(__file__).resolve(), with_sha256=True
            ),
            "openpyxl_version": openpyxl_version,
            "phone_font": "Noto Sans",
        },
        "output": file_fingerprint(output_path, with_sha256=True),
        "verification": {
            "round_trip_openpyxl": "passed",
            "expected_sheets": expected_sheets,
            "review_rows": 27,
            "occurrence_rows": 31,
            "formula_contract": "passed",
            "data_validation_present": data_validation_present,
        },
        "runtime": runtime_snapshot(PROJECT_ROOT),
    }
    atomic_write_json(manifest_path, manifest)
    return manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="공통발음 r2 연구자 검토용 XLSX 생성"
    )
    parser.add_argument("--no-path-review", type=Path, required=True)
    parser.add_argument("--jamo-review", type=Path, required=True)
    parser.add_argument("--staged-occurrences", type=Path, required=True)
    parser.add_argument("--staged-manifest", type=Path, required=True)
    parser.add_argument("--model-bundle", type=Path, required=True)
    parser.add_argument("--lexicon", type=Path, required=True)
    parser.add_argument("--original-probe", type=Path, required=True)
    parser.add_argument("--addendum-probe", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    return parser.parse_args()


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    args = parse_args()
    manifest = build_workbook(
        no_path_review_path=args.no_path_review.resolve(),
        jamo_review_path=args.jamo_review.resolve(),
        staged_occurrences_path=args.staged_occurrences.resolve(),
        staged_manifest_path=args.staged_manifest.resolve(),
        model_bundle_path=args.model_bundle.resolve(),
        lexicon_path=args.lexicon.resolve(),
        original_probe_path=args.original_probe.resolve(),
        addendum_probe_path=args.addendum_probe.resolve(),
        output_path=args.output.resolve(),
    )
    print(json.dumps(manifest["counts"], ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
