"""REVIEW.csv를 사용자가 입력하기 쉬운 REVIEW.xlsx로 만든다.

CSV는 기계 판독 정본으로 그대로 보존하고, XLSX는 편의용 보조 산출물이다.
사용자가 명시적으로 요청한 openpyxl 우회 경로를 사용한다.
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from pipeline_common import atomic_write_json, file_fingerprint, now_iso

BUNDLE_SCHEMA_VERSION = "mfa_r2_flat_review_bundle.v2"


def read_rows(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open(encoding="utf-8-sig", newline="") as stream:
        reader = csv.reader(stream)
        rows = list(reader)
    if len(rows) < 2:
        raise RuntimeError(f"검토 CSV가 비어 있음: {path}")
    return rows[0], rows[1:]


def create_workbook(
    input_csv: Path,
    output_xlsx: Path,
    manifest_path: Path | None = None,
    bundle_manifest_path: Path | None = None,
) -> dict[str, object]:
    input_csv = input_csv.resolve()
    output_xlsx = output_xlsx.resolve()
    manifest_path = (
        manifest_path.resolve()
        if manifest_path is not None
        else output_xlsx.with_name(
            "REVIEW_XLSX_TEMPLATE_MANIFEST.json"
        )
    )
    bundle_manifest_path = (
        bundle_manifest_path.resolve()
        if bundle_manifest_path is not None
        else None
    )
    if output_xlsx.exists():
        raise FileExistsError(f"기존 XLSX를 덮어쓰지 않음: {output_xlsx}")
    if manifest_path.exists():
        raise FileExistsError(
            f"기존 XLSX manifest를 덮어쓰지 않음: {manifest_path}"
        )
    if (
        bundle_manifest_path is not None
        and not bundle_manifest_path.is_file()
    ):
        raise FileNotFoundError(bundle_manifest_path)
    headers, rows = read_rows(input_csv)
    input_csv_record = file_fingerprint(
        input_csv, with_sha256=True
    )
    if bundle_manifest_path is not None:
        bundle = json.loads(
            bundle_manifest_path.read_text(encoding="utf-8-sig")
        )
        bundle_supporting = (
            bundle.get("supporting_files")
            if isinstance(bundle, dict)
            else None
        )
        bundle_review = (
            bundle_supporting.get("REVIEW.csv")
            if isinstance(bundle_supporting, dict)
            else None
        )
        if (
            not isinstance(bundle, dict)
            or bundle.get("schema_version") != BUNDLE_SCHEMA_VERSION
            or bundle.get("status") != "success"
            or bundle.get("flat_layout") is not True
            or not isinstance(bundle_review, dict)
            or bundle_review.get("sha256")
            != input_csv_record["sha256"]
            or bundle_review.get("bytes") != input_csv_record["bytes"]
        ):
            raise RuntimeError(
                "bundle manifest와 REVIEW.csv fingerprint 불일치"
            )
    required = {
        "review_order",
        "year",
        "utt_id",
        "linkage_status",
        "tier_structure_status",
        "boundary_status",
        "csv_searchability_status",
        "overall_infrastructure_decision",
        "notes",
    }
    missing = required - set(headers)
    if missing:
        raise RuntimeError(f"REVIEW.csv 필수 열 누락: {sorted(missing)}")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "검토입력"
    guide = workbook.create_sheet("안내")
    sheet.append(headers)
    for row in rows:
        typed_row: list[object] = []
        for header, value in zip(headers, row, strict=True):
            if header in {"review_order", "year"} and value:
                typed_row.append(int(value))
            else:
                typed_row.append(value)
        sheet.append(typed_row)

    dark = "1F4E78"
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    header_fill = PatternFill("solid", fgColor=dark)
    pass_fill = PatternFill("solid", fgColor="E2F0D9")
    issue_fill = PatternFill("solid", fgColor="FCE4D6")
    thin_gray = Side(style="thin", color="D9E2F3")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = Border(bottom=thin_gray)
    sheet.row_dimensions[1].height = 34
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False

    widths = {
        "review_order": 11,
        "year": 8,
        "utt_id": 34,
        "speaker_id": 18,
        "session_id": 20,
        "linkage_status": 16,
        "tier_structure_status": 18,
        "boundary_status": 16,
        "csv_searchability_status": 20,
        "overall_infrastructure_decision": 24,
        "notes": 44,
        "wav_file": 48,
        "textgrid_file": 48,
        "lab_file": 48,
        "csv_file": 48,
    }
    index = {name: offset + 1 for offset, name in enumerate(headers)}
    for name, width in widths.items():
        if name in index:
            sheet.column_dimensions[
                sheet.cell(1, index[name]).column_letter
            ].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top", wrap_text=(
                    cell.column == index.get("notes")
                )
            )
    for name in (
        "linkage_status",
        "tier_structure_status",
        "boundary_status",
        "csv_searchability_status",
        "overall_infrastructure_decision",
        "notes",
    ):
        column = index[name]
        for row_num in range(2, sheet.max_row + 1):
            sheet.cell(row_num, column).fill = input_fill

    component_validation = DataValidation(
        type="list",
        formula1='"미검토,통과,문제있음,해당없음"',
        allow_blank=False,
    )
    overall_validation = DataValidation(
        type="list",
        formula1='"미검토,인프라 통과,수정 후 재검토"',
        allow_blank=False,
    )
    sheet.add_data_validation(component_validation)
    sheet.add_data_validation(overall_validation)
    for name in (
        "linkage_status",
        "tier_structure_status",
        "boundary_status",
        "csv_searchability_status",
    ):
        letter = sheet.cell(1, index[name]).column_letter
        component_validation.add(f"{letter}2:{letter}{sheet.max_row}")
    overall_letter = sheet.cell(
        1, index["overall_infrastructure_decision"]
    ).column_letter
    overall_validation.add(
        f"{overall_letter}2:{overall_letter}{sheet.max_row}"
    )
    sheet.conditional_formatting.add(
        f"{overall_letter}2:{overall_letter}{sheet.max_row}",
        FormulaRule(
            formula=[f'${overall_letter}2="인프라 통과"'],
            fill=pass_fill,
        ),
    )
    sheet.conditional_formatting.add(
        f"{overall_letter}2:{overall_letter}{sheet.max_row}",
        FormulaRule(
            formula=[f'${overall_letter}2="수정 후 재검토"'],
            fill=issue_fill,
        ),
    )
    for name in ("wav_file", "textgrid_file", "lab_file", "csv_file"):
        column = index.get(name)
        if column is None:
            continue
        for row_num in range(2, sheet.max_row + 1):
            cell = sheet.cell(row_num, column)
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"

    table = Table(
        displayName="MfaR2InfrastructureReview",
        ref=f"A1:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    guide_rows = [
        ["MFA r2 인프라 수용 파일럿 검토 안내", ""],
        ["검토 목적", "향후 연구 검색·수동 판정을 위한 인프라가 연결되는지 확인"],
        ["이번 단계에서 확인", "WAV/TextGrid/CSV/LAB 연결, 4-tier 구조와 처음·끝 경계, CSV 검색 편의성"],
        ["이번 단계에서 하지 않음", "ㄴ 삽입 등 구체적 음운 실현 여부 판정"],
        ["phones tier 해석", "MFA 정렬·탐색 보조층이며 실제 실현 판정값이 아님"],
        [
            "생성 기준",
            (
                "REVIEW.csv는 XLSX 생성 기준으로 보존하며, 작성된 XLSX는 "
                "검증 후 기계 판독 결정표로 회수"
            ),
        ],
        [
            "입력 방법",
            (
                "노란 열의 드롭다운을 선택하고 필요한 경우 notes에 기록. "
                "다른 열과 파일 이름은 변경하지 않음"
            ),
        ],
    ]
    for row in guide_rows:
        guide.append(row)
    guide.merge_cells("A1:B1")
    guide["A1"].fill = header_fill
    guide["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    guide["A1"].alignment = Alignment(horizontal="center")
    guide.column_dimensions["A"].width = 23
    guide.column_dimensions["B"].width = 85
    for row in range(2, guide.max_row + 1):
        guide.cell(row, 1).font = Font(bold=True)
        guide.cell(row, 1).fill = PatternFill(
            "solid", fgColor="D9EAF7"
        )
        guide.cell(row, 2).alignment = Alignment(wrap_text=True)
        guide.row_dimensions[row].height = 32
    guide.sheet_view.showGridLines = False
    guide.freeze_panes = "A2"

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)
    check = load_workbook(output_xlsx, data_only=False)
    try:
        if (
            check.sheetnames != ["검토입력", "안내"]
            or check["검토입력"].max_row != len(rows) + 1
            or check["검토입력"].max_column != len(headers)
        ):
            raise RuntimeError("저장 후 XLSX 구조 검증 실패")
    finally:
        check.close()
    manifest: dict[str, object] = {
        "schema_version": "mfa_r2_review_xlsx_template.v1",
        "status": "success",
        "created_at": now_iso(),
        "input_review_csv": input_csv_record,
        "source_bundle_manifest": (
            file_fingerprint(
                bundle_manifest_path, with_sha256=True
            )
            if bundle_manifest_path is not None
            else None
        ),
        "output_review_xlsx": file_fingerprint(
            output_xlsx, with_sha256=True
        ),
        "rows": len(rows),
        "columns": len(headers),
        "sheets": ["검토입력", "안내"],
        "canonical_machine_readable_review": "REVIEW.csv",
        "workbook_role": "researcher_input_convenience_copy",
    }
    atomic_write_json(manifest_path, manifest)
    return manifest


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-csv", type=Path, required=True)
    parser.add_argument("--output-xlsx", type=Path, required=True)
    parser.add_argument("--manifest", type=Path)
    parser.add_argument("--bundle-manifest", type=Path)
    args = parser.parse_args()
    manifest = create_workbook(
        args.input_csv,
        args.output_xlsx,
        args.manifest,
        args.bundle_manifest,
    )
    print(json.dumps(manifest, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
