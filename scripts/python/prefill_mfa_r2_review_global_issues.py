"""Safely prefill repeated global issues in the MFA r2 review workbook.

The first researcher-reviewed data row is preserved byte-for-value at the
cell level.  Only rows whose ``review_order`` is 2 or greater are eligible,
and the script refuses to overwrite non-default researcher decisions.
"""

from __future__ import annotations

import argparse
import hashlib
import json
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

SHEET_NAME = "검토입력"
GUIDE_SHEET_NAME = "안내"
TABLE_NAME = "MfaR2InfrastructureReview"
DEFAULT_STATUS = "미검토"
GLOBAL_NOTE = "[전역 G-TIER-01, G-CSV-01]"

REQUIRED_HEADERS = (
    "review_order",
    "year",
    "utt_id",
    "speaker_id",
    "session_id",
    "linkage_status",
    "tier_structure_status",
    "boundary_status",
    "csv_searchability_status",
    "overall_infrastructure_decision",
    "notes",
    "wav_file",
    "textgrid_file",
    "lab_file",
    "csv_file",
)
MUTABLE_FIELDS = (
    "linkage_status",
    "tier_structure_status",
    "boundary_status",
    "csv_searchability_status",
    "overall_infrastructure_decision",
    "notes",
)
TARGET_FIELDS = (
    "tier_structure_status",
    "csv_searchability_status",
    "overall_infrastructure_decision",
    "notes",
)
LINK_FIELDS = ("wav_file", "textgrid_file", "lab_file", "csv_file")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def cell_snapshot(sheet: Any, row: int, index: dict[str, int]) -> dict[str, Any]:
    return {
        field: sheet.cell(row, index[field]).value
        for field in REQUIRED_HEADERS
    }


def immutable_snapshot(
    sheet: Any, index: dict[str, int]
) -> list[dict[str, Any]]:
    immutable = [
        field for field in REQUIRED_HEADERS if field not in MUTABLE_FIELDS
    ]
    return [
        {
            field: sheet.cell(row, index[field]).value
            for field in immutable
        }
        for row in range(2, sheet.max_row + 1)
    ]


def add_global_issue_guide(guide: Any) -> None:
    definitions = {
        "전역 이슈 G-TIER-01": (
            "legacy morphemes 시간분할은 words 경계와 다르고 형태소 시간경계로 "
            "오해될 수 있음. 후속본에서는 CSV 형태소 tagging을 발화 전체 길이의 "
            "단일 morph_analysis 구간으로 제시."
        ),
        "전역 이슈 G-CSV-01": (
            "tagged_roman은 표시·기초검색용으로 유지하되 형태소 시작/끝과 좌우 "
            "음운형태 환경을 안정적으로 찾을 구조화 morph_tokens 및 "
            "morph_boundaries 자료가 필요."
        ),
    }
    existing = {
        str(guide.cell(row, 1).value or "")
        for row in range(1, guide.max_row + 1)
    }
    style_source_row = guide.max_row
    for label, description in definitions.items():
        if label in existing:
            continue
        guide.append([label, description])
        row = guide.max_row
        for column in (1, 2):
            source = guide.cell(style_source_row, column)
            target = guide.cell(row, column)
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)
        guide.row_dimensions[row].height = guide.row_dimensions[
            style_source_row
        ].height


def prefill(input_path: Path, output_path: Path, manifest_path: Path) -> dict:
    input_path = input_path.resolve()
    output_path = output_path.resolve()
    manifest_path = manifest_path.resolve()
    if not input_path.is_file():
        raise FileNotFoundError(input_path)
    if output_path.exists():
        raise FileExistsError(output_path)
    if manifest_path.exists():
        raise FileExistsError(manifest_path)

    workbook = load_workbook(input_path, data_only=False, read_only=False)
    try:
        if workbook.sheetnames != [SHEET_NAME, GUIDE_SHEET_NAME]:
            raise RuntimeError(
                f"예상하지 못한 시트 구성: {workbook.sheetnames}"
            )
        sheet = workbook[SHEET_NAME]
        guide = workbook[GUIDE_SHEET_NAME]
        headers = [
            str(sheet.cell(1, column).value or "")
            for column in range(1, sheet.max_column + 1)
        ]
        if tuple(headers) != REQUIRED_HEADERS:
            raise RuntimeError(f"열 구성 불일치: {headers}")
        if sheet.max_row != 61 or sheet.max_column != 15:
            raise RuntimeError(
                f"검토표 크기 불일치: {sheet.max_row}x{sheet.max_column}"
            )
        index = {
            header: position + 1
            for position, header in enumerate(headers)
        }
        row_one_before = cell_snapshot(sheet, 2, index)
        immutable_before = immutable_snapshot(sheet, index)

        conflicts: list[dict[str, Any]] = []
        for row in range(3, sheet.max_row + 1):
            review_order = sheet.cell(
                row, index["review_order"]
            ).value
            expected_order = row - 1
            if review_order != expected_order:
                raise RuntimeError(
                    f"review_order 불일치: row={row}, value={review_order}"
                )
            expected = {
                "tier_structure_status": DEFAULT_STATUS,
                "csv_searchability_status": DEFAULT_STATUS,
                "overall_infrastructure_decision": DEFAULT_STATUS,
                "notes": (None, ""),
            }
            for field, allowed in expected.items():
                value = sheet.cell(row, index[field]).value
                allowed_values = (
                    allowed if isinstance(allowed, tuple) else (allowed,)
                )
                if value not in allowed_values:
                    conflicts.append(
                        {
                            "review_order": review_order,
                            "field": field,
                            "value": value,
                        }
                    )
        if conflicts:
            raise RuntimeError(
                "기존 연구자 입력을 덮어쓸 수 없음: "
                + json.dumps(conflicts[:10], ensure_ascii=False)
            )

        for row in range(3, sheet.max_row + 1):
            sheet.cell(
                row, index["tier_structure_status"]
            ).value = "문제있음"
            sheet.cell(
                row, index["csv_searchability_status"]
            ).value = "문제있음"
            sheet.cell(
                row, index["overall_infrastructure_decision"]
            ).value = "수정 후 재검토"
            sheet.cell(row, index["notes"]).value = GLOBAL_NOTE

        add_global_issue_guide(guide)
        if TABLE_NAME not in sheet.tables:
            table = Table(
                displayName=TABLE_NAME,
                ref=f"A1:O{sheet.max_row}",
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)

        if cell_snapshot(sheet, 2, index) != row_one_before:
            raise RuntimeError("1번 검토 행이 변경됨")
        if immutable_snapshot(sheet, index) != immutable_before:
            raise RuntimeError("원본 연결/식별 열이 변경됨")

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
    finally:
        workbook.close()

    check = load_workbook(output_path, data_only=False, read_only=False)
    try:
        sheet = check[SHEET_NAME]
        guide = check[GUIDE_SHEET_NAME]
        headers = [
            str(sheet.cell(1, column).value or "")
            for column in range(1, sheet.max_column + 1)
        ]
        index = {
            header: position + 1
            for position, header in enumerate(headers)
        }
        if cell_snapshot(sheet, 2, index) != row_one_before:
            raise RuntimeError("재로딩 후 1번 검토 행이 변경됨")
        if immutable_snapshot(sheet, index) != immutable_before:
            raise RuntimeError("재로딩 후 원본 연결/식별 열이 변경됨")
        if any(
            sheet.cell(row, index["tier_structure_status"]).value
            != "문제있음"
            or sheet.cell(
                row, index["csv_searchability_status"]
            ).value
            != "문제있음"
            or sheet.cell(
                row, index["overall_infrastructure_decision"]
            ).value
            != "수정 후 재검토"
            or sheet.cell(row, index["notes"]).value != GLOBAL_NOTE
            for row in range(3, sheet.max_row + 1)
        ):
            raise RuntimeError("전역 이슈 사전입력 재검증 실패")
        if any(
            sheet.cell(row, index[field]).value != DEFAULT_STATUS
            for row in range(3, sheet.max_row + 1)
            for field in ("linkage_status", "boundary_status")
        ):
            raise RuntimeError("개별 검토 대상 열이 변경됨")
        hyperlink_count = sum(
            1
            for row in range(2, sheet.max_row + 1)
            for field in LINK_FIELDS
            if sheet.cell(row, index[field]).hyperlink is not None
        )
        validation_count = len(
            sheet.data_validations.dataValidation
        )
        if hyperlink_count != 240:
            raise RuntimeError(
                f"하이퍼링크 수 불일치: {hyperlink_count}"
            )
        if validation_count != 2:
            raise RuntimeError(
                f"드롭다운 수 불일치: {validation_count}"
            )
        if TABLE_NAME not in sheet.tables:
            raise RuntimeError("검토표 Excel table 복원 실패")
        guide_labels = {
            str(guide.cell(row, 1).value or "")
            for row in range(1, guide.max_row + 1)
        }
        if not {
            "전역 이슈 G-TIER-01",
            "전역 이슈 G-CSV-01",
        }.issubset(guide_labels):
            raise RuntimeError("안내 시트의 전역 이슈 정의 누락")
    finally:
        check.close()

    result = {
        "schema_version": "mfa_r2_review_global_prefill.v1",
        "status": "success",
        "created_at": datetime.now().astimezone().isoformat(),
        "input": {
            "path": str(input_path),
            "bytes": input_path.stat().st_size,
            "sha256": sha256(input_path),
        },
        "output": {
            "path": str(output_path),
            "bytes": output_path.stat().st_size,
            "sha256": sha256(output_path),
        },
        "preserved_researcher_row": {
            "review_order": 1,
            "values": row_one_before,
        },
        "prefilled_review_orders": [2, 60],
        "prefilled_rows": 59,
        "global_issue_codes": ["G-TIER-01", "G-CSV-01"],
        "individual_review_remaining": [
            "linkage_status",
            "boundary_status",
        ],
        "hyperlinks_verified": 240,
        "dropdown_validations_verified": 2,
        "table_verified": TABLE_NAME,
        "source_or_mfa_results_modified": False,
    }
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return result


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--manifest", type=Path, required=True)
    args = parser.parse_args()
    result = prefill(args.input, args.output, args.manifest)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
