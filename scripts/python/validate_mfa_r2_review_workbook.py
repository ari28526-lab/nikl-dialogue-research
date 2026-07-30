"""작성된 MFA r2 인프라 검토 workbook을 원본 REVIEW.csv와 대조한다.

연구자가 수정할 수 있는 열만 받아들이고 파일·발화 연결 열은 원본 CSV와
행 단위로 전수 비교한다. 구체적인 음운 실현 판정은 이 스키마에 포함하지
않는다.
"""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pipeline_common import (
    atomic_text_writer,
    atomic_write_json,
    file_fingerprint,
    now_iso,
)


SCHEMA_VERSION = "mfa_r2_infrastructure_researcher_review.v1"
BUNDLE_SCHEMA_VERSION = "mfa_r2_flat_review_bundle.v2"
EDITABLE_FIELDS = (
    "linkage_status",
    "tier_structure_status",
    "boundary_status",
    "csv_searchability_status",
    "overall_infrastructure_decision",
    "notes",
)
COMPONENT_FIELDS = EDITABLE_FIELDS[:4]
COMPONENT_VALUES = {"미검토", "통과", "문제있음", "해당없음"}
OVERALL_VALUES = {"미검토", "인프라 통과", "수정 후 재검토"}


def read_csv_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(encoding="utf-8-sig", newline="") as stream:
        reader = csv.DictReader(stream)
        return list(reader.fieldnames or []), list(reader)


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_workbook_rows(
    path: Path,
) -> tuple[list[str], list[dict[str, str]]]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        if workbook.sheetnames != ["검토입력", "안내"]:
            raise RuntimeError(
                f"workbook 시트 계약 불일치: {workbook.sheetnames}"
            )
        sheet = workbook["검토입력"]
        headers = [
            cell_text(sheet.cell(1, column).value)
            for column in range(1, sheet.max_column + 1)
        ]
        if not headers or any(not header for header in headers):
            raise RuntimeError("workbook 헤더가 비어 있음")
        rows = [
            {
                header: cell_text(sheet.cell(row_number, column).value)
                for column, header in enumerate(headers, 1)
            }
            for row_number in range(2, sheet.max_row + 1)
        ]
        return headers, rows
    finally:
        workbook.close()


def write_decisions(
    path: Path,
    headers: list[str],
    rows: list[dict[str, str]],
) -> None:
    with atomic_text_writer(
        path, encoding="utf-8-sig", newline=""
    ) as (stream, _temp):
        writer = csv.DictWriter(
            stream, fieldnames=headers, lineterminator="\n"
        )
        writer.writeheader()
        writer.writerows(rows)


def validate_review_workbook(
    *,
    review_csv: Path,
    bundle_manifest_path: Path,
    workbook_path: Path,
    decision_csv: Path,
    report_path: Path,
    reviewer_id: str,
) -> dict[str, object]:
    review_csv = review_csv.resolve()
    bundle_manifest_path = bundle_manifest_path.resolve()
    workbook_path = workbook_path.resolve()
    decision_csv = decision_csv.resolve()
    report_path = report_path.resolve()
    if not reviewer_id.strip():
        raise ValueError("reviewer_id는 비어 있을 수 없음")
    for path in (review_csv, bundle_manifest_path, workbook_path):
        if not path.is_file():
            raise FileNotFoundError(path)
    for path in (decision_csv, report_path):
        if path.exists():
            raise FileExistsError(f"기존 검토 결과를 덮어쓰지 않음: {path}")

    bundle = json.loads(
        bundle_manifest_path.read_text(encoding="utf-8-sig")
    )
    if not isinstance(bundle, dict):
        raise RuntimeError("bundle manifest가 JSON object가 아님")
    supporting = bundle.get("supporting_files")
    machine_evidence = bundle.get("machine_gate_evidence")
    review_record = (
        supporting.get("REVIEW.csv")
        if isinstance(supporting, dict)
        else None
    )
    year_contracts = (
        machine_evidence.get("year_contracts")
        if isinstance(machine_evidence, dict)
        else None
    )
    review_actual = file_fingerprint(review_csv, with_sha256=True)
    if (
        bundle.get("schema_version")
        != BUNDLE_SCHEMA_VERSION
        or bundle.get("status") != "success"
        or bundle.get("flat_layout") is not True
        or bundle.get("review_scope")
        != "infrastructure_acceptance_only"
        or bundle.get("realization_judgment_performed") is not False
        or not isinstance(review_record, dict)
        or review_record.get("sha256") != review_actual["sha256"]
        or review_record.get("bytes") != review_actual["bytes"]
        or not isinstance(year_contracts, dict)
        or not year_contracts
    ):
        raise RuntimeError("bundle manifest/REVIEW.csv 기계 gate 불일치")

    csv_headers, source_rows = read_csv_rows(review_csv)
    workbook_headers, workbook_rows = read_workbook_rows(workbook_path)
    if workbook_headers != csv_headers:
        raise RuntimeError("workbook과 REVIEW.csv 헤더/순서 불일치")
    if len(workbook_rows) != len(source_rows) or not source_rows:
        raise RuntimeError(
            "workbook과 REVIEW.csv 행 수 불일치 또는 빈 검토표"
        )
    missing_editable = set(EDITABLE_FIELDS) - set(csv_headers)
    if missing_editable:
        raise RuntimeError(
            f"검토 입력 열 누락: {sorted(missing_editable)}"
        )

    immutable_fields = [
        field for field in csv_headers if field not in EDITABLE_FIELDS
    ]
    normalized: list[dict[str, str]] = []
    issues: list[dict[str, object]] = []
    for row_number, (source, reviewed) in enumerate(
        zip(source_rows, workbook_rows, strict=True),
        start=2,
    ):
        changed = [
            field
            for field in immutable_fields
            if cell_text(reviewed.get(field))
            != cell_text(source.get(field))
        ]
        if changed:
            raise RuntimeError(
                f"{row_number}행 불변 연결 열 변경: {changed}"
            )
        for field in COMPONENT_FIELDS:
            if reviewed[field] not in COMPONENT_VALUES:
                raise RuntimeError(
                    f"{row_number}행 {field} 허용값 아님: "
                    f"{reviewed[field]!r}"
                )
        overall = reviewed["overall_infrastructure_decision"]
        if overall not in OVERALL_VALUES:
            raise RuntimeError(
                f"{row_number}행 overall 허용값 아님: {overall!r}"
            )
        components = [reviewed[field] for field in COMPONENT_FIELDS]
        if (
            overall == "인프라 통과"
            and any(
                value not in {"통과", "해당없음"}
                for value in components
            )
        ):
            issues.append(
                {
                    "row": row_number,
                    "utt_id": reviewed.get("utt_id", ""),
                    "type": "inconsistent_pass",
                    "detail": components,
                }
            )
        if (
            overall == "수정 후 재검토"
            and not reviewed.get("notes", "").strip()
        ):
            issues.append(
                {
                    "row": row_number,
                    "utt_id": reviewed.get("utt_id", ""),
                    "type": "missing_issue_notes",
                }
            )
        normalized.append(
            {
                field: (
                    reviewed[field]
                    if field in EDITABLE_FIELDS
                    else source[field]
                )
                for field in csv_headers
            }
        )

    component_unreviewed = sum(
        row[field] == "미검토"
        for row in normalized
        for field in COMPONENT_FIELDS
    )
    overall_counts = Counter(
        row["overall_infrastructure_decision"]
        for row in normalized
    )
    reviewed_years = sorted(
        {row["year"] for row in normalized}
    )
    if reviewed_years != sorted(year_contracts):
        raise RuntimeError(
            "검토 행 연도와 bundle machine contract 연도 불일치"
        )
    all_approved = (
        not issues
        and component_unreviewed == 0
        and overall_counts["미검토"] == 0
        and overall_counts["수정 후 재검토"] == 0
        and overall_counts["인프라 통과"] == len(normalized)
    )
    status = (
        "approved"
        if all_approved
        else (
            "invalid"
            if issues
            else (
                "incomplete"
                if component_unreviewed or overall_counts["미검토"]
                else "changes_required"
            )
        )
    )

    write_decisions(decision_csv, csv_headers, normalized)
    report: dict[str, object] = {
        "schema_version": SCHEMA_VERSION,
        "status": status,
        "allow_bulk_mfa": all_approved,
        "reviewed_at": now_iso(),
        "reviewer_id": reviewer_id.strip(),
        "review_scope": "infrastructure_acceptance_only",
        "realization_judgment_performed": False,
        "source_review_csv": file_fingerprint(
            review_csv, with_sha256=True
        ),
        "source_bundle_manifest": file_fingerprint(
            bundle_manifest_path, with_sha256=True
        ),
        "year_contracts": year_contracts,
        "completed_workbook": file_fingerprint(
            workbook_path, with_sha256=True
        ),
        "normalized_decision_csv": file_fingerprint(
            decision_csv, with_sha256=True
        ),
        "counts": {
            "rows": len(normalized),
            "years": dict(
                sorted(Counter(row["year"] for row in normalized).items())
            ),
            "speakers": len(
                {
                    (row["year"], row["speaker_id"])
                    for row in normalized
                }
            ),
            "sessions": len(
                {
                    (row["year"], row["session_id"])
                    for row in normalized
                }
            ),
            "component_unreviewed": component_unreviewed,
            "overall_decisions": dict(sorted(overall_counts.items())),
            "consistency_issues": len(issues),
        },
        "issues": issues,
    }
    atomic_write_json(report_path, report)
    return report


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--review-csv", type=Path, required=True)
    parser.add_argument("--bundle-manifest", type=Path, required=True)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--decision-csv", type=Path, required=True)
    parser.add_argument("--report", type=Path, required=True)
    parser.add_argument("--reviewer-id", required=True)
    args = parser.parse_args()
    report = validate_review_workbook(
        review_csv=args.review_csv,
        bundle_manifest_path=args.bundle_manifest,
        workbook_path=args.workbook,
        decision_csv=args.decision_csv,
        report_path=args.report,
        reviewer_id=args.reviewer_id,
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if report["status"] == "approved" else 1


if __name__ == "__main__":
    raise SystemExit(main())
