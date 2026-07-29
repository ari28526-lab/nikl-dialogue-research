"""Render a compact XLSX range preview with openpyxl and Pillow.

This is a visual QA fallback for environments where the artifact-tool loader
or headless Office renderer is unavailable.  It does not modify the workbook.
"""
from __future__ import annotations

import argparse
import math
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries
from PIL import Image, ImageDraw, ImageFont


REGULAR_FONT = Path(r"C:\Windows\Fonts\malgun.ttf")
BOLD_FONT = Path(r"C:\Windows\Fonts\malgunbd.ttf")
IPA_REGULAR_FONT = Path(r"C:\Windows\Fonts\NotoSans-Regular.ttf")
IPA_BOLD_FONT = Path(r"C:\Windows\Fonts\NotoSans-Bold.ttf")


def rgb(value: object, default: str) -> str:
    color_type = getattr(value, "type", None)
    raw = getattr(value, "rgb", None)
    if color_type == "rgb" and raw:
        return f"#{str(raw)[-6:]}"
    return default


def column_pixels(width: float | None) -> int:
    return max(35, int((width if width is not None else 8.43) * 7 + 8))


def row_pixels(height: float | None) -> int:
    points = height if height is not None else 15
    return max(20, int(points * 96 / 72))


def wrap_lines(
    draw: ImageDraw.ImageDraw,
    text: str,
    font: ImageFont.FreeTypeFont,
    width: int,
    max_lines: int,
) -> list[str]:
    if not text:
        return []
    output: list[str] = []
    for paragraph in text.splitlines() or [""]:
        current = ""
        for char in paragraph:
            candidate = current + char
            if draw.textlength(candidate, font=font) <= width or not current:
                current = candidate
                continue
            output.append(current)
            current = char
            if len(output) >= max_lines:
                break
        if len(output) >= max_lines:
            break
        output.append(current)
    if len(output) > max_lines:
        output = output[:max_lines]
    if output and len("\n".join(output)) < len(text):
        last = output[-1]
        while last and draw.textlength(
            f"{last}…", font=font
        ) > width:
            last = last[:-1]
        output[-1] = f"{last}…"
    return output


def render_range(
    *,
    input_path: Path,
    sheet_name: str,
    cell_range: str,
    output_path: Path,
    scale: float,
) -> None:
    workbook = load_workbook(input_path, data_only=False, read_only=False)
    sheet = workbook[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    col_widths = [
        column_pixels(
            sheet.column_dimensions[get_column_letter(col)].width
        )
        for col in range(min_col, max_col + 1)
    ]
    row_heights = [
        row_pixels(sheet.row_dimensions[row].height)
        for row in range(min_row, max_row + 1)
    ]
    x_positions = [0]
    for width in col_widths:
        x_positions.append(x_positions[-1] + width)
    y_positions = [0]
    for height in row_heights:
        y_positions.append(y_positions[-1] + height)

    image = Image.new(
        "RGB",
        (x_positions[-1] + 2, y_positions[-1] + 2),
        "white",
    )
    draw = ImageDraw.Draw(image)
    merged_by_cell = {}
    merged_non_anchor = set()
    for merged in sheet.merged_cells.ranges:
        if (
            merged.max_col < min_col
            or merged.min_col > max_col
            or merged.max_row < min_row
            or merged.min_row > max_row
        ):
            continue
        anchor = (merged.min_row, merged.min_col)
        merged_by_cell[anchor] = merged
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                if (row, col) != anchor:
                    merged_non_anchor.add((row, col))

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            if (row, col) in merged_non_anchor:
                continue
            cell = sheet.cell(row, col)
            x1 = x_positions[col - min_col]
            y1 = y_positions[row - min_row]
            merged = merged_by_cell.get((row, col))
            if merged is not None:
                span_max_col = min(merged.max_col, max_col)
                span_max_row = min(merged.max_row, max_row)
                x2 = x_positions[span_max_col - min_col + 1]
                y2 = y_positions[span_max_row - min_row + 1]
            else:
                x2 = x_positions[col - min_col + 1]
                y2 = y_positions[row - min_row + 1]
            fill_color = (
                rgb(cell.fill.fgColor, "#FFFFFF")
                if cell.fill.fill_type
                else "#FFFFFF"
            )
            draw.rectangle((x1, y1, x2, y2), fill=fill_color)
            draw.rectangle((x1, y1, x2, y2), outline="#D7DEE6", width=1)
            value = "" if cell.value is None else str(cell.value)
            font_size = max(8, int(float(cell.font.sz or 9)))
            is_ipa_font = (cell.font.name or "").lower() == "noto sans"
            if is_ipa_font:
                font_path = (
                    IPA_BOLD_FONT if cell.font.bold else IPA_REGULAR_FONT
                )
            else:
                font_path = BOLD_FONT if cell.font.bold else REGULAR_FONT
            font = ImageFont.truetype(str(font_path), font_size)
            text_color = rgb(cell.font.color, "#17212B")
            padding = 4
            usable_width = max(10, x2 - x1 - 2 * padding)
            line_height = max(11, int(font_size * 1.45))
            max_lines = max(1, math.floor((y2 - y1 - 2 * padding) / line_height))
            lines = wrap_lines(
                draw, value, font, usable_width, max_lines
            )
            text_block_height = len(lines) * line_height
            vertical = cell.alignment.vertical or "top"
            if vertical == "center":
                ty = y1 + max(padding, (y2 - y1 - text_block_height) // 2)
            elif vertical == "bottom":
                ty = y2 - padding - text_block_height
            else:
                ty = y1 + padding
            horizontal = cell.alignment.horizontal or "left"
            for line in lines:
                length = draw.textlength(line, font=font)
                if horizontal == "center":
                    tx = x1 + max(padding, (x2 - x1 - length) / 2)
                elif horizontal == "right":
                    tx = x2 - padding - length
                else:
                    tx = x1 + padding
                draw.text((tx, ty), line, font=font, fill=text_color)
                ty += line_height
    workbook.close()

    if scale != 1:
        image = image.resize(
            (
                max(1, int(image.width * scale)),
                max(1, int(image.height * scale)),
            ),
            Image.Resampling.LANCZOS,
        )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    image.save(output_path, format="PNG")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="XLSX range PNG preview")
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--sheet", required=True)
    parser.add_argument("--range", dest="cell_range", required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--scale", type=float, default=1.0)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    render_range(
        input_path=args.input.resolve(),
        sheet_name=args.sheet,
        cell_range=args.cell_range,
        output_path=args.output.resolve(),
        scale=args.scale,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
