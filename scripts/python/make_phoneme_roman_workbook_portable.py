"""Rewrite the phoneme Roman pilot workbook with folder-relative review links.

The source workbook is kept untouched.  The output must live beside the 12 WAV,
12 original four-tier TextGrids, and 12 optional five-tier TextGrids when used.
The user previously approved the project's openpyxl fallback for review files.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


REVIEW_SHEET = "발화_검토"
LINK_HEADERS = ("WAV", "기존_4tier", "새_5tier")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def expected_targets(year: str, utt_id: str) -> dict[str, str]:
    prefix = f"{year}__{utt_id}"
    return {
        "WAV": f"{prefix}.wav",
        "기존_4tier": f"{prefix}.TextGrid",
        "새_5tier": f"{prefix}__phoneme_r_auto.TextGrid",
    }


def rewrite_links(source: Path, output: Path, review_root: Path) -> dict[str, object]:
    if source.resolve() == output.resolve():
        raise ValueError("원본 workbook은 덮어쓸 수 없습니다")
    workbook = load_workbook(source, data_only=False, read_only=False)
    if REVIEW_SHEET not in workbook.sheetnames:
        raise RuntimeError(f"검토 sheet 없음: {REVIEW_SHEET}")
    sheet = workbook[REVIEW_SHEET]
    headers = {str(cell.value): cell.column for cell in sheet[1] if cell.value}
    required = {"연도", "utt_id", *LINK_HEADERS}
    missing_headers = sorted(required - set(headers))
    if missing_headers:
        raise RuntimeError(f"필수 열 없음: {missing_headers}")

    missing_files: list[str] = []
    rewritten = 0
    for row_index in range(2, sheet.max_row + 1):
        year = str(sheet.cell(row_index, headers["연도"]).value)
        utt_id = str(sheet.cell(row_index, headers["utt_id"]).value)
        for header, filename in expected_targets(year, utt_id).items():
            target = review_root / filename
            if not target.is_file():
                missing_files.append(filename)
                continue
            cell = sheet.cell(row_index, headers[header])
            cell.hyperlink = filename
            cell.style = "Hyperlink"
            rewritten += 1
    if missing_files:
        raise FileNotFoundError(f"Dropbox 검토 파일 누락: {missing_files}")

    output.parent.mkdir(parents=True, exist_ok=True)
    temp = output.with_name(f".{output.name}.{uuid.uuid4().hex}.partial")
    workbook.save(temp)
    os.replace(temp, output)

    # Round-trip and portability verification.
    check = load_workbook(output, data_only=False, read_only=False)
    check_sheet = check[REVIEW_SHEET]
    check_headers = {str(cell.value): cell.column for cell in check_sheet[1] if cell.value}
    targets: list[str] = []
    for row_index in range(2, check_sheet.max_row + 1):
        for header in LINK_HEADERS:
            cell = check_sheet.cell(row_index, check_headers[header])
            if cell.hyperlink is None:
                raise RuntimeError(f"링크 누락: {cell.coordinate}")
            target = str(cell.hyperlink.target)
            if Path(target).is_absolute() or ":" in target or "/" in target or "\\" in target:
                raise RuntimeError(f"절대/하위경로 링크 남음: {cell.coordinate}={target}")
            if not (review_root / target).is_file():
                raise RuntimeError(f"링크 대상 없음: {target}")
            targets.append(target)
    expected_link_count = (check_sheet.max_row - 1) * len(LINK_HEADERS)
    if len(targets) != expected_link_count:
        raise RuntimeError(
            f"휴대형 링크 수 불일치: {len(targets)}/{expected_link_count}"
        )

    return {
        "schema_version": "phoneme_roman_portable_workbook.v1",
        "status": "success",
        "recorded_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "source_workbook": str(source.resolve()),
        "output_workbook": str(output.resolve()),
        "source_sha256": sha256_file(source),
        "output_sha256": sha256_file(output),
        "review_rows": check_sheet.max_row - 1,
        "relative_links": len(targets),
        "unique_link_targets": len(set(targets)),
        "missing_link_targets": 0,
        "source_overwritten": False,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--review-root", type=Path, required=True)
    parser.add_argument("--manifest", type=Path, required=True)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    manifest = rewrite_links(
        args.source.resolve(), args.output.resolve(), args.review_root.resolve()
    )
    args.manifest.parent.mkdir(parents=True, exist_ok=True)
    args.manifest.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(json.dumps(manifest, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
