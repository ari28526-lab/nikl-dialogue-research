"""Audit whether readiness-v4 policy rows have a reusable researcher decision."""

from __future__ import annotations

import argparse
import csv
import gzip
import json
import re
import sys
from pathlib import Path
from typing import Iterator

sys.path.insert(0, str(Path(__file__).resolve().parent))

from pipeline_common import atomic_write_json, file_fingerprint, now_iso, runtime_snapshot


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SCHEMA_VERSION = "common_pron_r3_policy_decision_reuse_audit.v1"
NAME_PATTERN = re.compile(r"review|decision|approval|ledger", re.IGNORECASE)
EXTENSIONS = {".csv", ".json", ".xlsx"}


def clean(value: object) -> str:
    return str(value or "").strip()


def scalar_strings(value: object) -> Iterator[str]:
    if isinstance(value, dict):
        for key, item in value.items():
            yield clean(key)
            yield from scalar_strings(item)
    elif isinstance(value, list):
        for item in value:
            yield from scalar_strings(item)
    elif value is not None:
        yield clean(value)


def exact_hits_csv(path: Path, targets: set[str]) -> set[str]:
    for encoding in ("utf-8-sig", "cp949"):
        try:
            hits: set[str] = set()
            with path.open("r", encoding=encoding, newline="") as stream:
                for row in csv.reader(stream):
                    hits.update(value for value in map(clean, row) if value in targets)
            return hits
        except UnicodeDecodeError:
            continue
    return set()


def exact_hits_json(path: Path, targets: set[str]) -> set[str]:
    try:
        data = json.loads(path.read_text(encoding="utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError):
        return set()
    return {value for value in scalar_strings(data) if value in targets}


def exact_hits_xlsx(path: Path, targets: set[str]) -> set[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment contract
        raise RuntimeError("openpyxl is required to audit XLSX decision ledgers") from exc
    hits: set[str] = set()
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                hits.update(value for value in map(clean, row) if value in targets)
    finally:
        workbook.close()
    return hits


def audit(readiness_path: Path, ledger_root: Path, output_path: Path) -> dict[str, object]:
    targets: dict[str, dict[str, str]] = {}
    with gzip.open(readiness_path, "rt", encoding="utf-8-sig", newline="") as stream:
        for row in csv.DictReader(stream):
            if clean(row.get("planning_requires_policy_decision")) == "true":
                targets[row["token"]] = row
    if len(targets) != 35 or sum(int(row["total_occurrences"]) for row in targets.values()) != 163:
        raise RuntimeError("readiness-v4 policy target accounting differs")

    bound = {
        token: clean(row.get("manual_decision_id"))
        for token, row in targets.items()
        if clean(row.get("manual_decision_id"))
    }
    ledger_files = sorted(
        path
        for path in ledger_root.rglob("*")
        if path.is_file()
        and path.suffix.lower() in EXTENSIONS
        and NAME_PATTERN.search(path.name)
        and path.stat().st_size <= 100_000_000
    )
    hit_files: list[dict[str, object]] = []
    target_set = set(targets)
    for path in ledger_files:
        if path.suffix.lower() == ".csv":
            hits = exact_hits_csv(path, target_set)
        elif path.suffix.lower() == ".json":
            hits = exact_hits_json(path, target_set)
        else:
            hits = exact_hits_xlsx(path, target_set)
        if hits:
            hit_files.append(
                {
                    "file": file_fingerprint(path, with_sha256=True),
                    "exact_target_tokens": sorted(hits),
                }
            )

    exact_tokens = sorted({token for record in hit_files for token in record["exact_target_tokens"]})
    reusable = sorted(set(bound))
    status = "no_reusable_decisions_found" if not reusable else "reusable_decisions_require_binding_audit"
    report = {
        "schema_version": SCHEMA_VERSION,
        "status": status,
        "recorded_at": now_iso(),
        "scope": {
            "reuse_requires_explicit_manual_decision_id": True,
            "generated_r2_pronunciation_is_not_researcher_approval": True,
            "no_decision_inferred_from_output_similarity": True,
        },
        "inputs": {
            "readiness_v4": file_fingerprint(readiness_path, with_sha256=True),
            "ledger_root": str(ledger_root.resolve()),
        },
        "counts": {
            "policy_types": len(targets),
            "policy_occurrences": sum(int(row["total_occurrences"]) for row in targets.values()),
            "manual_decision_id_bound_types": len(bound),
            "ledger_files_scanned": len(ledger_files),
            "files_with_exact_token_hits": len(hit_files),
            "exact_token_hit_types": len(exact_tokens),
            "reusable_decision_types": len(reusable),
        },
        "exact_token_hits": hit_files,
        "reusable_decision_tokens": reusable,
        "unresolved_policy_tokens": sorted(set(targets) - set(reusable)),
        "runtime": runtime_snapshot(PROJECT_ROOT),
    }
    atomic_write_json(output_path, report)
    return report


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--readiness-v4", type=Path, required=True)
    parser.add_argument("--ledger-root", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    report = audit(args.readiness_v4.resolve(), args.ledger_root.resolve(), args.output.resolve())
    print(json.dumps(report["counts"], ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
