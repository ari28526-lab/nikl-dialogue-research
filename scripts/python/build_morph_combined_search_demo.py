"""60발화 구조화 표의 조합 검색 결과를 작은 검토 workbook으로 만든다."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

MAX_EXAMPLES_PER_QUERY = 2
REVIEW_HEADERS = ("조건_일치", "파일_연결", "표시_이해", "판정", "메모")


def read_csv(path: Path) -> list[dict[str, str]]:
    with open(path, encoding="utf-8-sig", newline="") as stream:
        return list(csv.DictReader(stream))


def write_csv(
    path: Path, rows: Iterable[Mapping[str, object]], fieldnames: list[str]
) -> None:
    with open(path, "w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(
            stream,
            fieldnames=fieldnames,
            extrasaction="ignore",
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(rows)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def truthy(value: object) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes"}


def json_count(value: object) -> int:
    try:
        parsed = json.loads(str(value or "[]"))
    except json.JSONDecodeError:
        return 0
    return len(parsed) if isinstance(parsed, list) else 0


def common_record(
    row: Mapping[str, str], *, matched_table: str, evidence: str
) -> dict[str, str]:
    return {
        "utt_id": row["utt_id"],
        "year": row["year"],
        "matched_table": matched_table,
        "evidence": evidence,
        "boundary_scope": row.get("boundary_scope", ""),
        "left_morph": row.get("left_morph_surface", ""),
        "left_pos": row.get("left_pos", ""),
        "left_coda": row.get("left_coda_jamo", ""),
        "right_morph": row.get("right_morph_surface", ""),
        "right_pos": row.get("right_pos", ""),
        "right_onset_zero": row.get("right_onset_zero", ""),
        "right_nucleus": row.get("right_nucleus_jamo", ""),
        "target_morph": row.get("morph_surface", ""),
        "target_pos": row.get("pos", ""),
        "morph_position": (
            f"{row.get('morph_idx_in_eojeol', '')}/"
            f"{row.get('morph_count_in_eojeol', '')}"
            if row.get("morph_idx_in_eojeol")
            else ""
        ),
        "unit_surface": row.get("unit_surface", ""),
        "unit_roman": row.get("unit_roman", ""),
        "components": "",
        "align_warn": row.get("align_warn", ""),
    }


def boundary_query(
    rows: list[dict[str, str]],
    predicate: Callable[[Mapping[str, str]], bool],
    evidence: Callable[[Mapping[str, str]], str],
) -> list[dict[str, str]]:
    return [
        common_record(row, matched_table="morph_boundaries", evidence=evidence(row))
        for row in rows
        if predicate(row)
    ]


def unit_query(
    rows: list[dict[str, str]],
    predicate: Callable[[Mapping[str, str]], bool],
    evidence: Callable[[Mapping[str, str]], str],
    component_field: str = "",
) -> list[dict[str, str]]:
    result: list[dict[str, str]] = []
    for row in rows:
        if not predicate(row):
            continue
        record = common_record(row, matched_table="morph_units", evidence=evidence(row))
        if component_field:
            record["components"] = row.get(component_field, "")
        result.append(record)
    return result


def build_queries(
    *,
    masters: list[dict[str, str]],
    morphs: list[dict[str, str]],
    units: list[dict[str, str]],
    boundaries: list[dict[str, str]],
) -> tuple[list[dict[str, str]], dict[str, list[dict[str, str]]]]:
    high_vowels = {"ㅣ", "ㅑ", "ㅕ", "ㅛ", "ㅠ"}
    catalog = [
        {
            "query_id": "Q1_N_INSERTION_LIKE",
            "query_name": "ㄴ 삽입 유사 철자 환경",
            "research_question": "좌측 종성 뒤 우측 ㅇ+i/y계 모음 환경인가",
            "filter": "left_coda!='' AND right_onset_zero AND right_nucleus IN (ㅣ,ㅑ,ㅕ,ㅛ,ㅠ)",
            "interpretation": "광의 후보. 실제 ㄴ 삽입 판정이 아님",
        },
        {
            "query_id": "Q2_MORPH_POS_POSITION",
            "query_name": "형태소+품사+어절내 위치",
            "research_question": "이/JKS·JKC가 어절 마지막 형태소인가",
            "filter": "morph_surface='이' AND pos IN (JKS,JKC) AND morph_idx=morph_count",
            "interpretation": "동형 표면형을 POS와 위치로 좁히는 예",
        },
        {
            "query_id": "Q3_COMPLEX_CODA",
            "query_name": "겹받침 구성 자모",
            "research_question": "종성 slot이 둘 이상의 철자 구성 자모를 가지는가",
            "filter": "unit_type='hangul' AND len(coda_components)>1",
            "interpretation": "slot을 실제 phone 둘로 판정하지 않음",
        },
        {
            "query_id": "Q4_COMPOUND_NUCLEUS",
            "query_name": "복합모음 구성 자모",
            "research_question": "중성 slot이 둘 이상의 철자 구성 자모를 가지는가",
            "filter": "unit_type='hangul' AND len(nucleus_components)>1",
            "interpretation": "철자 구성 검색",
        },
        {
            "query_id": "Q5_STANDALONE_JAMO",
            "query_name": "독립 자모 형태소",
            "research_question": "형태소 첫 unit이 독립 자모인가",
            "filter": "unit_type='jamo' AND unit_idx_in_morph=1",
            "interpretation": "ㄴ·ㄹ 등 독립 표면형을 누락하지 않는 예",
        },
        {
            "query_id": "Q6_LITERAL_ALIGN_WARN",
            "query_name": "literal+형태 정렬 경고",
            "research_question": "기호·영문 literal이 있고 form-tagged 대응 경고도 있는가",
            "filter": "has_literal AND align_warn!=''",
            "interpretation": "자동 해석 대신 사람이 확인할 예외 후보",
        },
        {
            "query_id": "Q7_INTER_EOJEOL_ENV",
            "query_name": "어절 사이 좌우 환경",
            "research_question": "어절 경계에서 좌 종성+우 무음초성인가",
            "filter": "boundary_scope='inter_eojeol' AND left_coda!='' AND right_onset_zero",
            "interpretation": "어절 내부 경계와 구분한 후보",
        },
    ]

    grouped_units: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in units:
        grouped_units[row["utt_id"]].append(row)
    queries: dict[str, list[dict[str, str]]] = {}
    queries["Q1_N_INSERTION_LIKE"] = boundary_query(
        boundaries,
        lambda row: bool(row["left_coda_jamo"])
        and truthy(row["right_onset_zero"])
        and row["right_nucleus_jamo"] in high_vowels,
        lambda row: (
            f"{row['left_morph_surface']}/{row['left_pos']}[종성 {row['left_coda_jamo']}] "
            f"{row['boundary_scope']} {row['right_morph_surface']}/{row['right_pos']}"
            f"[초성 ㅇ, 중성 {row['right_nucleus_jamo']}]"
        ),
    )
    queries["Q2_MORPH_POS_POSITION"] = [
        common_record(
            row,
            matched_table="morph_tokens",
            evidence=(
                f"{row['morph_surface']}/{row['pos']} · 어절내 형태소 "
                f"{row['morph_idx_in_eojeol']}/{row['morph_count_in_eojeol']}"
            ),
        )
        for row in morphs
        if row["morph_surface"] == "이"
        and row["pos"] in {"JKS", "JKC"}
        and row["morph_idx_in_eojeol"] == row["morph_count_in_eojeol"]
    ]
    queries["Q3_COMPLEX_CODA"] = unit_query(
        units,
        lambda row: row["unit_type"] == "hangul"
        and json_count(row["coda_components_json"]) > 1,
        lambda row: (
            f"{row['morph_surface']}/{row['pos']}의 {row['unit_surface']} · "
            f"종성 {row['coda_jamo']}={row['coda_components_json']}"
        ),
        "coda_components_json",
    )
    queries["Q4_COMPOUND_NUCLEUS"] = unit_query(
        units,
        lambda row: row["unit_type"] == "hangul"
        and json_count(row["nucleus_components_json"]) > 1,
        lambda row: (
            f"{row['morph_surface']}/{row['pos']}의 {row['unit_surface']} · "
            f"중성 {row['nucleus_jamo']}={row['nucleus_components_json']}"
        ),
        "nucleus_components_json",
    )
    queries["Q5_STANDALONE_JAMO"] = unit_query(
        units,
        lambda row: row["unit_type"] == "jamo"
        and row["unit_idx_in_morph"] == "1",
        lambda row: (
            f"{row['morph_surface']}/{row['pos']} · 독립 자모 "
            f"{row['standalone_jamo']}→{row['unit_roman']} · unit "
            f"{row['unit_idx_in_morph']}/{row['unit_count_in_morph']}"
        ),
    )
    literal_master_ids = {
        utt_id
        for utt_id, rows in grouped_units.items()
        if any(row["unit_type"] == "literal" for row in rows)
    }
    queries["Q6_LITERAL_ALIGN_WARN"] = []
    for row in masters:
        if row["utt_id"] not in literal_master_ids or not row["align_warn"]:
            continue
        literals = sorted(
            {
                unit["unit_surface"]
                for unit in grouped_units[row["utt_id"]]
                if unit["unit_type"] == "literal"
            }
        )
        record = common_record(
            row,
            matched_table="utterance_master+morph_units",
            evidence=(
                f"literal={json.dumps(literals, ensure_ascii=False)} · "
                f"align_warn={row['align_warn']}"
            ),
        )
        record["components"] = json.dumps(literals, ensure_ascii=False)
        queries["Q6_LITERAL_ALIGN_WARN"].append(record)
    queries["Q7_INTER_EOJEOL_ENV"] = boundary_query(
        boundaries,
        lambda row: row["boundary_scope"] == "inter_eojeol"
        and bool(row["left_coda_jamo"])
        and truthy(row["right_onset_zero"]),
        lambda row: (
            f"{row['left_morph_surface']}/{row['left_pos']}[종성 {row['left_coda_jamo']}] "
            f"| {row['right_morph_surface']}/{row['right_pos']}"
            f"[초성 ㅇ, 중성 {row['right_nucleus_jamo']}]"
        ),
    )
    return catalog, queries


def select_review_results(
    *,
    catalog: list[dict[str, str]],
    queries: Mapping[str, list[dict[str, str]]],
    selected_ids: set[str],
    masters_by_id: Mapping[str, Mapping[str, str]],
) -> list[dict[str, object]]:
    results: list[dict[str, object]] = []
    result_order = 0
    for query in catalog:
        query_id = query["query_id"]
        all_hits = queries[query_id]
        review_hits = [row for row in all_hits if row["utt_id"] in selected_ids]
        review_hits.sort(
            key=lambda row: (row["year"], row["utt_id"], row["evidence"])
        )
        if not review_hits:
            raise RuntimeError(f"12발화 묶음에 조합 검색 예가 없음: {query_id}")
        for example_index, hit in enumerate(
            review_hits[:MAX_EXAMPLES_PER_QUERY], 1
        ):
            result_order += 1
            master = masters_by_id[hit["utt_id"]]
            prefix = f"{hit['year']}__{hit['utt_id']}"
            results.append(
                {
                    "result_order": result_order,
                    "query_id": query_id,
                    "query_name": query["query_name"],
                    "all_hit_rows": len(all_hits),
                    "all_hit_utterances": len({row["utt_id"] for row in all_hits}),
                    "review_hit_rows": len(review_hits),
                    "example_index": example_index,
                    "year": hit["year"],
                    "utt_id": hit["utt_id"],
                    "form": master["form"],
                    **hit,
                    "wav": f"{prefix}.wav",
                    "textgrid": f"{prefix}.TextGrid",
                    "csv": f"{prefix}.csv",
                }
            )
    return results


def add_hyperlink(cell, filename: str) -> None:
    cell.value = filename
    cell.hyperlink = filename
    cell.style = "Hyperlink"


def create_workbook(
    path: Path,
    *,
    catalog: list[dict[str, str]],
    results: list[dict[str, object]],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "검색_결과"
    sheet.sheet_view.showGridLines = False
    headers = [
        "결과순서",
        "검색ID",
        "검색유형",
        "60발화_적중행",
        "연도",
        "utt_id",
        "발화",
        "적중근거",
        *REVIEW_HEADERS,
        "WAV",
        "TextGrid",
        "CSV",
    ]
    sheet.append(headers)
    for result in results:
        row_index = sheet.max_row + 1
        sheet.append(
            [
                result["result_order"],
                result["query_id"],
                result["query_name"],
                result["all_hit_rows"],
                result["year"],
                result["utt_id"],
                result["form"],
                result["evidence"],
                "",
                "",
                "",
                "",
                "",
                None,
                None,
                None,
            ]
        )
        add_hyperlink(sheet.cell(row_index, 14), str(result["wav"]))
        add_hyperlink(sheet.cell(row_index, 15), str(result["textgrid"]))
        add_hyperlink(sheet.cell(row_index, 16), str(result["csv"]))

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:P{sheet.max_row}"
    widths = [9, 26, 25, 13, 8, 31, 42, 60, 13, 13, 13, 13, 42, 24, 27, 24]
    for column, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + column)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.row_dimensions[row[0].row].height = 62
    check_validation = DataValidation(
        type="list", formula1='"정상,문제,보류"', allow_blank=True
    )
    decision_validation = DataValidation(
        type="list", formula1='"승인,수정 필요,보류"', allow_blank=True
    )
    sheet.add_data_validation(check_validation)
    check_validation.add(f"I2:K{sheet.max_row}")
    sheet.add_data_validation(decision_validation)
    decision_validation.add(f"L2:L{sheet.max_row}")

    guide = workbook.create_sheet("검색식_안내")
    guide.sheet_view.showGridLines = False
    guide.append(
        ["검색ID", "연구 질문", "기계 조건", "해석 한계"]
    )
    for query in catalog:
        guide.append(
            [
                query["query_id"],
                query["research_question"],
                query["filter"],
                query["interpretation"],
            ]
        )
    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    guide.freeze_panes = "A2"
    for column, width in zip("ABCD", (28, 52, 82, 52)):
        guide.column_dimensions[column].width = width
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if row[0].row > 1:
            guide.row_dimensions[row[0].row].height = 58

    detail = workbook.create_sheet("세부_근거")
    detail.sheet_view.showGridLines = False
    detail_fields = [
        "result_order",
        "query_id",
        "year",
        "utt_id",
        "matched_table",
        "boundary_scope",
        "left_morph",
        "left_pos",
        "left_coda",
        "right_morph",
        "right_pos",
        "right_onset_zero",
        "right_nucleus",
        "target_morph",
        "target_pos",
        "morph_position",
        "unit_surface",
        "unit_roman",
        "components",
        "align_warn",
    ]
    detail.append(detail_fields)
    for result in results:
        detail.append([result.get(field, "") for field in detail_fields])
    for cell in detail[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    detail.freeze_panes = "A2"
    detail.auto_filter.ref = f"A1:T{detail.max_row}"
    for column in range(1, len(detail_fields) + 1):
        detail.column_dimensions[chr(64 + column)].width = 18
    detail.column_dimensions["D"].width = 31
    detail.column_dimensions["T"].width = 40
    for row in detail.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(path)


def validate_workbook(path: Path, *, expected_rows: int, bundle_root: Path) -> None:
    workbook = load_workbook(path, data_only=False)
    if workbook.sheetnames != ["검색_결과", "검색식_안내", "세부_근거"]:
        raise RuntimeError(f"sheet 불일치: {workbook.sheetnames}")
    sheet = workbook["검색_결과"]
    if sheet.max_row != expected_rows + 1:
        raise RuntimeError(f"결과 행 수 불일치: {sheet.max_row - 1}")
    link_count = 0
    for row in range(2, sheet.max_row + 1):
        for column in range(14, 17):
            cell = sheet.cell(row, column)
            if not cell.hyperlink:
                raise RuntimeError(f"hyperlink 누락: row={row}, col={column}")
            if not (bundle_root / str(cell.value)).is_file():
                raise RuntimeError(f"연결 파일 없음: {cell.value}")
            link_count += 1
    if link_count != expected_rows * 3:
        raise RuntimeError(f"link 수 불일치: {link_count}")
    if len(sheet.data_validations.dataValidation) != 2:
        raise RuntimeError("dropdown 수 불일치")
    workbook.close()


def build_demo(
    *, morph_root: Path, bundle_root: Path, output_root: Path
) -> dict[str, object]:
    morph_root = morph_root.resolve()
    bundle_root = bundle_root.resolve()
    output_root = output_root.resolve()
    partial = output_root.with_name(output_root.name + ".partial")
    if output_root.exists() or partial.exists():
        raise FileExistsError(f"기존 출력 덮어쓰기 금지: {output_root}")
    partial.mkdir(parents=True)
    try:
        masters = read_csv(morph_root / "utterance_master_v2.csv")
        morphs = read_csv(morph_root / "morph_tokens.csv")
        units = read_csv(morph_root / "morph_units.csv")
        boundaries = read_csv(morph_root / "morph_boundaries.csv")
        bundle_manifest_path = bundle_root / "BUNDLE_MANIFEST.json"
        bundle_manifest = json.loads(bundle_manifest_path.read_text(encoding="utf-8"))
        selected_ids = {
            str(row["utt_id"]) for row in bundle_manifest["selection"]
        }
        masters_by_id = {row["utt_id"]: row for row in masters}
        catalog, queries = build_queries(
            masters=masters,
            morphs=morphs,
            units=units,
            boundaries=boundaries,
        )
        results = select_review_results(
            catalog=catalog,
            queries=queries,
            selected_ids=selected_ids,
            masters_by_id=masters_by_id,
        )
        catalog_path = partial / "COMBINED_SEARCH_CATALOG.csv"
        results_path = partial / "COMBINED_SEARCH_RESULTS.csv"
        workbook_path = partial / "COMBINED_SEARCH_DEMO.xlsx"
        readme_path = partial / "COMBINED_SEARCH_README.md"
        write_csv(catalog_path, catalog, list(catalog[0]))
        result_fields = list(results[0])
        write_csv(results_path, results, result_fields)
        create_workbook(workbook_path, catalog=catalog, results=results)
        validate_workbook(
            workbook_path,
            expected_rows=len(results),
            bundle_root=bundle_root,
        )
        readme_path.write_text(
            """# 조합 검색 데모 검토

1. `COMBINED_SEARCH_DEMO.xlsx`를 연다.
2. `검색식_안내`에서 Q1--Q7의 조건과 해석 한계를 먼저 읽는다.
3. `검색_결과`에서 같은 검색ID의 첫 결과부터 적중 근거를 확인한다.
4. WAV·TextGrid·CSV 링크가 같은 발화를 여는지 확인한다.
5. `조건_일치`, `파일_연결`, `표시_이해`, `판정`, 필요한 경우 `메모`만 입력한다.

이 검토는 형태소·철자 환경으로 후보를 찾고 연결하는 인프라를 본다.
실제 ㄴ 삽입, 연음, 경음화 등의 실현 여부는 여기서 판정하지 않는다.
""",
            encoding="utf-8",
        )
        files = []
        for path in (catalog_path, results_path, workbook_path, readme_path):
            files.append(
                {
                    "relative_path": path.name,
                    "bytes": path.stat().st_size,
                    "sha256": sha256_file(path),
                }
            )
        manifest = {
            "schema_version": "morph_combined_search_demo.v1",
            "status": "success",
            "created_at": datetime.now().astimezone().isoformat(),
            "morph_root": str(morph_root),
            "bundle_root": str(bundle_root),
            "bundle_manifest_sha256": sha256_file(bundle_manifest_path),
            "counts": {
                "queries": len(catalog),
                "result_rows": len(results),
                "linked_payloads": len(results) * 3,
            },
            "query_counts": {
                query["query_id"]: {
                    "hit_rows": len(queries[query["query_id"]]),
                    "hit_utterances": len(
                        {row["utt_id"] for row in queries[query["query_id"]]}
                    ),
                    "review_examples": sum(
                        row["query_id"] == query["query_id"] for row in results
                    ),
                }
                for query in catalog
            },
            "files": files,
        }
        manifest_path = partial / "COMBINED_SEARCH_MANIFEST.json"
        manifest_path.write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        partial.replace(output_root)
        return manifest
    except Exception:
        # 실패 근거를 조사할 수 있게 partial을 보존한다.
        raise


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--morph-root", type=Path, required=True)
    parser.add_argument("--bundle-root", type=Path, required=True)
    parser.add_argument("--output-root", type=Path, required=True)
    args = parser.parse_args()
    try:
        manifest = build_demo(
            morph_root=args.morph_root,
            bundle_root=args.bundle_root,
            output_root=args.output_root,
        )
    except Exception as exc:
        print(f"[FAIL] {type(exc).__name__}: {exc}", file=sys.stderr)
        return 1
    print(json.dumps(manifest, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
