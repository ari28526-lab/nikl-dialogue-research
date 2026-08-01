"""6개년 MFA 파일럿에 검색용 로마자 음소 보조층을 만든다.

정본 4-tier, MFA DB, WAV, CSV는 읽기 전용이다. 전체 60발화 결과는 D:의
별도 output root에 만들고, 사용자 검토용 12발화 5-tier 사본과 workbook만
Dropbox의 기존 평면 검토 폴더에 추가한다.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import sys
import uuid
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).resolve().parent))

from phoneme_roman import (  # noqa: E402
    ROMAN_SYSTEM_VERSION,
    SCHEMA_VERSION,
    align_phone_to_reference,
    build_phone_inventory,
    classify_phone,
    expand_roman_eojeol,
    load_acoustic_meta,
    model_group_lookup,
    sequence_edit_count,
    split_roman_eojeols,
)
from pipeline_common import (  # noqa: E402
    atomic_write_json,
    file_fingerprint,
    git_commit,
    now_iso,
    sha256_file,
)
from retrofit_textgrid_2020_2024 import parse_mfa_textgrid  # noqa: E402


PILOT_SCHEMA_VERSION = "phoneme_roman_pilot.v1"
TARGET_TIERS = [
    "words",
    "phones_mfa",
    "utterance",
    "utterance_search",
]
FIVE_TIERS = TARGET_TIERS + ["phoneme_r_auto"]
SILENCE = {"", "sil", "sp", "spn", "<eps>", "<unk>"}


def read_csv(path: Path) -> list[dict[str, str]]:
    with open(path, encoding="utf-8-sig", newline="") as stream:
        return list(csv.DictReader(stream))


def write_csv(path: Path, rows: Sequence[Mapping[str, object]]) -> None:
    if not rows:
        raise ValueError(f"빈 CSV를 쓰지 않음: {path}")
    fields = list(rows[0])
    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_name(f".{path.name}.{uuid.uuid4().hex}.partial")
    with open(temp, "x", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields, extrasaction="raise")
        writer.writeheader()
        writer.writerows(rows)
        stream.flush()
        os.fsync(stream.fileno())
    os.replace(temp, path)


def interval_key(begin: float, end: float, label: str) -> tuple[float, float, str]:
    return round(float(begin), 6), round(float(end), 6), str(label)


def content(intervals: Iterable[tuple[float, float, str]]):
    return [row for row in intervals if row[2].strip().lower() not in SILENCE]


def _same_intervals(
    left: Sequence[tuple[float, float, str]],
    right: Sequence[tuple[float, float, str]],
) -> bool:
    if len(left) != len(right):
        return False
    return all(
        abs(a[0] - b[0]) <= 1e-6
        and abs(a[1] - b[1]) <= 1e-6
        and a[2] == b[2]
        for a, b in zip(left, right)
    )


def write_five_tier(
    source: Path,
    destination: Path,
    labels: Mapping[tuple[float, float, str], str],
) -> dict[str, object]:
    duration, tiers = parse_mfa_textgrid(source)
    if duration is None or list(tiers) != TARGET_TIERS:
        raise RuntimeError(
            f"4-tier source 계약 불일치: {source} tiers={list(tiers)}"
        )
    phones = tiers["phones_mfa"]
    new_tier = [
        (
            begin,
            end,
            "" if not label.strip() else labels[interval_key(begin, end, label)],
        )
        for begin, end, label in phones
    ]
    destination.parent.mkdir(parents=True, exist_ok=True)
    temp = destination.with_name(
        f".{destination.name}.{uuid.uuid4().hex}.partial"
    )
    # 앞 네 tier를 재직렬화하면 원본의 1 microsecond 허용오차가 새 빈 구간으로
    # 물질화될 수 있다. 원문 네 tier는 byte 그대로 두고 top-level size만 5로
    # 바꾼 뒤, phones_mfa와 정확히 같은 경계의 다섯 번째 tier만 덧붙인다.
    source_text = source.read_text(encoding="utf-8")
    updated, replacements = re.subn(
        r"(?m)^size = 4\s*$", "size = 5", source_text, count=1
    )
    if replacements != 1:
        raise RuntimeError(f"top-level tier size=4를 찾지 못함: {source}")
    tier_lines = [
        '    item [5]:',
        '        class = "IntervalTier"',
        '        name = "phoneme_r_auto"',
        '        xmin = 0',
        f'        xmax = {duration:.6f}',
        f'        intervals: size = {len(new_tier)}',
    ]
    for index, (begin, end, label) in enumerate(new_tier, 1):
        escaped = str(label).replace('"', '""')
        tier_lines.extend(
            [
                f"        intervals [{index}]:",
                f"            xmin = {begin:.6f}",
                f"            xmax = {end:.6f}",
                f'            text = "{escaped}"',
            ]
        )
    with open(temp, "x", encoding="utf-8", newline="\n") as stream:
        stream.write(updated.rstrip("\r\n") + "\n" + "\n".join(tier_lines) + "\n")
        stream.flush()
        os.fsync(stream.fileno())
    out_duration, out_tiers = parse_mfa_textgrid(temp)
    if out_duration is None or abs(out_duration - duration) > 1e-6:
        raise RuntimeError(f"5-tier duration 불일치: {destination}")
    if list(out_tiers) != FIVE_TIERS:
        raise RuntimeError(f"5-tier 순서 불일치: {list(out_tiers)}")
    for name in TARGET_TIERS:
        if not _same_intervals(tiers[name], out_tiers[name]):
            raise RuntimeError(f"기존 tier 변경 감지: {name} {source}")
    if not _same_intervals(new_tier, out_tiers["phoneme_r_auto"]):
        raise RuntimeError(f"phoneme tier 왕복 불일치: {destination}")
    os.replace(temp, destination)
    return {
        "path": str(destination.resolve()),
        "duration": duration,
        "tier_names": FIVE_TIERS,
        "original_four_tiers_unchanged": True,
        "phoneme_boundaries_equal_phones_mfa": True,
        "intervals": len(new_tier),
    }


def load_search_rows(pilot_root: Path) -> dict[str, dict[str, str]]:
    result: dict[str, dict[str, str]] = {}
    for year in map(str, range(2020, 2026)):
        path = pilot_root / "csv" / year / "search_master_selected.csv"
        for row in read_csv(path):
            utt_id = row["utt_id"]
            if utt_id in result:
                raise RuntimeError(f"중복 utt_id: {utt_id}")
            result[utt_id] = row
    return result


def review_ids(review_root: Path) -> dict[str, str]:
    result: dict[str, str] = {}
    for path in sorted(review_root.glob("*.wav")):
        year, marker, utt_id = path.stem.partition("__")
        if marker != "__" or not year.isdigit() or not utt_id:
            raise RuntimeError(f"검토 WAV 이름 규약 밖: {path.name}")
        result[utt_id] = year
    if len(result) != 12:
        raise RuntimeError(f"검토 WAV가 12개가 아님: {len(result)}")
    return result


def build_pilot(
    *,
    project_root: Path,
    pilot_root: Path,
    acoustic_model: Path,
    output_root: Path,
    review_root: Path,
) -> dict[str, object]:
    if output_root.exists():
        raise FileExistsError(f"새 output root만 허용: {output_root}")
    partial = output_root.with_name(
        f".{output_root.name}.{uuid.uuid4().hex}.partial"
    )
    partial.mkdir(parents=True)
    selection_path = pilot_root / "selection_manifest.csv"
    selection_rows = read_csv(selection_path)
    if len(selection_rows) != 60:
        raise RuntimeError(f"selection 60행 계약 불일치: {len(selection_rows)}")
    search_rows = load_search_rows(pilot_root)
    if set(search_rows) != {row["utt_id"] for row in selection_rows}:
        raise RuntimeError("selection과 search selected utt_id 불일치")

    meta = load_acoustic_meta(acoustic_model)
    group_lookup = model_group_lookup(meta)
    inventory = build_phone_inventory(meta)
    if len(inventory) != 107:
        raise RuntimeError(f"동결 non-silence phone 107개가 아님: {len(inventory)}")

    inventory_rows: list[dict[str, object]] = []
    for phone in inventory:
        inventory_rows.append(
            {
                "phone_mfa": phone.phone_mfa,
                "phone_class_r_auto": phone.phone_class_r_auto,
                "comparison_key": phone.comparison_key,
                "model_group_id": phone.model_group_id,
                "model_group_r": phone.model_group_r,
                "has_length": phone.has_length,
                "secondary_articulation": phone.secondary_articulation,
                "unreleased": phone.unreleased,
                "mapping_source": "frozen_acoustic_meta.phone_groups+project_rules.v1",
                "realization_judgment": "not_performed",
            }
        )
    for symbol in ("sil", "spn"):
        inventory_rows.append(
            {
                "phone_mfa": symbol,
                "phone_class_r_auto": "",
                "comparison_key": "",
                "model_group_id": "",
                "model_group_r": "NON_SPEECH" if symbol == "sil" else "OOV",
                "has_length": False,
                "secondary_articulation": "",
                "unreleased": False,
                "mapping_source": "frozen_acoustic_meta.reserved_symbol",
                "realization_judgment": "not_performed",
            }
        )

    all_intervals: list[dict[str, object]] = []
    all_correspondence: list[dict[str, object]] = []
    summaries: list[dict[str, object]] = []
    textgrid_reports: list[dict[str, object]] = []
    status_counts: Counter[str] = Counter()
    observed_phone_counts: Counter[str] = Counter()

    schema_root = (
        pilot_root / "research_schema_v1_20260731" / "textgrid_research_v1"
    )
    five_root = partial / "textgrid_5tier_optional"
    for selection in sorted(
        selection_rows, key=lambda row: (row["year"], row["utt_id"])
    ):
        year = selection["year"]
        utt_id = selection["utt_id"]
        session_id = selection["session_id"]
        row = search_rows[utt_id]
        source_tg = schema_root / year / session_id / f"{utt_id}.TextGrid"
        duration, tiers = parse_mfa_textgrid(source_tg)
        if duration is None or list(tiers) != TARGET_TIERS:
            raise RuntimeError(f"research TextGrid 계약 실패: {source_tg}")
        words = content(tiers["words"])
        labeled_phones = content(tiers["phones_mfa"])
        observed_phone_counts.update(label for _, _, label in labeled_phones)

        raw_map = json.loads(selection["eojeol_map_json"])
        word_to_source = {
            int(item["mfa_word_index"]): int(item["source_eojeol_index"])
            for item in raw_map
            if item.get("included_in_mfa")
        }
        if set(word_to_source) != set(range(len(words))):
            raise RuntimeError(
                f"eojeol map/words 불일치: {utt_id} map={sorted(word_to_source)} "
                f"words={len(words)}"
            )
        # eojeol_map은 실제 MFA 입력인 pron_reference_form 기준이다. 숫자/기호를
        # original_form에서 회복한 발화는 raw form_roman과 어절 수가 다를 수
        # 있으므로, 시간 투영의 철자 근거도 같은 입력 기준 열을 사용한다.
        orth_eojeols = split_roman_eojeols(row["pron_reference_form_roman"])
        pron_eojeols = split_roman_eojeols(row["pron_reference_roman"])
        required_source = max(word_to_source.values(), default=-1)
        if required_source >= len(orth_eojeols) or required_source >= len(pron_eojeols):
            raise RuntimeError(
                f"로마자 어절 coverage 불일치: {utt_id} source={required_source} "
                f"orth={len(orth_eojeols)} pron={len(pron_eojeols)}"
            )

        phone_label_by_interval: dict[tuple[float, float, str], str] = {}
        utterance_phone_labels: list[str] = []
        utterance_lexical_labels: list[str] = []
        review_count = 0
        compatible_count = 0
        reference_only_count = 0
        assigned_phone_keys: set[tuple[float, float, str]] = set()
        operation_index = 0
        for word_index, (word_begin, word_end, word_label) in enumerate(words):
            source_index = word_to_source[word_index]
            orth_value = orth_eojeols[source_index]
            pron_value = pron_eojeols[source_index]
            orth_units = expand_roman_eojeol(orth_value)
            reference_units = expand_roman_eojeol(pron_value)
            orth_pron_edits = sequence_edit_count(orth_units, reference_units)
            orth_pron_status = (
                "same_expanded_sequence" if orth_pron_edits == 0 else "rule_changed"
            )
            phone_intervals = [
                (begin, end, label)
                for begin, end, label in labeled_phones
                if begin >= word_begin - 1e-6 and end <= word_end + 1e-6
            ]
            phone_classes = [
                classify_phone(label, group_lookup)
                for _, _, label in phone_intervals
            ]
            ops = align_phone_to_reference(phone_classes, reference_units)
            phone_to_op = {
                op.phone_index: op for op in ops if op.phone_index is not None
            }
            for local_phone_index, ((begin, end, raw_phone), phone_class) in enumerate(
                zip(phone_intervals, phone_classes)
            ):
                op = phone_to_op[local_phone_index]
                reference = (
                    reference_units[op.reference_index]
                    if op.reference_index is not None
                    else None
                )
                lexical = reference.display if reference is not None else ""
                usable = op.status in {"exact", "position_compatible"}
                if usable:
                    compatible_count += 1
                    grid_label = lexical
                else:
                    review_count += 1
                    grid_label = f"?{lexical or ('+' + phone_class.phone_class_r_auto)}"
                status_counts[op.status] += 1
                key = interval_key(begin, end, raw_phone)
                if key in phone_label_by_interval:
                    raise RuntimeError(f"phone interval 중복 배정: {utt_id} {key}")
                phone_label_by_interval[key] = grid_label
                assigned_phone_keys.add(key)
                utterance_phone_labels.append(phone_class.phone_class_r_auto)
                utterance_lexical_labels.append(grid_label)
                interval_row = {
                    "year": year,
                    "utt_id": utt_id,
                    "session_id": session_id,
                    "word_index": word_index + 1,
                    "source_eojeol_index": source_index + 1,
                    "word": word_label,
                    "begin": round(begin, 6),
                    "end": round(end, 6),
                    "phone_mfa": raw_phone,
                    "phone_class_r_auto": phone_class.phone_class_r_auto,
                    "phoneme_lexical_r_auto": lexical,
                    "textgrid_label": grid_label,
                    "mapping_status": op.status,
                    "automatic_use": "usable" if usable else "review",
                    "model_group_r": phone_class.model_group_r,
                    "has_length": phone_class.has_length,
                    "secondary_articulation": phone_class.secondary_articulation,
                    "unreleased": phone_class.unreleased,
                    "pron_source_token": reference.source_token if reference else "",
                    "pron_component_index": reference.component_index if reference else "",
                    "pron_component_count": reference.component_count if reference else "",
                    "orth_roman_eojeol": orth_value,
                    "orth_reference_source_field": "pron_reference_form_roman",
                    "form_roman_raw": row["form_roman"],
                    "pron_reference_roman_eojeol": pron_value,
                    "orth_pron_status": orth_pron_status,
                    "orth_pron_edit_count": orth_pron_edits,
                    "pron_reference_source": row["pron_reference_source"],
                    "pron_reference_status": row["pron_reference_status"],
                    "realization_judgment": "not_performed",
                    "schema_version": SCHEMA_VERSION,
                }
                all_intervals.append(interval_row)
            for op in ops:
                operation_index += 1
                phone_interval = (
                    phone_intervals[op.phone_index]
                    if op.phone_index is not None
                    else None
                )
                phone_class = (
                    phone_classes[op.phone_index]
                    if op.phone_index is not None
                    else None
                )
                reference = (
                    reference_units[op.reference_index]
                    if op.reference_index is not None
                    else None
                )
                if op.status == "reference_only":
                    reference_only_count += 1
                    review_count += 1
                    status_counts[op.status] += 1
                all_correspondence.append(
                    {
                        "year": year,
                        "utt_id": utt_id,
                        "word_index": word_index + 1,
                        "word": word_label,
                        "operation_index": operation_index,
                        "operation": op.operation,
                        "mapping_status": op.status,
                        "begin": round(phone_interval[0], 6) if phone_interval else "",
                        "end": round(phone_interval[1], 6) if phone_interval else "",
                        "phone_mfa": phone_interval[2] if phone_interval else "",
                        "phone_class_r_auto": (
                            phone_class.phone_class_r_auto if phone_class else ""
                        ),
                        "phoneme_lexical_r_auto": reference.display if reference else "",
                        "pron_source_token": reference.source_token if reference else "",
                        "pron_syllable_index": (
                            reference.syllable_index if reference else ""
                        ),
                        "orth_roman_eojeol": orth_value,
                        "pron_reference_roman_eojeol": pron_value,
                        "orth_pron_status": orth_pron_status,
                        "alignment_cost": op.cost,
                        "realization_judgment": "not_performed",
                    }
                )

        expected_keys = {
            interval_key(begin, end, label)
            for begin, end, label in labeled_phones
        }
        if assigned_phone_keys != expected_keys:
            raise RuntimeError(
                f"phone interval coverage 불일치: {utt_id} "
                f"missing={expected_keys-assigned_phone_keys} "
                f"extras={assigned_phone_keys-expected_keys}"
            )
        for begin, end, label in tiers["phones_mfa"]:
            if label.strip():
                continue
            all_intervals.append(
                {
                    "year": year,
                    "utt_id": utt_id,
                    "session_id": session_id,
                    "word_index": "",
                    "source_eojeol_index": "",
                    "word": "",
                    "begin": round(begin, 6),
                    "end": round(end, 6),
                    "phone_mfa": "",
                    "phone_class_r_auto": "",
                    "phoneme_lexical_r_auto": "",
                    "textgrid_label": "",
                    "mapping_status": "non_speech_blank",
                    "automatic_use": "not_applicable",
                    "model_group_r": "NON_SPEECH",
                    "has_length": False,
                    "secondary_articulation": "",
                    "unreleased": False,
                    "pron_source_token": "",
                    "pron_component_index": "",
                    "pron_component_count": "",
                    "orth_roman_eojeol": "",
                    "orth_reference_source_field": "pron_reference_form_roman",
                    "form_roman_raw": row["form_roman"],
                    "pron_reference_roman_eojeol": "",
                    "orth_pron_status": "",
                    "orth_pron_edit_count": "",
                    "pron_reference_source": row["pron_reference_source"],
                    "pron_reference_status": row["pron_reference_status"],
                    "realization_judgment": "not_performed",
                    "schema_version": SCHEMA_VERSION,
                }
            )

        destination = five_root / year / session_id / f"{utt_id}.TextGrid"
        final_destination = output_root / destination.relative_to(partial)
        textgrid_reports.append(
            {
                "year": year,
                "utt_id": utt_id,
                **write_five_tier(source_tg, destination, phone_label_by_interval),
            }
        )
        total_labeled = len(labeled_phones)
        summaries.append(
            {
                "year": year,
                "utt_id": utt_id,
                "session_id": session_id,
                "speaker_id": selection["speaker_id"],
                "form": row["form"],
                "form_roman": row["form_roman"],
                "pron_reference_form": row["pron_reference_form"],
                "pron_reference_form_roman": row["pron_reference_form_roman"],
                "pron_reference_hangul": row["pron_reference_hangul"],
                "pron_reference_roman": row["pron_reference_roman"],
                "pron_reference_source": row["pron_reference_source"],
                "pron_reference_status": row["pron_reference_status"],
                "phone_class_r_sequence": " ".join(utterance_phone_labels),
                "phoneme_lexical_r_sequence": " ".join(utterance_lexical_labels),
                "labeled_phone_intervals": total_labeled,
                "exact_or_position_compatible": compatible_count,
                "review_phone_intervals": review_count - reference_only_count,
                "reference_only_units": reference_only_count,
                "automatic_coverage": (
                    round(compatible_count / total_labeled, 6)
                    if total_labeled
                    else 0.0
                ),
                "pilot_status": (
                    "auto_compatible"
                    if review_count == 0
                    else "researcher_review_needed"
                ),
                "realization_judgment": "not_performed",
                "source_textgrid": str(source_tg.resolve()),
                "five_tier_textgrid": str(final_destination.resolve()),
            }
        )

    unknown_observed = set(observed_phone_counts) - set(group_lookup)
    if unknown_observed:
        raise RuntimeError(f"관측 phone inventory 밖 기호: {sorted(unknown_observed)}")
    if any(row["phone_mfa"] == "spn" for row in all_intervals):
        raise RuntimeError("파일럿 실제 interval에 spn 존재")

    write_csv(partial / "PHONE_ROMAN_INVENTORY.csv", inventory_rows)
    write_csv(partial / "PHONE_ROMAN_INTERVALS.csv", all_intervals)
    write_csv(partial / "PHONEME_ROMAN_CORRESPONDENCE.csv", all_correspondence)
    write_csv(partial / "UTTERANCE_PHONEME_ROMAN_SUMMARY.csv", summaries)

    review_set = review_ids(review_root)
    summary_by_utt = {str(row["utt_id"]): row for row in summaries}
    review_summaries = [
        summary_by_utt[utt_id]
        for utt_id in sorted(review_set, key=lambda uid: (review_set[uid], uid))
    ]
    review_corr = [
        row for row in all_correspondence if str(row["utt_id"]) in review_set
    ]
    review_delivery = partial / "review_delivery"
    review_delivery.mkdir()
    delivered_tg: dict[str, Path] = {}
    for row in review_summaries:
        year = str(row["year"])
        utt_id = str(row["utt_id"])
        source = Path(str(row["five_tier_textgrid"]).replace(str(output_root), str(partial)))
        destination = review_delivery / f"{year}__{utt_id}__phoneme_r_auto.TextGrid"
        shutil.copy2(source, destination)
        delivered_tg[utt_id] = review_root / destination.name

    workbook_path = review_delivery / "PHONEME_ROMAN_PILOT.xlsx"
    build_workbook(
        workbook_path,
        review_summaries=review_summaries,
        review_correspondence=review_corr,
        inventory_rows=inventory_rows,
        review_root=review_root,
        delivered_textgrids=delivered_tg,
    )
    verify_workbook(workbook_path, expected_rows=12)
    readme = review_delivery / "PHONEME_ROMAN_README.md"
    readme.write_text(
        "# MFA 로마자 음소 보조층 파일럿\n\n"
        "- 기존 `phones_mfa` IPA와 시간은 변경하지 않았다.\n"
        "- `phone_class_r_auto`는 동결 acoustic phone을 검색용 로마자로 범주화한다.\n"
        "- `phoneme_lexical_r_auto`는 철자·규칙 예측발음을 참조해 MFA 시간에 "
        "투영한 자동 후보이며 실제 실현 판정이 아니다.\n"
        "- `?`로 시작하는 5-tier 라벨은 자동 대응이 완전 일치하지 않아 검토가 필요하다.\n"
        "- 기본 연구 TextGrid는 계속 4-tier이며, 이 5-tier 파일은 선택적 사본이다.\n",
        encoding="utf-8",
    )
    review_csv = review_delivery / "PHONEME_ROMAN_REVIEW.csv"
    write_csv(review_csv, review_summaries)

    output_files = [
        partial / "PHONE_ROMAN_INVENTORY.csv",
        partial / "PHONE_ROMAN_INTERVALS.csv",
        partial / "PHONEME_ROMAN_CORRESPONDENCE.csv",
        partial / "UTTERANCE_PHONEME_ROMAN_SUMMARY.csv",
    ]
    def final_output_fingerprint(path: Path) -> dict[str, object]:
        fingerprint = file_fingerprint(path, with_sha256=True)
        fingerprint["path"] = str(
            (output_root / path.relative_to(partial)).resolve()
        )
        return fingerprint

    manifest = {
        "schema_version": PILOT_SCHEMA_VERSION,
        "status": "success_researcher_review_pending",
        "recorded_at": now_iso(),
        "purpose": "additive Romanized phoneme candidates for search, not realization judgment",
        "contracts": {
            "source_textgrid_tiers": TARGET_TIERS,
            "optional_copy_tiers": FIVE_TIERS,
            "phones_mfa_unchanged": True,
            "default_four_tier_unchanged": True,
            "realization_judgment_performed": False,
            "roman_system": ROMAN_SYSTEM_VERSION,
            "aux_schema": SCHEMA_VERSION,
        },
        "inputs": {
            "pilot_root": str(pilot_root.resolve()),
            "selection_manifest": file_fingerprint(
                selection_path, with_sha256=True
            ),
            "acoustic_model": file_fingerprint(acoustic_model, with_sha256=True),
            "review_root": str(review_root.resolve()),
        },
        "counts": {
            "utterances": len(summaries),
            "review_utterances": len(review_summaries),
            "inventory_non_silence_phones": len(inventory),
            "observed_unique_phones": len(observed_phone_counts),
            "labeled_phone_intervals": sum(observed_phone_counts.values()),
            "correspondence_rows": len(all_correspondence),
            "five_tier_textgrids": len(textgrid_reports),
            "mapping_status": dict(sorted(status_counts.items())),
        },
        "outputs": [final_output_fingerprint(path) for path in output_files],
        "review_delivery": {
            "files": len(list(review_delivery.iterdir())),
            "workbook": final_output_fingerprint(workbook_path),
        },
        "runtime": {
            "git_commit": git_commit(project_root),
            "python": sys.executable,
        },
    }
    atomic_write_json(partial / "PILOT_MANIFEST.json", manifest)
    os.replace(partial, output_root)

    # Dropbox는 새 이름만 원자적 복사한다. 기존 REVIEW와 payload는 건드리지 않는다.
    delivered: list[dict[str, object]] = []
    final_review_delivery = output_root / "review_delivery"
    for source in sorted(final_review_delivery.iterdir()):
        destination = review_root / source.name
        if destination.exists():
            raise FileExistsError(f"검토 전달 파일이 이미 존재: {destination}")
        temp = destination.with_name(
            f".{destination.name}.{uuid.uuid4().hex}.partial"
        )
        shutil.copy2(source, temp)
        if sha256_file(source) != sha256_file(temp):
            raise RuntimeError(f"Dropbox 복사 SHA 불일치: {destination}")
        os.replace(temp, destination)
        delivered.append(file_fingerprint(destination, with_sha256=True))
    delivery_manifest = {
        "schema_version": "phoneme_roman_delivery.v1",
        "status": "success",
        "recorded_at": now_iso(),
        "source_output": str(output_root.resolve()),
        "destination": str(review_root.resolve()),
        "existing_review_files_modified": False,
        "files": delivered,
    }
    atomic_write_json(review_root / "PHONEME_ROMAN_DELIVERY_MANIFEST.json", delivery_manifest)
    return {
        "manifest": output_root / "PILOT_MANIFEST.json",
        "workbook": review_root / "PHONEME_ROMAN_PILOT.xlsx",
        "counts": manifest["counts"],
        "delivered_files": len(delivered) + 1,
    }


def _style_table(ws, headers: Sequence[str], rows: Sequence[Sequence[object]]) -> None:
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)


def build_workbook(
    path: Path,
    *,
    review_summaries: Sequence[Mapping[str, object]],
    review_correspondence: Sequence[Mapping[str, object]],
    inventory_rows: Sequence[Mapping[str, object]],
    review_root: Path,
    delivered_textgrids: Mapping[str, Path],
) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "발화_검토"
    headers = [
        "번호",
        "연도",
        "utt_id",
        "발화",
        "철자_로마자",
        "MFA입력_철자_로마자",
        "예측발음_로마자",
        "MFA_phone_로마자열",
        "자동_음소_로마자열",
        "자동일치_분절",
        "검토필요_phone",
        "참조만_토큰",
        "자동_coverage",
        "WAV",
        "기존_4tier",
        "새_5tier",
        "phone_범주",
        "철자_예측_참조",
        "경계",
        "판정",
        "메모",
    ]
    review_rows = []
    for index, row in enumerate(review_summaries, 1):
        year = str(row["year"])
        utt_id = str(row["utt_id"])
        review_rows.append(
            [
                index,
                int(year),
                utt_id,
                row["form"],
                row["form_roman"],
                row["pron_reference_form_roman"],
                row["pron_reference_roman"],
                row["phone_class_r_sequence"],
                row["phoneme_lexical_r_sequence"],
                row["exact_or_position_compatible"],
                row["review_phone_intervals"],
                row["reference_only_units"],
                row["automatic_coverage"],
                "WAV 열기",
                "기존 4-tier",
                "새 5-tier",
                "",
                "",
                "",
                "",
                "",
            ]
        )
    _style_table(ws, headers, review_rows)
    widths = {
        "A": 7,
        "B": 8,
        "C": 34,
        "D": 24,
        "E": 42,
        "F": 42,
        "G": 42,
        "H": 45,
        "I": 45,
        "J": 12,
        "K": 13,
        "L": 12,
        "M": 13,
        "N": 13,
        "O": 15,
        "P": 15,
        "Q": 15,
        "R": 17,
        "S": 12,
        "T": 14,
        "U": 32,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.column_dimensions["A"].hidden = False
    for row_index, row in enumerate(review_summaries, 2):
        year = str(row["year"])
        utt_id = str(row["utt_id"])
        wav = review_root / f"{year}__{utt_id}.wav"
        old_tg = review_root / f"{year}__{utt_id}.TextGrid"
        new_tg = delivered_textgrids[utt_id]
        for column, target in ((14, wav), (15, old_tg), (16, new_tg)):
            if target.parent.resolve() != review_root.resolve():
                raise RuntimeError(f"검토 링크 대상이 전달 폴더 밖에 있음: {target}")
            cell = ws.cell(row=row_index, column=column)
            # Workbook과 36개 검토 파일을 같은 Dropbox 폴더에 두고
            # 다른 컴퓨터에서도 열 수 있도록 절대경로를 저장하지 않는다.
            cell.hyperlink = target.name
            cell.style = "Hyperlink"
        ws.cell(row=row_index, column=13).number_format = "0.0%"
    normal_validation = DataValidation(
        type="list", formula1='"정상,문제,보류"', allow_blank=True
    )
    decision_validation = DataValidation(
        type="list", formula1='"승인,수정필요,보류"', allow_blank=True
    )
    ws.add_data_validation(normal_validation)
    normal_validation.add(f"Q2:S{ws.max_row}")
    ws.add_data_validation(decision_validation)
    decision_validation.add(f"T2:T{ws.max_row}")
    warning_fill = PatternFill("solid", fgColor="FCE4D6")
    ws.conditional_formatting.add(
        f"K2:K{ws.max_row}",
        FormulaRule(formula=["K2>0"], fill=warning_fill),
    )
    ws.conditional_formatting.add(
        f"L2:L{ws.max_row}",
        FormulaRule(formula=["L2>0"], fill=warning_fill),
    )

    detail = workbook.create_sheet("대응_세부")
    detail_headers = [
        "연도",
        "utt_id",
        "어절",
        "어절번호",
        "operation",
        "status",
        "시작",
        "끝",
        "MFA_IPA",
        "phone_class_r_auto",
        "phoneme_lexical_r_auto",
        "예측원토큰",
        "철자_로마자_어절",
        "예측발음_로마자_어절",
        "철자_예측_관계",
        "실현판정",
    ]
    detail_rows = [
        [
            int(str(row["year"])),
            row["utt_id"],
            row["word"],
            row["word_index"],
            row["operation"],
            row["mapping_status"],
            row["begin"],
            row["end"],
            row["phone_mfa"],
            row["phone_class_r_auto"],
            row["phoneme_lexical_r_auto"],
            row["pron_source_token"],
            row["orth_roman_eojeol"],
            row["pron_reference_roman_eojeol"],
            row["orth_pron_status"],
            row["realization_judgment"],
        ]
        for row in review_correspondence
    ]
    _style_table(detail, detail_headers, detail_rows)
    for col, width in {
        "A": 8,
        "B": 34,
        "C": 18,
        "D": 10,
        "E": 15,
        "F": 20,
        "G": 11,
        "H": 11,
        "I": 13,
        "J": 22,
        "K": 25,
        "L": 15,
        "M": 35,
        "N": 35,
        "O": 22,
        "P": 20,
    }.items():
        detail.column_dimensions[col].width = width
    for column in (7, 8):
        for cell in detail.iter_cols(
            min_col=column, max_col=column, min_row=2, max_row=detail.max_row
        ):
            cell[0].number_format = "0.000000"

    inventory_ws = workbook.create_sheet("기호_사전")
    inv_headers = [
        "MFA_IPA",
        "phone_class_r_auto",
        "비교키",
        "model_group_id",
        "model_group_r",
        "장음표지",
        "이차조음",
        "미방출",
        "근거",
        "실현판정",
    ]
    inv_rows = [
        [
            row["phone_mfa"],
            row["phone_class_r_auto"],
            row["comparison_key"],
            row["model_group_id"],
            row["model_group_r"],
            row["has_length"],
            row["secondary_articulation"],
            row["unreleased"],
            row["mapping_source"],
            row["realization_judgment"],
        ]
        for row in inventory_rows
    ]
    _style_table(inventory_ws, inv_headers, inv_rows)
    for col, width in {
        "A": 15,
        "B": 22,
        "C": 14,
        "D": 15,
        "E": 22,
        "F": 12,
        "G": 20,
        "H": 12,
        "I": 48,
        "J": 20,
    }.items():
        inventory_ws.column_dimensions[col].width = width

    method = workbook.create_sheet("방법_읽기")
    method_rows = [
        ("목적", "MFA 정렬 phone을 실제 연구 검색에 편한 프로젝트 로마자로 보조 표시"),
        ("phones_mfa", "동결 MFA가 사전/G2P 발음열을 음성에 강제정렬한 IPA 원값. 변경하지 않음"),
        ("phone_class_r_auto", "MFA IPA의 기계적 로마자 범주. acoustic meta phone_groups와 명시 규칙 사용"),
        ("phoneme_lexical_r_auto", "철자 로마자·규칙 예측발음 로마자를 함께 참조해 MFA 시간에 투영한 자동 음소 후보"),
        ("철자_로마자", "형태·철자 환경 검색용. 실제 발음과 동일시하지 않음"),
        ("예측발음_로마자", "필수 규칙 기반 검색 참조. 사전의 모든 변이나 실제 실현을 뜻하지 않음"),
        ("? 라벨", "phone과 예측발음의 자동 대응이 exact/position-compatible이 아니므로 연구자 검토 필요"),
        ("새 5-tier", "기존 네 tier를 그대로 복제하고 phoneme_r_auto만 덧붙인 선택적 사본"),
        ("실현 여부", "이 단계에서는 판정하지 않음. WAV와 TextGrid를 보고 연구 단계에서 별도 판정"),
        ("아침 1단계", "발화_검토 1행의 WAV, 기존 4-tier, 새 5-tier를 차례로 열기"),
        ("아침 2단계", "MFA_phone_로마자열과 자동_음소_로마자열이 이해 가능한지 확인"),
        ("아침 3단계", "P:R을 입력하고, 문제일 때만 대응_세부에서 같은 utt_id를 필터"),
        ("주의", "자동 보조층을 실제 음성 실현의 관찰값 또는 정답 음소 전사로 인용하지 않음"),
    ]
    _style_table(method, ["항목", "설명"], method_rows)
    method.column_dimensions["A"].width = 24
    method.column_dimensions["B"].width = 95
    method.row_dimensions[1].height = 28

    path.parent.mkdir(parents=True, exist_ok=True)
    temp = path.with_name(f".{path.name}.{uuid.uuid4().hex}.partial")
    workbook.save(temp)
    os.replace(temp, path)


def verify_workbook(path: Path, *, expected_rows: int) -> None:
    workbook = load_workbook(path, data_only=False, read_only=False)
    expected_sheets = ["발화_검토", "대응_세부", "기호_사전", "방법_읽기"]
    if workbook.sheetnames != expected_sheets:
        raise RuntimeError(f"workbook sheet 불일치: {workbook.sheetnames}")
    review = workbook["발화_검토"]
    if review.max_row - 1 != expected_rows:
        raise RuntimeError(f"workbook 검토행 불일치: {review.max_row-1}")
    links = sum(
        1
        for row in review.iter_rows(min_row=2)
        for cell in row
        if cell.hyperlink is not None
    )
    if links != expected_rows * 3:
        raise RuntimeError(f"workbook 링크 불일치: {links}")
    nonportable_links = []
    for row in review.iter_rows(min_row=2):
        for cell in row:
            if cell.hyperlink is None:
                continue
            target = str(cell.hyperlink.target)
            if Path(target).is_absolute() or ":" in target or "/" in target or "\\" in target:
                nonportable_links.append(f"{cell.coordinate}={target}")
    if nonportable_links:
        raise RuntimeError(f"workbook 비휴대형 절대경로 링크: {nonportable_links[:5]}")
    if len(review.data_validations.dataValidation) != 2:
        raise RuntimeError("workbook dropdown 수 불일치")
    errors = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith(
                    ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
                ):
                    errors.append(f"{sheet.title}!{cell.coordinate}={cell.value}")
    if errors:
        raise RuntimeError(f"workbook 오류 셀: {errors[:10]}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--project-root", type=Path, required=True)
    parser.add_argument("--pilot-root", type=Path, required=True)
    parser.add_argument("--acoustic-model", type=Path, required=True)
    parser.add_argument("--output-root", type=Path, required=True)
    parser.add_argument("--review-root", type=Path, required=True)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    result = build_pilot(
        project_root=args.project_root.resolve(),
        pilot_root=args.pilot_root.resolve(),
        acoustic_model=args.acoustic_model.resolve(),
        output_root=args.output_root.resolve(),
        review_root=args.review_root.resolve(),
    )
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
