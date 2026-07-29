"""Validate a researcher-filled common-pronunciation r2 workbook.

The clean workbook is an immutable template.  A researcher must save a copy
and may change only the three yellow columns in ``발음검토``:

* R: researcher_decision
* S: researcher_custom_phone
* U: researcher_notes

This validator never edits the workbook, D: review ledgers, G2P shards, or the
final dictionary.  It compares every immutable cell, formula, hyperlink,
merged range, table, and data-validation contract with the clean template.
Only a complete set of 27 affirmative decisions inside the frozen acoustic
phone inventory can produce normalized decision and correction CSV files.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from build_common_pron_mfa_lexicon import (  # noqa: E402
    acoustic_phone_inventory,
    phone_inventory_contract,
)
from build_common_pron_researcher_review_xlsx import (  # noqa: E402
    REVIEW_HEADERS,
    SCHEMA_VERSION as WORKBOOK_SCHEMA_VERSION,
)
from pipeline_common import (  # noqa: E402
    atomic_text_writer,
    atomic_write_json,
    file_fingerprint,
    runtime_snapshot,
)


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SCHEMA_VERSION = "common_pron_researcher_decision_validation.v1"
REVIEW_SHEET = "발음검토"
EXPECTED_SHEETS = (
    "검토안내",
    REVIEW_SHEET,
    "발화근거",
    "모델후보원자료",
    "근거자료",
    "동결모델프로브",
    "재현계약",
)
EDITABLE_COLUMNS = {"R", "S", "U"}
EDITABLE_COORDINATES = {
    f"{column}{row}"
    for column in EDITABLE_COLUMNS
    for row in range(2, 29)
}
ALLOWED_DECISIONS = {
    "pending",
    "approve_recommended",
    "approve_alternative",
    "approve_custom",
    "hold",
    "reject",
}
AFFIRMATIVE_DECISIONS = {
    "approve_recommended",
    "approve_alternative",
    "approve_custom",
}
FORBIDDEN_PRONUNCIATION_PHONES = {"sil", "spn"}

DECISION_FIELDS = (
    "review_order",
    "category",
    "token",
    "model_input",
    "model_candidate_phone",
    "recommendation_action",
    "researcher_decision",
    "approved_pron_phones_mfa",
    "approved_phone_source",
    "approved_phone_provenance",
    "researcher_notes",
    "source_handling",
    "source_url",
    "reason",
    "example_utt_id",
    "review_wav",
)
CORRECTION_FIELDS = (
    "review_order",
    "token",
    "correction_kind",
    "raw_search_token",
    "normalized_search_token",
    "source_notation",
    "approved_pron_phones_mfa",
    "researcher_decision",
    "researcher_notes",
    "example_utt_id",
)
CORRECTION_SPECS = {
    "외곬수적인": {
        "correction_kind": "source_spelling",
        "raw_search_token": "외곬수적인",
        "normalized_search_token": "외골수적인",
        "source_notation": "원 JSON 표기 보존",
    },
    "천구백칤비육": {
        "correction_kind": "numeric_placeholder",
        "raw_search_token": "천구백칤비육",
        "normalized_search_token": "천구백칠십육",
        "source_notation": "form 1976년; raw placeholder 보존",
    },
}


class WorkbookIntegrityError(RuntimeError):
    """The filled workbook changed an immutable research contract."""


def clean(value: object) -> str:
    return str(value or "").strip()


def normalize_phone(value: object) -> str:
    return " ".join(clean(value).split())


def phone_tokens(value: object) -> tuple[str, ...]:
    normalized = normalize_phone(value)
    return tuple(normalized.split()) if normalized else ()


def lexical_phone_inventory(loader_inventory: set[str]) -> set[str]:
    """Remove alignment-only symbols from dictionary pronunciation phones."""
    return set(loader_inventory) - FORBIDDEN_PRONUNCIATION_PHONES


def fingerprint_matches(path: Path, expected: dict[str, Any]) -> bool:
    actual = file_fingerprint(path, with_sha256=True)
    return (
        actual["bytes"] == expected.get("bytes")
        and actual["sha256"] == expected.get("sha256")
    )


def hyperlink_target(cell) -> str:
    hyperlink = cell.hyperlink
    if hyperlink is None:
        return ""
    return clean(hyperlink.target or hyperlink.location)


def table_contract(sheet) -> dict[str, str]:
    return {
        name: clean(sheet.tables[name].ref)
        for name in sorted(sheet.tables)
    }


def validation_contract(sheet) -> list[tuple[str, str, str]]:
    return sorted(
        (
            clean(validation.sqref),
            clean(validation.type),
            clean(validation.formula1),
        )
        for validation in sheet.data_validations.dataValidation
    )


def compare_workbook_contract(template, filled) -> int:
    """Return the number of immutable cells compared."""
    if tuple(template.sheetnames) != EXPECTED_SHEETS:
        raise WorkbookIntegrityError(
            f"template sheet contract mismatch: {template.sheetnames}"
        )
    if tuple(filled.sheetnames) != EXPECTED_SHEETS:
        raise WorkbookIntegrityError(
            f"filled sheet contract mismatch: {filled.sheetnames}"
        )

    compared = 0
    for sheet_name in EXPECTED_SHEETS:
        left = template[sheet_name]
        right = filled[sheet_name]
        if (left.max_row, left.max_column) != (
            right.max_row,
            right.max_column,
        ):
            raise WorkbookIntegrityError(
                f"sheet dimension changed: {sheet_name} "
                f"{left.max_row}x{left.max_column} -> "
                f"{right.max_row}x{right.max_column}"
            )
        left_merges = sorted(str(item) for item in left.merged_cells.ranges)
        right_merges = sorted(str(item) for item in right.merged_cells.ranges)
        if left_merges != right_merges:
            raise WorkbookIntegrityError(
                f"merged range contract changed: {sheet_name}"
            )
        if table_contract(left) != table_contract(right):
            raise WorkbookIntegrityError(
                f"table contract changed: {sheet_name}"
            )
        if validation_contract(left) != validation_contract(right):
            raise WorkbookIntegrityError(
                f"data validation contract changed: {sheet_name}"
            )

        for row in range(1, left.max_row + 1):
            for column in range(1, left.max_column + 1):
                left_cell = left.cell(row, column)
                right_cell = right.cell(row, column)
                if (
                    sheet_name == REVIEW_SHEET
                    and left_cell.coordinate in EDITABLE_COORDINATES
                ):
                    continue
                compared += 1
                if left_cell.value != right_cell.value:
                    raise WorkbookIntegrityError(
                        "immutable cell changed: "
                        f"{sheet_name}!{left_cell.coordinate}"
                    )
                if hyperlink_target(left_cell) != hyperlink_target(right_cell):
                    raise WorkbookIntegrityError(
                        "immutable hyperlink changed: "
                        f"{sheet_name}!{left_cell.coordinate}"
                    )
    return compared


def review_rows_from_sheet(sheet) -> list[dict[str, object]]:
    headers = tuple(clean(sheet.cell(1, column).value) for column in range(1, 25))
    if headers != REVIEW_HEADERS:
        raise WorkbookIntegrityError(
            f"review headers changed: {headers}"
        )
    rows: list[dict[str, object]] = []
    for row_number in range(2, 29):
        row = {
            header: sheet.cell(row_number, column).value
            for column, header in enumerate(REVIEW_HEADERS, start=1)
        }
        rows.append(row)
    tokens = [clean(row["token"]) for row in rows]
    categories = Counter(clean(row["category"]) for row in rows)
    if len(rows) != 27 or len(tokens) != len(set(tokens)):
        raise WorkbookIntegrityError("review rows are missing or duplicated")
    if categories != Counter({"no_path": 23, "jamo_ls": 4}):
        raise WorkbookIntegrityError(
            f"review category contract changed: {dict(categories)}"
        )
    return rows


def validate_phone(
    *,
    token: str,
    label: str,
    value: object,
    inventory: set[str],
    required: bool,
) -> str:
    normalized = normalize_phone(value)
    if required and not normalized:
        raise WorkbookIntegrityError(
            f"{token}: {label} phone is required"
        )
    if not normalized:
        return ""
    phones = set(phone_tokens(normalized))
    forbidden = phones & FORBIDDEN_PRONUNCIATION_PHONES
    unknown = phones - inventory
    if forbidden:
        raise WorkbookIntegrityError(
            f"{token}: {label} contains forbidden phones {sorted(forbidden)}"
        )
    if unknown:
        raise WorkbookIntegrityError(
            f"{token}: {label} is outside acoustic inventory "
            f"{sorted(unknown)}"
        )
    return normalized


def normalize_decisions(
    rows: list[dict[str, object]], inventory: set[str]
) -> tuple[list[dict[str, str]], list[dict[str, str]], dict[str, Any]]:
    decisions: list[dict[str, str]] = []
    corrections: list[dict[str, str]] = []
    errors: list[str] = []
    counts: Counter[str] = Counter()

    for row in rows:
        token = clean(row["token"])
        decision = clean(row["researcher_decision"])
        notes = clean(row["researcher_notes"])
        action = clean(row["recommendation_action"])
        if decision not in ALLOWED_DECISIONS:
            errors.append(f"{token}: unsupported decision {decision!r}")
            continue
        counts[decision] += 1

        try:
            candidate = validate_phone(
                token=token,
                label="model_candidate",
                value=row["model_candidate_phone"],
                inventory=inventory,
                required=True,
            )
            recommended = validate_phone(
                token=token,
                label="recommended",
                value=row["recommended_phone"],
                inventory=inventory,
                required=True,
            )
            alternative = validate_phone(
                token=token,
                label="alternative",
                value=row["alternative_phone"],
                inventory=inventory,
                required=False,
            )
            custom = validate_phone(
                token=token,
                label="custom",
                value=row["researcher_custom_phone"],
                inventory=inventory,
                required=decision == "approve_custom",
            )
        except WorkbookIntegrityError as exc:
            errors.append(str(exc))
            continue

        selected = ""
        selected_source = ""
        if decision == "approve_recommended":
            selected = recommended
            selected_source = "recommended"
            if custom:
                errors.append(
                    f"{token}: custom phone must be blank for recommended approval"
                )
        elif decision == "approve_alternative":
            selected = alternative
            selected_source = "alternative"
            if not alternative:
                errors.append(
                    f"{token}: alternative approval has no alternative phone"
                )
            if custom:
                errors.append(
                    f"{token}: custom phone must be blank for alternative approval"
                )
        elif decision == "approve_custom":
            selected = custom
            selected_source = "custom"
        elif custom:
            errors.append(
                f"{token}: custom phone is present without custom approval"
            )

        requires_notes = (
            decision in {"approve_alternative", "approve_custom", "hold", "reject"}
            or (
                decision == "approve_recommended"
                and (
                    action != "accept_model_candidate"
                    or selected != candidate
                )
            )
        )
        if requires_notes and not notes:
            errors.append(
                f"{token}: this decision requires researcher notes"
            )

        provenance = ""
        if selected:
            provenance = (
                "researcher_workbook_same_frozen_candidate"
                if selected == candidate
                else "researcher_workbook_manual_same_inventory"
            )
        normalized = {
            "review_order": clean(row["review_order"]),
            "category": clean(row["category"]),
            "token": token,
            "model_input": clean(row["model_input"]),
            "model_candidate_phone": candidate,
            "recommendation_action": action,
            "researcher_decision": decision,
            "approved_pron_phones_mfa": selected,
            "approved_phone_source": selected_source,
            "approved_phone_provenance": provenance,
            "researcher_notes": notes,
            "source_handling": clean(row["source_handling"]),
            "source_url": clean(row["source_url"]),
            "reason": clean(row["reason"]),
            "example_utt_id": clean(row["example_utt_id"]),
            "review_wav": clean(row["review_wav"]),
        }
        decisions.append(normalized)

        if token in CORRECTION_SPECS and selected:
            spec = CORRECTION_SPECS[token]
            corrections.append(
                {
                    "review_order": clean(row["review_order"]),
                    "token": token,
                    **spec,
                    "approved_pron_phones_mfa": selected,
                    "researcher_decision": decision,
                    "researcher_notes": notes,
                    "example_utt_id": clean(row["example_utt_id"]),
                }
            )

    affirmative = sum(counts[value] for value in AFFIRMATIVE_DECISIONS)
    pending = counts["pending"]
    held = counts["hold"] + counts["reject"]
    ready = (
        not errors
        and len(decisions) == 27
        and affirmative == 27
        and pending == 0
        and held == 0
        and len(corrections) == 2
    )
    if errors:
        status = "invalid_decisions"
    elif pending:
        status = "incomplete_pending"
    elif held:
        status = "researcher_hold"
    elif ready:
        status = "ready_for_apply"
    else:
        status = "incomplete"
    return decisions, corrections, {
        "status": status,
        "ready_for_apply": ready,
        "decision_counts": dict(sorted(counts.items())),
        "affirmative_decisions": affirmative,
        "pending_decisions": pending,
        "hold_or_reject_decisions": held,
        "errors": errors,
    }


def validate_filled_workbook(
    *,
    template_path: Path,
    filled_path: Path,
    template_manifest: dict[str, Any],
    inventory: set[str],
) -> dict[str, Any]:
    if template_path.resolve() == filled_path.resolve():
        raise WorkbookIntegrityError(
            "filled workbook must be a Save As copy, not the clean template"
        )
    if template_manifest.get("schema_version") != WORKBOOK_SCHEMA_VERSION:
        raise WorkbookIntegrityError("template manifest schema mismatch")
    if template_manifest.get("status") != "success":
        raise WorkbookIntegrityError("template manifest is not successful")
    if not fingerprint_matches(
        template_path, template_manifest.get("output", {})
    ):
        raise WorkbookIntegrityError(
            "clean template fingerprint differs from its manifest"
        )

    template = load_workbook(
        template_path, data_only=False, read_only=False
    )
    filled = load_workbook(
        filled_path, data_only=False, read_only=False
    )
    try:
        immutable_cells = compare_workbook_contract(template, filled)
        rows = review_rows_from_sheet(filled[REVIEW_SHEET])
        decisions, corrections, decision_report = normalize_decisions(
            rows, inventory
        )
    finally:
        template.close()
        filled.close()
    return {
        "immutable_contract": {
            "status": "passed",
            "cells_compared": immutable_cells,
            "editable_cells_allowed": len(EDITABLE_COORDINATES),
            "editable_columns": sorted(EDITABLE_COLUMNS),
        },
        "decisions": decisions,
        "corrections": corrections,
        "decision_report": decision_report,
    }


def load_frozen_inventory(
    template_manifest: dict[str, Any]
) -> tuple[set[str], dict[str, Any], Path]:
    model_record = template_manifest.get("inputs", {}).get(
        "model_bundle", {}
    )
    model_path = Path(clean(model_record.get("path")))
    if not model_path.is_file() or not fingerprint_matches(
        model_path, model_record
    ):
        raise WorkbookIntegrityError(
            "frozen model bundle differs from workbook manifest"
        )
    model_bundle = json.loads(model_path.read_text(encoding="utf-8"))
    if model_bundle.get("status") != "success":
        raise WorkbookIntegrityError("frozen model bundle is not successful")
    acoustic_record = model_bundle.get("outputs", {}).get(
        "acoustic_model", {}
    )
    acoustic_path = Path(clean(acoustic_record.get("path")))
    if not acoustic_path.is_file() or not fingerprint_matches(
        acoustic_path, acoustic_record
    ):
        raise WorkbookIntegrityError(
            "frozen acoustic model differs from model bundle"
        )
    loader_inventory = acoustic_phone_inventory(acoustic_path)
    inventory = lexical_phone_inventory(loader_inventory)
    contract = phone_inventory_contract(inventory)
    expected = model_bundle.get("contract", {})
    if (
        contract["count"] != expected.get("phone_count")
        or contract["sorted_phone_sha256"]
        != expected.get("phone_sorted_sha256")
    ):
        raise WorkbookIntegrityError(
            "acoustic phone inventory contract mismatch"
        )
    return inventory, contract, model_path


def write_csv(path: Path, fields: tuple[str, ...], rows: list[dict]) -> None:
    with atomic_text_writer(
        path, encoding="utf-8-sig", newline=""
    ) as (stream, _):
        writer = csv.DictWriter(
            stream, fieldnames=fields, lineterminator="\n"
        )
        writer.writeheader()
        writer.writerows(rows)


def output_paths(manifest_path: Path) -> tuple[Path, Path]:
    suffix = ".manifest.json"
    if not manifest_path.name.endswith(suffix):
        raise ValueError("output manifest name must end with .manifest.json")
    stem = manifest_path.name[: -len(suffix)]
    return (
        manifest_path.with_name(f"{stem}.decisions.csv"),
        manifest_path.with_name(f"{stem}.corrections.csv"),
    )


def build_validation_outputs(
    *,
    template_path: Path,
    filled_path: Path,
    template_manifest_path: Path,
    output_manifest_path: Path,
) -> dict[str, Any]:
    decisions_path, corrections_path = output_paths(output_manifest_path)
    for path in (
        output_manifest_path,
        decisions_path,
        corrections_path,
    ):
        if path.exists():
            raise FileExistsError(f"validation output already exists: {path}")
    template_manifest = json.loads(
        template_manifest_path.read_text(encoding="utf-8")
    )
    inventory, inventory_contract, model_bundle_path = (
        load_frozen_inventory(template_manifest)
    )
    result = validate_filled_workbook(
        template_path=template_path,
        filled_path=filled_path,
        template_manifest=template_manifest,
        inventory=inventory,
    )
    report = result["decision_report"]
    outputs: dict[str, Any] = {}
    if report["ready_for_apply"]:
        write_csv(decisions_path, DECISION_FIELDS, result["decisions"])
        write_csv(
            corrections_path,
            CORRECTION_FIELDS,
            result["corrections"],
        )
        outputs = {
            "normalized_decisions": file_fingerprint(
                decisions_path, with_sha256=True
            ),
            "correction_registry": file_fingerprint(
                corrections_path, with_sha256=True
            ),
        }
    manifest = {
        "schema_version": SCHEMA_VERSION,
        "status": report["status"],
        "kind": "common_pron_r2_researcher_decision_validation",
        "ready_for_apply": report["ready_for_apply"],
        "inputs": {
            "clean_template": file_fingerprint(
                template_path, with_sha256=True
            ),
            "filled_workbook": file_fingerprint(
                filled_path, with_sha256=True
            ),
            "template_manifest": file_fingerprint(
                template_manifest_path, with_sha256=True
            ),
            "model_bundle": file_fingerprint(
                model_bundle_path, with_sha256=True
            ),
        },
        "phone_inventory_contract": inventory_contract,
        "immutable_contract": result["immutable_contract"],
        "decision_report": report,
        "counts": {
            "normalized_decisions": (
                len(result["decisions"])
                if report["ready_for_apply"]
                else 0
            ),
            "correction_registry_rows": (
                len(result["corrections"])
                if report["ready_for_apply"]
                else 0
            ),
        },
        "outputs": outputs,
        "implementation": {
            "validator_script": file_fingerprint(
                Path(__file__).resolve(), with_sha256=True
            )
        },
        "runtime": runtime_snapshot(PROJECT_ROOT),
    }
    atomic_write_json(output_manifest_path, manifest)
    return manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="공통발음 r2 연구자 작성 workbook 검증"
    )
    parser.add_argument("--template", type=Path, required=True)
    parser.add_argument("--filled", type=Path, required=True)
    parser.add_argument(
        "--template-manifest", type=Path, required=True
    )
    parser.add_argument("--output-manifest", type=Path, required=True)
    return parser.parse_args()


def main() -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    args = parse_args()
    manifest = build_validation_outputs(
        template_path=args.template.resolve(),
        filled_path=args.filled.resolve(),
        template_manifest_path=args.template_manifest.resolve(),
        output_manifest_path=args.output_manifest.resolve(),
    )
    print(
        json.dumps(
            {
                "status": manifest["status"],
                "ready_for_apply": manifest["ready_for_apply"],
                **manifest["decision_report"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0 if manifest["ready_for_apply"] else 3


if __name__ == "__main__":
    raise SystemExit(main())
