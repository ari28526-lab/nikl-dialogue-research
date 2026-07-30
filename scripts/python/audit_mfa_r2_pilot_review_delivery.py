"""최종 MFA r2 연구자 검토 폴더와 REVIEW.xlsx를 재현 가능하게 감사한다."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from package_mfa_r2_pilot_review import BUNDLE_SCHEMA_VERSION
from pipeline_common import (
    atomic_write_json,
    file_fingerprint,
    now_iso,
)
from recover_mfa_r2_pilot_review_bundle import validate_bundle

WORKBOOK_NAME = "REVIEW.xlsx"
WORKBOOK_MANIFEST_NAME = "REVIEW_XLSX_TEMPLATE_MANIFEST.json"
LINK_FIELDS = ("wav_file", "textgrid_file", "lab_file", "csv_file")


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(encoding="utf-8-sig", newline="") as stream:
        reader = csv.DictReader(stream)
        return list(reader.fieldnames or []), list(reader)


def audit_delivery(
    *,
    review_root: Path,
    report_path: Path,
) -> dict[str, Any]:
    review_root = review_root.resolve()
    report_path = report_path.resolve()
    if report_path.exists():
        raise FileExistsError(f"기존 최종 감사 보고서를 덮어쓰지 않음: {report_path}")

    _bundle, bundle_summary = validate_bundle(
        review_root,
        accepted_schemas={BUNDLE_SCHEMA_VERSION},
        allowed_extra_files={
            WORKBOOK_NAME,
            WORKBOOK_MANIFEST_NAME,
        },
    )
    review_csv = review_root / "REVIEW.csv"
    workbook_path = review_root / WORKBOOK_NAME
    workbook_manifest_path = review_root / WORKBOOK_MANIFEST_NAME
    headers, csv_rows = read_csv(review_csv)
    if len(csv_rows) != 60 or len(headers) != 15:
        raise RuntimeError(
            f"REVIEW.csv 구조 불일치: rows={len(csv_rows)} cols={len(headers)}"
        )
    workbook_manifest = json.loads(
        workbook_manifest_path.read_text(encoding="utf-8-sig")
    )
    if not isinstance(workbook_manifest, dict):
        raise RuntimeError("XLSX template manifest가 object가 아님")
    csv_fp = file_fingerprint(review_csv, with_sha256=True)
    bundle_fp = file_fingerprint(
        review_root / "BUNDLE_MANIFEST.json",
        with_sha256=True,
    )
    xlsx_fp = file_fingerprint(workbook_path, with_sha256=True)
    if (
        workbook_manifest.get("schema_version")
        != "mfa_r2_review_xlsx_template.v1"
        or workbook_manifest.get("status") != "success"
        or workbook_manifest.get("rows") != 60
        or workbook_manifest.get("columns") != 15
        or workbook_manifest.get("sheets") != ["검토입력", "안내"]
        or workbook_manifest.get("input_review_csv", {}).get("sha256")
        != csv_fp["sha256"]
        or workbook_manifest.get("source_bundle_manifest", {}).get(
            "sha256"
        )
        != bundle_fp["sha256"]
        or workbook_manifest.get("output_review_xlsx", {}).get("sha256")
        != xlsx_fp["sha256"]
    ):
        raise RuntimeError("XLSX 생성 manifest와 최종 파일 fingerprint 불일치")

    workbook = load_workbook(
        workbook_path,
        data_only=False,
        read_only=False,
    )
    try:
        if workbook.sheetnames != ["검토입력", "안내"]:
            raise RuntimeError(f"workbook 시트 불일치: {workbook.sheetnames}")
        sheet = workbook["검토입력"]
        workbook_headers = [
            str(sheet.cell(1, column).value or "")
            for column in range(1, sheet.max_column + 1)
        ]
        if (
            workbook_headers != headers
            or sheet.max_row != 61
            or sheet.max_column != 15
        ):
            raise RuntimeError("workbook 행·열·헤더가 REVIEW.csv와 다름")
        index = {
            header: position + 1
            for position, header in enumerate(headers)
        }
        hyperlink_count = 0
        for row_number, csv_row in enumerate(csv_rows, start=2):
            for field in LINK_FIELDS:
                cell = sheet.cell(row_number, index[field])
                expected = csv_row[field]
                if (
                    cell.value != expected
                    or cell.hyperlink is None
                    or cell.hyperlink.target != expected
                    or not (review_root / expected).is_file()
                ):
                    raise RuntimeError(
                        f"{row_number}행 {field} 파일/링크 불일치"
                    )
                hyperlink_count += 1
        validation_count = len(
            sheet.data_validations.dataValidation
        )
        if validation_count != 2:
            raise RuntimeError(
                f"workbook dropdown validation 수 불일치: {validation_count}"
            )
        if "MfaR2InfrastructureReview" not in sheet.tables:
            raise RuntimeError("workbook 검토 표(table) 누락")
    finally:
        workbook.close()

    report: dict[str, Any] = {
        "schema_version": "mfa_r2_pilot_review_delivery_audit.v1",
        "status": "passed",
        "audited_at": now_iso(),
        "review_root": str(review_root),
        "flat_directory_files": len(list(review_root.iterdir())),
        "bundle": bundle_summary,
        "review_csv": csv_fp,
        "review_xlsx": xlsx_fp,
        "review_xlsx_manifest": file_fingerprint(
            workbook_manifest_path, with_sha256=True
        ),
        "workbook": {
            "sheets": ["검토입력", "안내"],
            "rows": len(csv_rows),
            "columns": len(headers),
            "hyperlinks_verified": hyperlink_count,
            "dropdown_validations": validation_count,
            "table": "MfaR2InfrastructureReview",
        },
        "review_scope": "infrastructure_acceptance_only",
        "realization_judgment_performed": False,
        "researcher_review_status": "pending",
        "source_or_d_results_modified": False,
    }
    atomic_write_json(report_path, report)
    return report


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--review-root", type=Path, required=True)
    parser.add_argument("--report", type=Path, required=True)
    args = parser.parse_args()
    report = audit_delivery(
        review_root=args.review_root,
        report_path=args.report,
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
