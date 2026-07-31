"""새 형태소 검색·TextGrid 스키마의 12발화 평면 검토 묶음을 만든다."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import shutil
import sys
import time
import wave
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Iterable, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from research_textgrid import (
    validate_research_textgrid,
    write_research_textgrid,
)
from retrofit_textgrid_2020_2024 import parse_mfa_textgrid

csv.field_size_limit(10_000_000)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

EXPLICIT_INCLUDE = {"SDRW2000000510.1.1.98"}
DETAIL_TABLES = (
    "morph_tokens",
    "morph_units",
    "morph_boundaries",
    "orth_components",
)
REVIEW_EDGE_PADDING_SECONDS = 0.05


def read_csv(path: Path) -> list[dict[str, str]]:
    with open(path, encoding="utf-8-sig", newline="") as stream:
        return list(csv.DictReader(stream))


def write_csv(
    path: Path, rows: list[Mapping[str, object]], fieldnames: list[str]
) -> None:
    with open(path, "w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(
            stream,
            fieldnames=fieldnames,
            extrasaction="ignore",
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(rows)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def copy_checked(source: Path, destination: Path) -> None:
    if not source.is_file():
        raise FileNotFoundError(source)
    shutil.copy2(source, destination)
    if (
        source.stat().st_size != destination.stat().st_size
        or sha256_file(source) != sha256_file(destination)
    ):
        raise RuntimeError(f"복사 검증 실패: {source} -> {destination}")


def promote_directory_with_retry(partial: Path, output_root: Path) -> None:
    delay = 0.25
    for attempt in range(12):
        try:
            os.replace(partial, output_root)
            return
        except PermissionError:
            if attempt == 11:
                raise
            time.sleep(delay)
            delay = min(delay * 1.7, 5.0)


def recover_completed_partial(
    partial: Path, output_root: Path
) -> dict[str, object] | None:
    manifest_path = partial / "BUNDLE_MANIFEST.json"
    if not manifest_path.is_file():
        return None
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    if manifest.get("status") != "success":
        return None
    for record in manifest.get("files", []):
        path = partial / str(record["relative_path"])
        if (
            not path.is_file()
            or path.stat().st_size != int(record["bytes"])
            or sha256_file(path) != record["sha256"]
        ):
            raise RuntimeError(f"완료 partial 검증 실패: {path}")
    validate_workbook(
        partial / "REVIEW.xlsx",
        int(manifest["counts"]["utterances"]),
    )
    promote_directory_with_retry(partial, output_root)
    return manifest


def score_rows(
    masters: list[dict[str, str]],
    units: list[dict[str, str]],
) -> dict[str, tuple[int, list[str]]]:
    units_by_id: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in units:
        units_by_id[row["utt_id"]].append(row)
    result: dict[str, tuple[int, list[str]]] = {}
    for row in masters:
        score = 0
        reasons: list[str] = []
        if row.get("align_warn"):
            score += 8
            reasons.append("form-tagged 어절수 경고")
        if row.get("legacy_tagged_roman_equal_v2") == "False":
            score += 3
            reasons.append("v1→v2 표시 변화")
        row_units = units_by_id[row["utt_id"]]
        if any(unit["unit_type"] == "jamo" for unit in row_units):
            score += 7
            reasons.append("독립 자모 형태소")
        if any(unit["unit_type"] == "literal" for unit in row_units):
            score += 4
            reasons.append("문장부호/literal")
        if any(
            unit.get("coda_components_json", "").count(",") >= 1
            for unit in row_units
        ):
            score += 5
            reasons.append("겹받침 구성 성분")
        if any(unit.get("onset_zero") == "True" for unit in row_units):
            score += 1
            reasons.append("무음 초성 ㅇ")
        score += min(int(row.get("morph_count_structured") or 0), 12) // 4
        if row["utt_id"] in EXPLICIT_INCLUDE:
            score += 100
            reasons.insert(0, "기존 연구자 검토 사례")
        result[row["utt_id"]] = (
            score,
            reasons or ["일반 대조 사례"],
        )
    return result


def select_twelve(
    masters: list[dict[str, str]],
    units: list[dict[str, str]],
) -> list[dict[str, str]]:
    scored = score_rows(masters, units)
    by_year: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in masters:
        by_year[row["year"]].append(row)
    selected: list[dict[str, str]] = []
    for year in map(str, range(2020, 2026)):
        candidates = sorted(
            by_year[year],
            key=lambda row: (
                -scored[row["utt_id"]][0],
                row["utt_id"],
            ),
        )
        if len(candidates) < 2:
            raise RuntimeError(f"{year}: 검토 후보가 2개 미만")
        chosen = candidates[:2]
        for row in chosen:
            copied = dict(row)
            copied["_selection_score"] = str(scored[row["utt_id"]][0])
            copied["_selection_reason"] = ", ".join(
                scored[row["utt_id"]][1]
            )
            selected.append(copied)
    if len(selected) != 12 or Counter(row["year"] for row in selected) != {
        str(year): 2 for year in range(2020, 2026)
    }:
        raise RuntimeError("12발화·연도별 2개 선택 계약 실패")
    return selected


def wav_duration(path: Path) -> float:
    with wave.open(str(path), "rb") as stream:
        return stream.getnframes() / stream.getframerate()


def blank_review_padding(
    intervals: list[tuple[float, float, str]],
    *,
    duration: float,
) -> list[tuple[float, float, str]]:
    """검토용으로 추가된 좌우 무음 구간의 label을 명시적으로 비운다.

    padding 안쪽의 시간·라벨은 바꾸지 않는다. 경계를 가로지른 interval은
    0.05초 경계에서 나누며, 새로 추가된 무음 바깥 조각만 빈 label로 둔다.
    """

    left = REVIEW_EDGE_PADDING_SECONDS
    right = float(duration) - REVIEW_EDGE_PADDING_SECONDS
    if right <= left:
        raise RuntimeError(f"검토 WAV가 padding보다 짧음: {duration}")
    result: list[tuple[float, float, str]] = []
    for begin, end, label in intervals:
        points = [float(begin)]
        points.extend(
            cut
            for cut in (left, right)
            if float(begin) + 1e-9 < cut < float(end) - 1e-9
        )
        points.append(float(end))
        for piece_begin, piece_end in zip(points, points[1:]):
            piece_label = (
                ""
                if piece_end <= left + 1e-9
                or piece_begin >= right - 1e-9
                else label
            )
            result.append((piece_begin, piece_end, piece_label))
    return result


def _hyperlink(cell, filename: str) -> None:
    cell.value = filename
    cell.hyperlink = filename
    cell.style = "Hyperlink"


def create_workbook(
    path: Path,
    review_rows: list[dict[str, object]],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "검토"
    sheet.sheet_view.showGridLines = False
    headers = [
        "검토순서",
        "연도",
        "utt_id",
        "발화",
        "선정이유",
        "WAV_LAB_일치",
        "words_phones_정상",
        "양끝_경계_보임",
        "MORPH_읽기_쉬움",
        "MORPH_R_검색_가능",
        "결정",
        "메모",
        "WAV",
        "TextGrid",
        "LAB",
        "CSV",
    ]
    sheet.append(headers)
    for review in review_rows:
        row_index = sheet.max_row + 1
        values = [
            review["review_order"],
            review["year"],
            review["utt_id"],
            review["form"],
            review["selection_reason"],
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            None,
            None,
            None,
            None,
        ]
        sheet.append(values)
        _hyperlink(sheet.cell(row_index, 13), str(review["wav"]))
        _hyperlink(sheet.cell(row_index, 14), str(review["textgrid"]))
        _hyperlink(sheet.cell(row_index, 15), str(review["lab"]))
        _hyperlink(sheet.cell(row_index, 16), str(review["csv"]))
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 30
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:P{sheet.max_row}"
    widths = {
        "A": 10,
        "B": 8,
        "C": 32,
        "D": 38,
        "E": 32,
        "F": 15,
        "G": 18,
        "H": 16,
        "I": 18,
        "J": 20,
        "K": 13,
        "L": 45,
        "M": 25,
        "N": 25,
        "O": 25,
        "P": 25,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.row_dimensions[row[0].row].height = 60
    check_validation = DataValidation(
        type="list",
        formula1='"정상,문제,보류"',
        allow_blank=True,
    )
    decision_validation = DataValidation(
        type="list",
        formula1='"승인,수정 필요,보류"',
        allow_blank=True,
    )
    sheet.add_data_validation(check_validation)
    check_validation.add(f"F2:J{sheet.max_row}")
    sheet.add_data_validation(decision_validation)
    decision_validation.add(f"K2:K{sheet.max_row}")
    red_fill = PatternFill("solid", fgColor="F4CCCC")
    sheet.conditional_formatting.add(
        f"F2:K{sheet.max_row}",
        FormulaRule(formula=['F2="문제"'], fill=red_fill),
    )

    guide = workbook.create_sheet("열_안내")
    guide.sheet_view.showGridLines = False
    guide.append(["열", "무엇을 확인하나", "판단 기준"])
    guide_rows = [
        (
            "WAV_LAB_일치",
            "들리는 발화와 LAB 문장이 같은가",
            "같으면 정상. 구체적 실제 발음 판정은 하지 않음",
        ),
        (
            "words_phones_정상",
            "MFA words/phones_mfa 시간·라벨이 이전과 같이 보이는가",
            "큰 누락·역전·겹침이 없으면 정상",
        ),
        (
            "양끝_경계_보임",
            "모든 tier 왼쪽과 오른쪽에 0.05초 빈 구간이 보이는가",
            "WAV도 같은 padding이 있으므로 시간 대응이 유지되어야 함",
        ),
        (
            "MORPH_읽기_쉬움",
            "[MORPH]의 형태소 표면형/POS가 발화 수준에서 읽히는가",
            "형태소를 시간분할한 tier가 아님",
        ),
        (
            "MORPH_R_검색_가능",
            "[MORPH_R]의 공백/_/+/|/POS와 ⟨literal⟩이 이해되는가",
            "Praat 검색과 후보 확인에 쓸 수 있으면 정상",
        ),
        (
            "결정",
            "이 새 CSV·TextGrid 표시 계약을 전수에 적용할 수 있는가",
            "승인/수정 필요/보류 중 선택",
        ),
        (
            "메모",
            "문제가 있으면 반복 입력하지 말고 첫 사례에 전역 문제로 기록",
            "예: 모든 MORPH_R에서 독립 자모가 불편함",
        ),
    ]
    for row in guide_rows:
        guide.append(row)
    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    guide.row_dimensions[1].height = 28
    guide.freeze_panes = "A2"
    guide.column_dimensions["A"].width = 25
    guide.column_dimensions["B"].width = 55
    guide.column_dimensions["C"].width = 65
    for row in guide.iter_rows():
        for cell in row:
            if cell.row > 1:
                cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if row[0].row > 1:
            guide.row_dimensions[row[0].row].height = 58

    change = workbook.create_sheet("변경_요약")
    change.sheet_view.showGridLines = False
    change.append(["항목", "새 계약"])
    changes = [
        ("TextGrid tier", "words / phones_mfa / utterance / utterance_search"),
        ("형태소 시간분할", "삭제. [MORPH] 발화 수준 한 줄로 제공"),
        ("철자 로마자", "[MORPH_R] tagged_roman_v2"),
        ("독립 자모", "ㄴ→n, ㄹ→l처럼 별도 unit으로 보존"),
        ("문장부호·숫자", "⟨.⟩, ⟨?⟩, ⟨1⟩처럼 literal 명시"),
        ("겹받침", "음절 slot과 ㄹ+ㄱ 같은 구성 성분을 CSV에서 모두 보존"),
        ("MFA phone", "실제 실현 판정이 아닌 자동 정렬 보조값"),
    ]
    for row in changes:
        change.append(row)
    for cell in change[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    change.row_dimensions[1].height = 28
    change.column_dimensions["A"].width = 24
    change.column_dimensions["B"].width = 85
    for row in change.iter_rows():
        for cell in row:
            if cell.row > 1:
                cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if row[0].row > 1:
            change.row_dimensions[row[0].row].height = 42
    workbook.save(path)


def validate_workbook(path: Path, expected_rows: int) -> None:
    workbook = load_workbook(path, data_only=False)
    if workbook.sheetnames != ["검토", "열_안내", "변경_요약"]:
        raise RuntimeError(f"검토 workbook sheet 불일치: {workbook.sheetnames}")
    sheet = workbook["검토"]
    if sheet.max_row != expected_rows + 1:
        raise RuntimeError(
            f"검토 workbook 행 수 불일치: {sheet.max_row - 1}"
        )
    for row in range(2, sheet.max_row + 1):
        for column in range(13, 17):
            if not sheet.cell(row, column).hyperlink:
                raise RuntimeError(
                    f"검토 workbook hyperlink 누락: row={row} col={column}"
                )


def package_review(
    *,
    legacy_review_root: Path,
    morph_root: Path,
    output_root: Path,
) -> dict[str, object]:
    legacy_review_root = legacy_review_root.resolve()
    morph_root = morph_root.resolve()
    output_root = output_root.resolve()
    partial = output_root.with_name(output_root.name + ".partial")
    if output_root.exists():
        raise FileExistsError(f"기존 검토 묶음 덮어쓰기 금지: {output_root}")
    if partial.exists():
        recovered = recover_completed_partial(partial, output_root)
        if recovered is not None:
            return recovered
        raise FileExistsError(f"미완료 검토 묶음 조사 필요: {partial}")
    partial.mkdir(parents=True)

    masters = read_csv(morph_root / "utterance_master_v2.csv")
    details = {
        name: read_csv(morph_root / f"{name}.csv")
        for name in DETAIL_TABLES
    }
    selected = select_twelve(masters, details["morph_units"])
    selected_ids = {row["utt_id"] for row in selected}
    master_fields = list(masters[0])
    review_rows: list[dict[str, object]] = []
    manifest_files: list[dict[str, object]] = []
    boundary_counts = Counter()

    for review_order, row in enumerate(selected, 1):
        year = row["year"]
        utt_id = row["utt_id"]
        prefix = f"{year}__{utt_id}"
        source_wav = legacy_review_root / f"{prefix}.wav"
        source_lab = legacy_review_root / f"{prefix}.lab"
        source_tg = legacy_review_root / f"{prefix}.TextGrid"
        destination_wav = partial / f"{prefix}.wav"
        destination_lab = partial / f"{prefix}.lab"
        destination_tg = partial / f"{prefix}.TextGrid"
        destination_csv = partial / f"{prefix}.csv"
        copy_checked(source_wav, destination_wav)
        copy_checked(source_lab, destination_lab)
        duration, tiers = parse_mfa_textgrid(source_tg)
        words = tiers.get("words", [])
        phones = tiers.get("phones_mfa", tiers.get("phones", []))
        if not words or not phones:
            raise RuntimeError(f"{source_tg}: words/phones 누락")
        words = blank_review_padding(words, duration=float(duration))
        phones = blank_review_padding(phones, duration=float(duration))
        validation = write_research_textgrid(
            destination_tg,
            duration=float(duration),
            words=words,
            phones=phones,
            search_row=row,
        )
        if not validation["valid"]:
            raise RuntimeError(f"{destination_tg}: 새 TextGrid 검증 실패")
        boundary_counts["left"] += bool(validation["left_empty_boundary"])
        boundary_counts["right"] += bool(validation["right_empty_boundary"])
        if (
            abs(wav_duration(destination_wav) - float(duration)) > 0.001
        ):
            raise RuntimeError(f"{prefix}: padded WAV/TextGrid duration 불일치")
        write_csv(destination_csv, [row], master_fields)
        for path in (
            destination_wav,
            destination_tg,
            destination_lab,
            destination_csv,
        ):
            manifest_files.append(
                {
                    "relative_path": path.name,
                    "bytes": path.stat().st_size,
                    "sha256": sha256_file(path),
                }
            )
        review_rows.append(
            {
                "review_order": review_order,
                "year": year,
                "utt_id": utt_id,
                "form": row["form"],
                "selection_reason": row["_selection_reason"],
                "wav": destination_wav.name,
                "textgrid": destination_tg.name,
                "lab": destination_lab.name,
                "csv": destination_csv.name,
            }
        )

    for name, rows in details.items():
        filtered = [row for row in rows if row["utt_id"] in selected_ids]
        destination = partial / f"{name.upper()}.csv"
        write_csv(destination, filtered, list(rows[0]))
        manifest_files.append(
            {
                "relative_path": destination.name,
                "bytes": destination.stat().st_size,
                "sha256": sha256_file(destination),
            }
        )
    review_csv = partial / "REVIEW.csv"
    review_fields = [
        "review_order",
        "year",
        "utt_id",
        "form",
        "selection_reason",
        "wav",
        "textgrid",
        "lab",
        "csv",
    ]
    write_csv(review_csv, review_rows, review_fields)
    workbook_path = partial / "REVIEW.xlsx"
    create_workbook(workbook_path, review_rows)
    validate_workbook(workbook_path, 12)
    readme = partial / "README.md"
    readme.write_text(
        """# 새 CSV·TextGrid 검색 스키마 12발화 검토

1. `REVIEW.xlsx`를 연다.
2. 1번부터 WAV를 재생하고 LAB이 같은 발화인지 확인한다.
3. TextGrid를 열어 `words / phones_mfa / utterance / utterance_search`
   네 tier와 양끝 빈 경계를 확인한다.
4. `utterance_search`의 `[MORPH]`와 `[MORPH_R]`가 읽고 검색하기
   편한지 확인한다.
5. 엑셀의 다섯 확인 열과 `결정`, 필요한 경우 `메모`만 입력한다.

같은 문제가 반복되면 첫 사례에만 “전역 문제”라고 적어도 된다.
이 검토에서는 구체적인 음운 현상의 실제 실현 여부를 판정하지 않는다.
`phones_mfa`는 자동 정렬 보조값이다.
""",
        encoding="utf-8",
    )
    for path in (review_csv, workbook_path, readme):
        manifest_files.append(
            {
                "relative_path": path.name,
                "bytes": path.stat().st_size,
                "sha256": sha256_file(path),
            }
        )
    if boundary_counts != {"left": 12, "right": 12}:
        raise RuntimeError(f"12발화 양끝 padding gate 실패: {boundary_counts}")
    manifest = {
        "schema_version": "mfa_research_schema_review_bundle.v1",
        "status": "success",
        "created_at": datetime.now().astimezone().isoformat(),
        "legacy_review_root": str(legacy_review_root),
        "morph_root": str(morph_root),
        "selection": [
            {
                "review_order": row["review_order"],
                "year": row["year"],
                "utt_id": row["utt_id"],
                "reason": row["selection_reason"],
            }
            for row in review_rows
        ],
        "counts": {
            "utterances": 12,
            "years": 6,
            "per_year": 2,
            "payload_files": 48,
            "left_empty_boundary": 12,
            "right_empty_boundary": 12,
        },
        "files": sorted(
            manifest_files, key=lambda record: str(record["relative_path"])
        ),
    }
    manifest_path = partial / "BUNDLE_MANIFEST.json"
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    promote_directory_with_retry(partial, output_root)
    return manifest


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--legacy-review-root", type=Path, required=True)
    parser.add_argument("--morph-root", type=Path, required=True)
    parser.add_argument("--output-root", type=Path, required=True)
    args = parser.parse_args()
    try:
        manifest = package_review(
            legacy_review_root=args.legacy_review_root,
            morph_root=args.morph_root,
            output_root=args.output_root,
        )
    except Exception as exc:
        print(f"[FAIL] {type(exc).__name__}: {exc}", file=sys.stderr)
        return 1
    print(json.dumps(manifest, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
