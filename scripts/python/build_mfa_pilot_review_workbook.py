#!/usr/bin/env python3
"""Build an auditable Excel workbook for manual review of the MFA pilot."""

from __future__ import annotations

import argparse
import csv
import json
import os
from datetime import datetime
from pathlib import Path, PureWindowsPath
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


REQUIRED_COLUMNS = (
    "year",
    "speaker_id",
    "dialogue_id",
    "dialogue_speaker_ids",
    "co_speaker_ids",
    "utt_id",
    "form",
    "original_form",
    "pron_reference_hangul",
    "pron_reference_source",
    "tier_warning",
    "wav_relpath",
    "lab_relpath",
    "textgrid_relpath",
    "csv_relpath",
)

PRIORITY_UTTERANCES = {
    "SDRW2300001955.1.1.164": "이전 검토에서 찾기 어려웠던 발화",
    "SDRW2300000891.1.1.331": "숫자 1/일 발음 참조·화자 정보 확인",
}

INPUT_HEADERS = (
    "번호",
    "우선 확인",
    "연도",
    "발화 ID",
    "화자 ID",
    "공동 참여자 ID",
    "정규형",
    "원문",
    "발음 참조",
    "WAV",
    "TextGrid",
    "CSV",
    "LAB",
    "음성-문장 일치",
    "words 정렬",
    "phones 정렬",
    "morphemes 정렬",
    "처음·끝 경계",
    "발화 정보 표시",
    "화자·참여자 ID",
    "종합 판정",
    "재검토",
    "비고",
    "검토자",
    "검토일",
    "검토 완료",
    "대화 ID",
    "전체 화자 ID",
    "발음 근거",
    "자동 tier 경고",
    "형태소 정렬 상태",
    "TextGrid 정보 스키마",
    "원본 행 번호",
)

LINK_COLUMN_SOURCE = {
    10: "wav_relpath",
    11: "textgrid_relpath",
    12: "csv_relpath",
    13: "lab_relpath",
}

THIN_GRAY = Side(style="thin", color="D9E1E8")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
FORMULA_FILL = PatternFill("solid", fgColor="E2F0D9")
PRIORITY_FILL = PatternFill("solid", fgColor="FFE699")
GOOD_FILL = PatternFill("solid", fgColor="E2F0D9")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
BAD_FILL = PatternFill("solid", fgColor="FCE4D6")
GRAY_FILL = PatternFill("solid", fgColor="E7E6E6")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--index", required=True, type=Path, help="INDEX_ALL.csv")
    parser.add_argument("--bundle-root", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Archive an existing workbook before replacing it.",
    )
    return parser.parse_args()


def read_index(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        headers = list(reader.fieldnames or ())
        rows = list(reader)
    missing = sorted(set(REQUIRED_COLUMNS) - set(headers))
    if missing:
        raise ValueError(f"필수 INDEX 열 누락: {missing}")
    if not rows:
        raise ValueError("INDEX에 발화가 없습니다.")
    utt_ids = [row["utt_id"] for row in rows]
    duplicates = sorted({utt_id for utt_id in utt_ids if utt_ids.count(utt_id) > 1})
    if duplicates:
        raise ValueError(f"발화 ID 중복: {duplicates[:10]}")
    return headers, rows


def safe_bundle_path(bundle_root: Path, relative: str) -> Path:
    windows_path = PureWindowsPath(relative)
    if windows_path.is_absolute() or ".." in windows_path.parts:
        raise ValueError(f"안전하지 않은 상대경로: {relative}")
    candidate = bundle_root.joinpath(*windows_path.parts).resolve()
    root = bundle_root.resolve()
    if root != candidate and root not in candidate.parents:
        raise ValueError(f"bundle 밖으로 벗어난 경로: {relative}")
    return candidate


def safe_workbook_link_path(
    workbook_parent: Path,
    bundle_root: Path,
    relative: str,
) -> Path:
    """Excel 상대 링크를 해석하되 최종 대상이 bundle 안인지 확인한다."""
    windows_path = PureWindowsPath(relative)
    if windows_path.is_absolute():
        raise ValueError(f"절대경로 링크 금지: {relative}")
    candidate = workbook_parent.joinpath(*windows_path.parts).resolve()
    root = bundle_root.resolve()
    if root != candidate and root not in candidate.parents:
        raise ValueError(f"링크가 bundle 밖으로 벗어남: {relative}")
    return candidate


def validate_bundle_files(bundle_root: Path, rows: Iterable[dict[str, str]]) -> None:
    missing: list[str] = []
    for row in rows:
        for field in ("wav_relpath", "lab_relpath", "textgrid_relpath", "csv_relpath"):
            relative = row[field]
            if not safe_bundle_path(bundle_root, relative).is_file():
                missing.append(f"{row['utt_id']}:{field}:{relative}")
    if missing:
        raise FileNotFoundError(f"bundle 파일 누락 {len(missing)}건: {missing[:10]}")


def add_table(sheet, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def style_header(sheet, cell_range: str) -> None:
    for row in sheet[cell_range]:
        for cell in row:
            cell.fill = HEADER_FILL
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=THIN_GRAY)


def build_workbook(
    source_headers: list[str],
    rows: list[dict[str, str]],
    index_path: Path,
    bundle_root: Path,
    output_path: Path | None = None,
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.title = "MFA 파일럿 연구자 검토표"
    wb.properties.subject = "2020-2025 연도별 층화 MFA 파일럿 수동 검토"
    wb.properties.creator = "Codex (openpyxl 우회, 사용자 승인)"
    wb.properties.description = (
        "INDEX_ALL.csv를 보존하면서 연구자가 음성·TextGrid·CSV를 검토하고 "
        "판정을 기록할 수 있도록 생성한 워크북"
    )
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    guide = wb.create_sheet("안내")
    review = wb.create_sheet("검토입력")
    summary = wb.create_sheet("진행요약")
    raw = wb.create_sheet("원본인덱스")
    lists = wb.create_sheet("선택값")

    build_guide_sheet(guide, rows, index_path, bundle_root)
    link_prefix = ""
    if output_path is not None:
        link_prefix = os.path.relpath(
            bundle_root.resolve(),
            start=output_path.resolve().parent,
        )
    build_review_sheet(review, rows, link_prefix=link_prefix)
    build_summary_sheet(summary, rows)
    build_raw_sheet(raw, source_headers, rows)
    build_lists_sheet(lists)
    lists.sheet_state = "hidden"
    wb.active = wb.sheetnames.index("검토입력")
    return wb


def build_guide_sheet(sheet, rows, index_path: Path, bundle_root: Path) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = "5B9BD5"
    values = [
        ["MFA 파일럿 연구자 검토표", "2020–2025년, 연도별 10개, 총 60개"],
        ["목적", "대량 MFA 전에 음성·전사·TextGrid·형태소·발음 참조·화자 연결을 사람이 확인하고 기록"],
        ["사용 순서", "검토입력 시트에서 WAV와 TextGrid를 열고 노란 입력 열을 드롭다운으로 판정"],
        ["파일 링크", "워크북과 bundle의 상대 위치를 바꾸지 않아야 상대경로 링크가 작동"],
        ["판정 기준", "양호/일치 = 그대로 사용 가능, 수정 필요/불일치 = 대량 작업 전에 원인 분류 필요"],
        ["판단 불가", "현재 자료만으로 판정할 수 없을 때 사용하며 비고에 이유 기록"],
        ["공동 참여자", "co_speaker_ids는 같은 대화 문서의 다른 참여자이며 직접 수신자(addressee) 표지가 아님"],
        ["발음 참조", "pron_reference는 철자·규칙 기반 검색 보조 정보이며 사전 발음 또는 실제 음성 실현 판정값이 아님"],
        ["원본 보존", "원본인덱스 시트는 INDEX_ALL.csv의 값이며 검토 결과와 분리"],
        ["자동 집계", "진행요약은 검토입력 시트의 판정값을 수식으로 집계하며 Excel에서 열 때 재계산"],
        ["우선 사례 1", "SDRW2300001955.1.1.164 — 파일 찾기와 음성·전사 일치 확인"],
        ["우선 사례 2", "SDRW2300000891.1.1.331 — 1/일 발음 참조 및 화자·공동 참여자 확인"],
        ["원본 INDEX", str(index_path)],
        ["bundle root", str(bundle_root)],
        ["생성 시각", datetime.now().astimezone().isoformat(timespec="seconds")],
        ["원본 발화 수", len(rows)],
    ]
    sheet.append(["항목", "내용"])
    for row in values:
        sheet.append(row)
    style_header(sheet, f"A1:B1")
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 105
    sheet.freeze_panes = "A2"
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
        row[0].font = Font(bold=True, color="1F4E78")
        row[0].alignment = Alignment(vertical="top")
        row[1].alignment = Alignment(vertical="top", wrap_text=True)
    add_table(sheet, "GuideTable", f"A1:B{sheet.max_row}")


def build_review_sheet(
    sheet,
    rows: list[dict[str, str]],
    *,
    link_prefix: str = "",
) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = "70AD47"
    sheet.append(INPUT_HEADERS)
    for idx, row in enumerate(rows, start=1):
        priority = PRIORITY_UTTERANCES.get(row["utt_id"], "")
        sheet.append(
            [
                idx,
                priority,
                int(row["year"]),
                row["utt_id"],
                row["speaker_id"],
                row["co_speaker_ids"],
                row["form"],
                row["original_form"],
                row["pron_reference_hangul"],
                "WAV 열기",
                "TextGrid 열기",
                "CSV 열기",
                "LAB 열기",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                None,
                f'=IF(COUNTA(N{idx + 1}:U{idx + 1})=8,"완료","미완료")',
                row["dialogue_id"],
                row["dialogue_speaker_ids"],
                row["pron_reference_source"],
                row["tier_warning"],
                row.get("morph_analysis_align_status")
                or row.get("original_form_align_status", ""),
                row.get("utterance_info_schema")
                or row.get("pron_reference_align_status", ""),
                idx + 1,
            ]
        )
        excel_row = idx + 1
        for column_index, field in LINK_COLUMN_SOURCE.items():
            cell = sheet.cell(excel_row, column_index)
            target = (
                PureWindowsPath(link_prefix) / PureWindowsPath(row[field])
                if link_prefix
                else PureWindowsPath(row[field])
            )
            cell.hyperlink = str(target).replace("\\", "/")
            cell.style = "Hyperlink"

    max_row = len(rows) + 1
    style_header(sheet, f"A1:AG1")
    sheet.auto_filter.ref = f"A1:AG{max_row}"
    sheet.freeze_panes = "G2"
    add_table(sheet, "ReviewTable", f"A1:AG{max_row}")

    for row in sheet.iter_rows(min_row=2, max_row=max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column in range(2, 10) or cell.column == 23)
        if row[3].value in PRIORITY_UTTERANCES:
            for cell in row[:13]:
                cell.fill = PRIORITY_FILL

    for column in range(14, 26):
        for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=2, max_row=max_row):
            for item in cell:
                item.fill = INPUT_FILL
    for cell in sheet["Z"][1:]:
        cell.fill = FORMULA_FILL

    sheet.column_dimensions["A"].width = 7
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 9
    sheet.column_dimensions["D"].width = 31
    sheet.column_dimensions["E"].width = 15
    sheet.column_dimensions["F"].width = 20
    for column in ("G", "H", "I"):
        sheet.column_dimensions[column].width = 34
    for column in ("J", "K", "L", "M"):
        sheet.column_dimensions[column].width = 15
    for column in ("N", "O", "P", "Q", "R", "S", "T", "U", "V"):
        sheet.column_dimensions[column].width = 17
    sheet.column_dimensions["W"].width = 42
    sheet.column_dimensions["X"].width = 14
    sheet.column_dimensions["Y"].width = 13
    sheet.column_dimensions["Z"].width = 13
    for column in ("AA", "AB", "AC", "AD", "AE", "AF"):
        sheet.column_dimensions[column].width = 26
    sheet.column_dimensions["AG"].width = 12

    sheet.column_dimensions["AA"].hidden = True
    sheet.column_dimensions["AB"].hidden = True
    sheet.column_dimensions["AC"].hidden = True
    sheet.column_dimensions["AD"].hidden = True
    sheet.column_dimensions["AE"].hidden = True
    sheet.column_dimensions["AF"].hidden = True
    sheet.column_dimensions["AG"].hidden = True
    sheet.row_dimensions[1].height = 38
    for row_number in range(2, max_row + 1):
        sheet.row_dimensions[row_number].height = 43

    add_validations(sheet, max_row)
    add_conditional_formats(sheet, max_row)
    sheet["Y2"].number_format = "yyyy-mm-dd"
    for cell in sheet["Y"][1:]:
        cell.number_format = "yyyy-mm-dd"


def add_validations(sheet, max_row: int) -> None:
    audio = DataValidation(
        type="list",
        formula1='"일치,불일치,판단 불가"',
        allow_blank=True,
    )
    quality = DataValidation(
        type="list",
        formula1='"양호,수정 필요,판단 불가"',
        allow_blank=True,
    )
    overall = DataValidation(
        type="list",
        formula1='"통과,수정 필요,보류"',
        allow_blank=True,
    )
    revisit = DataValidation(
        type="list",
        formula1='"아니오,예"',
        allow_blank=True,
    )
    for validation in (audio, quality, overall, revisit):
        validation.error = "목록에서 값을 선택하세요."
        validation.errorTitle = "허용되지 않은 값"
        validation.prompt = "셀 오른쪽의 목록에서 판정값을 선택하세요."
        validation.promptTitle = "검토 판정"
        validation.showErrorMessage = True
        validation.showInputMessage = True
        sheet.add_data_validation(validation)
    audio.add(f"N2:N{max_row}")
    quality.add(f"O2:T{max_row}")
    overall.add(f"U2:U{max_row}")
    revisit.add(f"V2:V{max_row}")


def add_conditional_formats(sheet, max_row: int) -> None:
    sheet.conditional_formatting.add(
        f"N2:N{max_row}",
        CellIsRule(operator="equal", formula=['"일치"'], fill=GOOD_FILL),
    )
    sheet.conditional_formatting.add(
        f"N2:N{max_row}",
        CellIsRule(operator="equal", formula=['"불일치"'], fill=BAD_FILL),
    )
    sheet.conditional_formatting.add(
        f"O2:T{max_row}",
        CellIsRule(operator="equal", formula=['"양호"'], fill=GOOD_FILL),
    )
    sheet.conditional_formatting.add(
        f"O2:T{max_row}",
        CellIsRule(operator="equal", formula=['"수정 필요"'], fill=BAD_FILL),
    )
    sheet.conditional_formatting.add(
        f"N2:T{max_row}",
        CellIsRule(operator="equal", formula=['"판단 불가"'], fill=WARN_FILL),
    )
    sheet.conditional_formatting.add(
        f"U2:U{max_row}",
        CellIsRule(operator="equal", formula=['"통과"'], fill=GOOD_FILL),
    )
    sheet.conditional_formatting.add(
        f"U2:U{max_row}",
        CellIsRule(operator="equal", formula=['"수정 필요"'], fill=BAD_FILL),
    )
    sheet.conditional_formatting.add(
        f"U2:U{max_row}",
        CellIsRule(operator="equal", formula=['"보류"'], fill=WARN_FILL),
    )
    sheet.conditional_formatting.add(
        f"V2:V{max_row}",
        CellIsRule(operator="equal", formula=['"예"'], fill=BAD_FILL),
    )
    sheet.conditional_formatting.add(
        f"Z2:Z{max_row}",
        CellIsRule(operator="equal", formula=['"완료"'], fill=GOOD_FILL),
    )
    sheet.conditional_formatting.add(
        f"Z2:Z{max_row}",
        CellIsRule(operator="equal", formula=['"미완료"'], fill=GRAY_FILL),
    )
    sheet.conditional_formatting.add(
        f"A2:AG{max_row}",
        FormulaRule(formula=["$AD2<>\"\""], fill=WARN_FILL),
    )


def build_summary_sheet(sheet, rows: list[dict[str, str]]) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = "ED7D31"
    sheet.append(["전체 진행 요약", "값"])
    sheet.append(["총 발화", "=COUNTA('검토입력'!$D$2:$D$61)"])
    sheet.append(["검토 완료", '=COUNTIF(\'검토입력\'!$Z$2:$Z$61,"완료")'])
    sheet.append(["진행률", "=IFERROR(B3/B2,0)"])
    sheet.append(["통과", '=COUNTIF(\'검토입력\'!$U$2:$U$61,"통과")'])
    sheet.append(["수정 필요", '=COUNTIF(\'검토입력\'!$U$2:$U$61,"수정 필요")'])
    sheet.append(["보류", '=COUNTIF(\'검토입력\'!$U$2:$U$61,"보류")'])
    sheet.append(["재검토", '=COUNTIF(\'검토입력\'!$V$2:$V$61,"예")'])
    sheet.append(["상태", '=IF(B3=B2,"전체 검토 완료","검토 진행 중")'])
    sheet.append([])
    sheet.append(["연도", "대상", "완료", "진행률", "통과", "수정 필요", "보류"])
    years = sorted({int(row["year"]) for row in rows})
    for row_number, year in enumerate(years, start=12):
        sheet.append(
            [
                year,
                f'=COUNTIF(\'검토입력\'!$C$2:$C$61,A{row_number})',
                f'=COUNTIFS(\'검토입력\'!$C$2:$C$61,A{row_number},\'검토입력\'!$Z$2:$Z$61,"완료")',
                f"=IFERROR(C{row_number}/B{row_number},0)",
                f'=COUNTIFS(\'검토입력\'!$C$2:$C$61,A{row_number},\'검토입력\'!$U$2:$U$61,"통과")',
                f'=COUNTIFS(\'검토입력\'!$C$2:$C$61,A{row_number},\'검토입력\'!$U$2:$U$61,"수정 필요")',
                f'=COUNTIFS(\'검토입력\'!$C$2:$C$61,A{row_number},\'검토입력\'!$U$2:$U$61,"보류")',
            ]
        )
    style_header(sheet, "A1:B1")
    style_header(sheet, "A11:G11")
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 22
    for column in ("B", "C", "D", "E", "F", "G"):
        sheet.column_dimensions[column].width = 15
    sheet["B4"].number_format = "0%"
    for row_number in range(12, 12 + len(years)):
        sheet[f"D{row_number}"].number_format = "0%"
    add_table(sheet, "YearSummaryTable", f"A11:G{11 + len(years)}")


def build_raw_sheet(sheet, headers: list[str], rows: list[dict[str, str]]) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.tabColor = "A5A5A5"
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    style_header(sheet, f"A1:{sheet.cell(1, len(headers)).column_letter}1")
    sheet.freeze_panes = "A2"
    add_table(
        sheet,
        "RawIndexTable",
        f"A1:{sheet.cell(len(rows) + 1, len(headers)).coordinate}",
    )
    for column_number, header in enumerate(headers, start=1):
        width = 18
        if header in {"form", "original_form", "pron_reference_hangul"}:
            width = 35
        elif header.endswith("_relpath"):
            width = 42
        sheet.column_dimensions[sheet.cell(1, column_number).column_letter].width = width
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_lists_sheet(sheet) -> None:
    sheet.append(["범주", "값"])
    rows = [
        ["음성-문장", "일치"],
        ["음성-문장", "불일치"],
        ["음성-문장", "판단 불가"],
        ["품질", "양호"],
        ["품질", "수정 필요"],
        ["품질", "판단 불가"],
        ["종합", "통과"],
        ["종합", "수정 필요"],
        ["종합", "보류"],
        ["재검토", "아니오"],
        ["재검토", "예"],
    ]
    for row in rows:
        sheet.append(row)


def archive_existing(output: Path) -> Path | None:
    if not output.exists():
        return None
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archive_dir = output.parent / "_archive" / f"review_workbook_{stamp}"
    archive_dir.mkdir(parents=True, exist_ok=False)
    archived = archive_dir / output.name
    output.replace(archived)
    return archived


def verify_workbook(
    path: Path,
    expected_rows: int,
    bundle_root: Path,
) -> dict[str, object]:
    wb = load_workbook(path, data_only=False)
    expected_sheets = {"안내", "검토입력", "진행요약", "원본인덱스", "선택값"}
    if set(wb.sheetnames) != expected_sheets:
        raise ValueError(f"시트 구성 불일치: {wb.sheetnames}")
    review = wb["검토입력"]
    raw = wb["원본인덱스"]
    if review.max_row != expected_rows + 1:
        raise ValueError(f"검토행 수 불일치: {review.max_row - 1}")
    if raw.max_row != expected_rows + 1:
        raise ValueError(f"원본행 수 불일치: {raw.max_row - 1}")
    if len(review.data_validations.dataValidation) != 4:
        raise ValueError("드롭다운 검증 규칙 수 불일치")
    if not str(review["Z2"].value).startswith("=IF("):
        raise ValueError("검토 완료 수식 누락")
    if not str(wb["진행요약"]["B2"].value).startswith("="):
        raise ValueError("요약 수식 누락")
    hyperlink_count = 0
    for row_number in range(2, expected_rows + 2):
        for column_number in LINK_COLUMN_SOURCE:
            link = review.cell(row_number, column_number).hyperlink
            if link is None:
                raise ValueError(f"파일 링크 누락: {row_number},{column_number}")
            target = link.target
            if not safe_workbook_link_path(
                path.parent,
                bundle_root,
                target,
            ).is_file():
                raise ValueError(f"링크 대상 누락: {target}")
            hyperlink_count += 1
    priority_count = sum(
        1
        for row_number in range(2, expected_rows + 2)
        if review.cell(row_number, 4).value in PRIORITY_UTTERANCES
    )
    if priority_count != len(PRIORITY_UTTERANCES):
        raise ValueError(f"우선 사례 누락: {priority_count}")
    formula_errors = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and any(
                    token in value
                    for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A")
                ):
                    formula_errors.append(f"{sheet.title}!{cell.coordinate}:{value}")
    if formula_errors:
        raise ValueError(f"수식 오류 문자열 발견: {formula_errors[:10]}")
    return {
        "status": "passed",
        "rows": expected_rows,
        "sheets": wb.sheetnames,
        "hyperlinks": hyperlink_count,
        "data_validations": len(review.data_validations.dataValidation),
        "priority_rows": priority_count,
        "formula_error_strings": len(formula_errors),
        "file_bytes": path.stat().st_size,
    }


def main() -> int:
    args = parse_args()
    index_path = args.index.resolve()
    bundle_root = args.bundle_root.resolve()
    output = args.output.resolve()
    if not index_path.is_file():
        raise FileNotFoundError(index_path)
    if not bundle_root.is_dir():
        raise NotADirectoryError(bundle_root)
    source_headers, rows = read_index(index_path)
    validate_bundle_files(bundle_root, rows)
    output.parent.mkdir(parents=True, exist_ok=True)
    archived = None
    if output.exists():
        if not args.overwrite:
            raise FileExistsError(f"기존 파일 보호: {output}")
        archived = archive_existing(output)

    partial = output.with_name(f".{output.stem}.partial.xlsx")
    if partial.exists():
        partial.unlink()
    try:
        workbook = build_workbook(
            source_headers,
            rows,
            index_path,
            bundle_root,
            output_path=output,
        )
        workbook.save(partial)
        report = verify_workbook(partial, len(rows), bundle_root)
        os.replace(partial, output)
    finally:
        if partial.exists():
            partial.unlink()
    report["output"] = str(output)
    report["archived_previous"] = str(archived) if archived else None
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
