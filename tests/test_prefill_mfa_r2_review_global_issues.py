import csv
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts" / "python"))

from create_mfa_r2_review_workbook import create_workbook  # noqa: E402
from package_mfa_r2_pilot_review import REVIEW_FIELDS  # noqa: E402
from prefill_mfa_r2_review_global_issues import (  # noqa: E402
    GLOBAL_NOTE,
    TABLE_NAME,
    prefill,
)


class PrefillMfaR2ReviewGlobalIssuesTests(unittest.TestCase):
    def make_workbook(self, root: Path) -> Path:
        source = root / "REVIEW.csv"
        with source.open(
            "w", encoding="utf-8-sig", newline=""
        ) as stream:
            writer = csv.DictWriter(
                stream, fieldnames=REVIEW_FIELDS
            )
            writer.writeheader()
            for order in range(1, 61):
                row = {field: "" for field in REVIEW_FIELDS}
                row.update(
                    {
                        "review_order": order,
                        "year": 2020 + (order - 1) // 10,
                        "utt_id": f"U{order}",
                        "speaker_id": f"S{order}",
                        "session_id": f"D{order}",
                        "linkage_status": "미검토",
                        "tier_structure_status": "미검토",
                        "boundary_status": "미검토",
                        "csv_searchability_status": "미검토",
                        "overall_infrastructure_decision": "미검토",
                        "wav_file": f"{order}.wav",
                        "textgrid_file": f"{order}.TextGrid",
                        "lab_file": f"{order}.lab",
                        "csv_file": f"{order}.csv",
                    }
                )
                writer.writerow(row)
        workbook_path = root / "source.xlsx"
        create_workbook(
            source,
            workbook_path,
            manifest_path=root / "template.json",
        )
        workbook = load_workbook(workbook_path)
        try:
            sheet = workbook["검토입력"]
            sheet["F2"] = "통과"
            sheet["G2"] = "문제있음"
            sheet["H2"] = "통과"
            sheet["I2"] = "문제있음"
            sheet["K2"] = "연구자 상세 기록"
            del sheet.tables[TABLE_NAME]
            workbook.save(workbook_path)
        finally:
            workbook.close()
        return workbook_path

    def test_prefills_only_rows_two_through_sixty(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = self.make_workbook(root)
            output = root / "output.xlsx"
            manifest = root / "prefill.json"

            result = prefill(source, output, manifest)

            workbook = load_workbook(output)
            try:
                sheet = workbook["검토입력"]
                self.assertEqual(sheet["F2"].value, "통과")
                self.assertEqual(sheet["G2"].value, "문제있음")
                self.assertEqual(sheet["H2"].value, "통과")
                self.assertEqual(sheet["I2"].value, "문제있음")
                self.assertEqual(sheet["J2"].value, "미검토")
                self.assertEqual(
                    sheet["K2"].value, "연구자 상세 기록"
                )
                for row in range(3, 62):
                    self.assertEqual(
                        sheet.cell(row, 6).value, "미검토"
                    )
                    self.assertEqual(
                        sheet.cell(row, 7).value, "문제있음"
                    )
                    self.assertEqual(
                        sheet.cell(row, 8).value, "미검토"
                    )
                    self.assertEqual(
                        sheet.cell(row, 9).value, "문제있음"
                    )
                    self.assertEqual(
                        sheet.cell(row, 10).value,
                        "수정 후 재검토",
                    )
                    self.assertEqual(
                        sheet.cell(row, 11).value, GLOBAL_NOTE
                    )
                self.assertIn(TABLE_NAME, sheet.tables)
                self.assertEqual(
                    sum(
                        1
                        for row in sheet.iter_rows()
                        for cell in row
                        if cell.hyperlink is not None
                    ),
                    240,
                )
            finally:
                workbook.close()
            self.assertEqual(result["prefilled_rows"], 59)
            self.assertTrue(manifest.is_file())

    def test_refuses_to_overwrite_existing_target_decision(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = self.make_workbook(root)
            workbook = load_workbook(source)
            try:
                workbook["검토입력"]["G3"] = "통과"
                workbook.save(source)
            finally:
                workbook.close()

            with self.assertRaisesRegex(
                RuntimeError, "기존 연구자 입력"
            ):
                prefill(
                    source,
                    root / "output.xlsx",
                    root / "prefill.json",
                )


if __name__ == "__main__":
    unittest.main()
