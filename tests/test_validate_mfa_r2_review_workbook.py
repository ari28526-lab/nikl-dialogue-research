import csv
import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts" / "python"))

from create_mfa_r2_review_workbook import create_workbook  # noqa: E402
from package_mfa_r2_pilot_review import REVIEW_FIELDS  # noqa: E402
from pipeline_common import file_fingerprint  # noqa: E402
from validate_mfa_r2_review_workbook import (  # noqa: E402
    validate_review_workbook,
)


class ValidateMfaR2ReviewWorkbookTests(unittest.TestCase):
    def make_template(self, root: Path) -> tuple[Path, Path, Path]:
        source = root / "REVIEW.csv"
        row = {field: "" for field in REVIEW_FIELDS}
        row.update(
            {
                "review_order": "1",
                "year": "2020",
                "utt_id": "SDRW2000000001.1.1.1",
                "speaker_id": "SPK1",
                "session_id": "SDRW2000000001",
                "linkage_status": "미검토",
                "tier_structure_status": "미검토",
                "boundary_status": "미검토",
                "csv_searchability_status": "미검토",
                "overall_infrastructure_decision": "미검토",
                "wav_file": "2020__x.wav",
                "textgrid_file": "2020__x.TextGrid",
                "lab_file": "2020__x.lab",
                "csv_file": "2020__x.csv",
            }
        )
        with source.open(
            "w", encoding="utf-8-sig", newline=""
        ) as stream:
            writer = csv.DictWriter(
                stream, fieldnames=REVIEW_FIELDS
            )
            writer.writeheader()
            writer.writerow(row)
        workbook = root / "REVIEW.xlsx"
        bundle_manifest = root / "BUNDLE_MANIFEST.json"
        bundle_manifest.write_text(
            json.dumps(
                {
                    "schema_version": "mfa_r2_flat_review_bundle.v2",
                    "status": "success",
                    "flat_layout": True,
                    "review_scope": "infrastructure_acceptance_only",
                    "realization_judgment_performed": False,
                    "supporting_files": {
                        "REVIEW.csv": file_fingerprint(
                            source, with_sha256=True
                        )
                    },
                    "machine_gate_evidence": {
                        "year_contracts": {
                            "2020": {
                                "alignment_contract_id": "align",
                                "lab_input_contract_id": "lab",
                            }
                        }
                    },
                }
            ),
            encoding="utf-8",
        )
        create_workbook(source, workbook)
        return source, bundle_manifest, workbook

    def test_approved_review_preserves_links_and_emits_report(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source, bundle_manifest, workbook_path = self.make_template(root)
            workbook = load_workbook(workbook_path)
            sheet = workbook["검토입력"]
            headers = {
                cell.value: cell.column
                for cell in sheet[1]
            }
            for field in (
                "linkage_status",
                "tier_structure_status",
                "boundary_status",
                "csv_searchability_status",
            ):
                sheet.cell(2, headers[field]).value = "통과"
            sheet.cell(
                2, headers["overall_infrastructure_decision"]
            ).value = "인프라 통과"
            workbook.save(workbook_path)
            workbook.close()

            report = validate_review_workbook(
                review_csv=source,
                bundle_manifest_path=bundle_manifest,
                workbook_path=workbook_path,
                decision_csv=root / "decisions.csv",
                report_path=root / "review_report.json",
                reviewer_id="researcher",
            )

            self.assertEqual(report["status"], "approved")
            self.assertTrue(report["allow_bulk_mfa"])
            self.assertFalse(report["realization_judgment_performed"])

    def test_rejects_changed_immutable_file_link(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source, bundle_manifest, workbook_path = self.make_template(root)
            workbook = load_workbook(workbook_path)
            sheet = workbook["검토입력"]
            headers = {
                cell.value: cell.column
                for cell in sheet[1]
            }
            sheet.cell(2, headers["wav_file"]).value = "other.wav"
            workbook.save(workbook_path)
            workbook.close()

            with self.assertRaisesRegex(RuntimeError, "불변 연결 열"):
                validate_review_workbook(
                    review_csv=source,
                    bundle_manifest_path=bundle_manifest,
                    workbook_path=workbook_path,
                    decision_csv=root / "decisions.csv",
                    report_path=root / "review_report.json",
                    reviewer_id="researcher",
                )

    def test_inconsistent_pass_is_invalid(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source, bundle_manifest, workbook_path = self.make_template(root)
            workbook = load_workbook(workbook_path)
            sheet = workbook["검토입력"]
            headers = {
                cell.value: cell.column
                for cell in sheet[1]
            }
            sheet.cell(
                2, headers["overall_infrastructure_decision"]
            ).value = "인프라 통과"
            workbook.save(workbook_path)
            workbook.close()

            report = validate_review_workbook(
                review_csv=source,
                bundle_manifest_path=bundle_manifest,
                workbook_path=workbook_path,
                decision_csv=root / "decisions.csv",
                report_path=root / "review_report.json",
                reviewer_id="researcher",
            )
            self.assertEqual(report["status"], "invalid")
            self.assertFalse(report["allow_bulk_mfa"])


if __name__ == "__main__":
    unittest.main()
