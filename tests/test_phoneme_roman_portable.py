from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
PYTHON_DIR = ROOT / "scripts" / "python"
if str(PYTHON_DIR) not in sys.path:
    sys.path.insert(0, str(PYTHON_DIR))

import make_phoneme_roman_workbook_portable as portable  # noqa: E402


class PhonemeRomanPortableWorkbookTests(unittest.TestCase):
    def test_rewrites_absolute_links_without_overwriting_source(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            source = root / "source.xlsx"
            output = root / "portable.xlsx"
            year = "2020"
            utt_id = "SDRW2000000001.1.1.1"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = portable.REVIEW_SHEET
            sheet.append(["연도", "utt_id", *portable.LINK_HEADERS])
            sheet.append([int(year), utt_id, "WAV", "old", "new"])
            for column in range(3, 6):
                sheet.cell(2, column).hyperlink = f"D:\\private\\{column}"
            workbook.save(source)
            source_before = portable.sha256_file(source)

            for filename in portable.expected_targets(year, utt_id).values():
                (root / filename).write_bytes(b"test")

            manifest = portable.rewrite_links(source, output, root)
            self.assertEqual(manifest["relative_links"], 3)
            self.assertEqual(manifest["missing_link_targets"], 0)
            self.assertEqual(portable.sha256_file(source), source_before)

            check = load_workbook(output, data_only=False, read_only=False)
            row = check[portable.REVIEW_SHEET][2]
            self.assertEqual(
                [cell.hyperlink.target for cell in row[2:5]],
                list(portable.expected_targets(year, utt_id).values()),
            )


if __name__ == "__main__":
    unittest.main()
