from __future__ import annotations

import json
import shutil
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PYTHON_DIR = PROJECT_ROOT / "scripts" / "python"
if str(PYTHON_DIR) not in sys.path:
    sys.path.insert(0, str(PYTHON_DIR))

import apply_common_pron_researcher_decisions as apply  # noqa: E402
import build_common_pron_mfa_lexicon as lexicon  # noqa: E402
import common_pron_no_path_review as no_path  # noqa: E402
import validate_common_pron_researcher_review_xlsx as validator  # noqa: E402
from pipeline_common import file_fingerprint, sha256_file  # noqa: E402


TEMPLATE = (
    PROJECT_ROOT
    / "outputs"
    / "common_pron_r2_review_20260729"
    / "common_pron_r2_researcher_review_20260729_v5.xlsx"
)


def workbook_rows_and_inventory() -> tuple[list[dict], set[str]]:
    workbook = load_workbook(TEMPLATE, data_only=False, read_only=False)
    sheet = workbook["발음검토"]
    rows = []
    inventory: set[str] = set()
    for row_number in range(2, 29):
        row = {
            header: sheet.cell(row_number, column).value
            for column, header in enumerate(
                validator.REVIEW_HEADERS, start=1
            )
        }
        rows.append(row)
        for field in (
            "model_candidate_phone",
            "recommended_phone",
            "alternative_phone",
        ):
            inventory.update(validator.phone_tokens(row[field]))
    workbook.close()
    return rows, inventory


def create_ready_fixture(root: Path) -> dict:
    common_root = root / "common"
    release = (
        common_root / "releases" / "common_pron_mfa_r2_test"
    )
    review_root = release / "03_review"
    review_root.mkdir(parents=True)
    no_path_path = review_root / "g2p_no_path_researcher_review.csv"
    jamo_path = review_root / "jamo_ls_researcher_review.csv"
    workbook_rows, inventory = workbook_rows_and_inventory()

    no_path_rows = [
        {
            "surface": "읊어",
            "respelled": "을퍼",
            "rule_id": "standard_pron_rule14_eulph_v1",
            "evidence_source": "fixture",
            "evidence_detail": "fixture",
            "pron_phones_mfa": "ɨ ɭ pʰ ʌ",
            "approved_pron_phones_mfa": "ɨ ɭ pʰ ʌ",
            "approved_phone_evidence": "legacy_fixture",
            "decision": "approved",
            "notes": "preserved fixture approval",
        }
    ]
    for row in workbook_rows:
        if row["category"] != "no_path":
            continue
        no_path_rows.append(
            {
                "surface": row["token"],
                "respelled": row["model_input"],
                "rule_id": "fixture_rule",
                "evidence_source": "fixture",
                "evidence_detail": "fixture",
                "pron_phones_mfa": row["model_candidate_phone"],
                "approved_pron_phones_mfa": "",
                "approved_phone_evidence": "",
                "decision": "pending",
                "notes": "",
            }
        )
    jamo_rows = [
        {
            "token": row["token"],
            "model_input": row["model_input"],
            "pron_phones_mfa": row["model_candidate_phone"],
            "approved_pron_phones_mfa": "",
            "decision": "pending",
            "evidence_source": "",
            "notes": "",
        }
        for row in workbook_rows
        if row["category"] == "jamo_ls"
    ]
    apply.write_csv(no_path_path, no_path.REVIEW_FIELDS, no_path_rows)
    apply.write_csv(
        jamo_path, lexicon.SPECIAL_REVIEW_FIELDS, jamo_rows
    )

    filled = root / "FILLED.xlsx"
    shutil.copy2(TEMPLATE, filled)
    filled_workbook = load_workbook(
        filled, data_only=False, read_only=False
    )
    filled_sheet = filled_workbook["발음검토"]
    for row_number in range(2, 29):
        filled_sheet[f"R{row_number}"] = "approve_recommended"
        if (
            filled_sheet[f"L{row_number}"].value
            != "accept_model_candidate"
        ):
            filled_sheet[f"U{row_number}"] = "fixture researcher note"
    filled_workbook.save(filled)
    filled_workbook.close()

    template_manifest_path = root / "template.manifest.json"
    template_manifest = {
        "schema_version": validator.WORKBOOK_SCHEMA_VERSION,
        "status": "success",
        "output": file_fingerprint(TEMPLATE, with_sha256=True),
        "inputs": {
            "no_path_review": file_fingerprint(
                no_path_path, with_sha256=True
            ),
            "jamo_review": file_fingerprint(
                jamo_path, with_sha256=True
            ),
        },
    }
    template_manifest_path.write_text(
        json.dumps(template_manifest, ensure_ascii=False, indent=2)
        + "\n",
        encoding="utf-8",
    )
    result = validator.validate_filled_workbook(
        template_path=TEMPLATE,
        filled_path=filled,
        template_manifest=template_manifest,
        inventory=inventory,
    )
    if not result["decision_report"]["ready_for_apply"]:
        raise AssertionError(result["decision_report"])

    decisions_path = root / "validation.decisions.csv"
    corrections_path = root / "validation.corrections.csv"
    apply.write_csv(
        decisions_path, validator.DECISION_FIELDS, result["decisions"]
    )
    apply.write_csv(
        corrections_path,
        validator.CORRECTION_FIELDS,
        result["corrections"],
    )
    validation_manifest_path = root / "validation.manifest.json"
    validation_manifest = {
        "schema_version": validator.SCHEMA_VERSION,
        "status": "ready_for_apply",
        "ready_for_apply": True,
        "inputs": {
            "template_manifest": file_fingerprint(
                template_manifest_path, with_sha256=True
            ),
            "filled_workbook": file_fingerprint(
                filled, with_sha256=True
            ),
        },
        "outputs": {
            "normalized_decisions": file_fingerprint(
                decisions_path, with_sha256=True
            ),
            "correction_registry": file_fingerprint(
                corrections_path, with_sha256=True
            ),
        },
    }
    validation_manifest_path.write_text(
        json.dumps(validation_manifest, ensure_ascii=False, indent=2)
        + "\n",
        encoding="utf-8",
    )
    inventory_contract = validator.phone_inventory_contract(inventory)
    return {
        "common_root": common_root,
        "release": release,
        "no_path": no_path_path,
        "jamo": jamo_path,
        "validation_manifest": validation_manifest_path,
        "inventory": inventory,
        "inventory_contract": inventory_contract,
    }


class ApplyCommonPronResearcherDecisionsTests(unittest.TestCase):
    def load_plan(self, fixture: dict):
        return apply.load_application_plan(
            validation_manifest_path=fixture["validation_manifest"],
            no_path_review_path=fixture["no_path"],
            jamo_review_path=fixture["jamo"],
        )

    def frozen_inventory_patch(self, fixture: dict):
        return patch.object(
            apply.validator,
            "load_frozen_inventory",
            return_value=(
                fixture["inventory"],
                fixture["inventory_contract"],
                TEMPLATE,
            ),
        )

    def test_read_only_plan_does_not_create_transaction(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            fixture = create_ready_fixture(Path(tmp))
            with self.frozen_inventory_patch(fixture):
                plan = self.load_plan(fixture)

            self.assertEqual(plan["counts"]["normalized_decisions"], 27)
            self.assertEqual(plan["counts"]["no_path_new_approved"], 23)
            self.assertFalse(
                (
                    fixture["release"]
                    / "03_review"
                    / "decision_transactions"
                ).exists()
            )

    def test_apply_archives_and_promotes_both_ledgers(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            fixture = create_ready_fixture(Path(tmp))
            no_path_before = sha256_file(fixture["no_path"])
            jamo_before = sha256_file(fixture["jamo"])
            with self.frozen_inventory_patch(fixture):
                plan = self.load_plan(fixture)
                manifest = apply.apply_plan(
                    plan=plan,
                    release_root=fixture["release"],
                    no_path_review_path=fixture["no_path"],
                    jamo_review_path=fixture["jamo"],
                )

            self.assertEqual(manifest["status"], "applied")
            self.assertNotEqual(
                sha256_file(fixture["no_path"]), no_path_before
            )
            self.assertNotEqual(sha256_file(fixture["jamo"]), jamo_before)
            self.assertFalse(
                (
                    fixture["common_root"]
                    / "locks"
                    / f"{fixture['release'].name}.lock"
                ).exists()
            )
            archive = (
                fixture["release"]
                / "03_review"
                / "decision_transactions"
                / manifest["transaction_id"]
                / "archive"
            )
            self.assertEqual(
                sha256_file(
                    archive / fixture["no_path"].name
                ),
                no_path_before,
            )
            self.assertEqual(
                sha256_file(archive / fixture["jamo"].name),
                jamo_before,
            )
            no_path_rows = no_path.read_review(fixture["no_path"])
            jamo_rows = lexicon.read_special_review(fixture["jamo"])
            self.assertEqual(
                sum(row["decision"] == "approved" for row in no_path_rows),
                24,
            )
            self.assertEqual(
                sum(row["decision"] == "approved" for row in jamo_rows),
                4,
            )
            self.assertEqual(
                [row["surface"] for row in no_path_rows if row["surface"] == "읊어"],
                ["읊어"],
            )
            self.assertTrue(
                Path(
                    manifest["outputs"]["correction_registry"]["path"]
                ).is_file()
            )
            with self.frozen_inventory_patch(fixture):
                repeated = apply.apply_plan(
                    plan=plan,
                    release_root=fixture["release"],
                    no_path_review_path=fixture["no_path"],
                    jamo_review_path=fixture["jamo"],
                )
            self.assertEqual(repeated, manifest)

    def test_existing_runner_lock_blocks_apply_without_changes(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            fixture = create_ready_fixture(Path(tmp))
            no_path_before = sha256_file(fixture["no_path"])
            jamo_before = sha256_file(fixture["jamo"])
            lock = (
                fixture["common_root"]
                / "locks"
                / f"{fixture['release'].name}.lock"
            )
            lock.parent.mkdir(parents=True)
            lock.write_text('{"kind":"runner"}\n', encoding="utf-8")
            with self.frozen_inventory_patch(fixture):
                plan = self.load_plan(fixture)
                with self.assertRaisesRegex(
                    apply.DecisionApplicationError, "lock exists"
                ):
                    apply.apply_plan(
                        plan=plan,
                        release_root=fixture["release"],
                        no_path_review_path=fixture["no_path"],
                        jamo_review_path=fixture["jamo"],
                    )

            self.assertEqual(
                sha256_file(fixture["no_path"]), no_path_before
            )
            self.assertEqual(sha256_file(fixture["jamo"]), jamo_before)

    def test_second_promotion_failure_rolls_back_both_ledgers(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            fixture = create_ready_fixture(Path(tmp))
            no_path_before = sha256_file(fixture["no_path"])
            jamo_before = sha256_file(fixture["jamo"])
            original_promote = apply.promote_copy
            calls = {"count": 0}

            def fail_second(source, destination, txid):
                calls["count"] += 1
                if calls["count"] == 2:
                    raise OSError("synthetic second promotion failure")
                return original_promote(source, destination, txid)

            with self.frozen_inventory_patch(fixture):
                plan = self.load_plan(fixture)
                with patch.object(
                    apply, "promote_copy", side_effect=fail_second
                ):
                    with self.assertRaisesRegex(
                        OSError, "synthetic second promotion failure"
                    ):
                        apply.apply_plan(
                            plan=plan,
                            release_root=fixture["release"],
                            no_path_review_path=fixture["no_path"],
                            jamo_review_path=fixture["jamo"],
                        )

            self.assertEqual(
                sha256_file(fixture["no_path"]), no_path_before
            )
            self.assertEqual(sha256_file(fixture["jamo"]), jamo_before)
            failure_manifests = list(
                fixture["release"].glob(
                    "03_review/decision_transactions/*/"
                    "failure_manifest.json"
                )
            )
            self.assertEqual(len(failure_manifests), 1)
            failure = json.loads(
                failure_manifests[0].read_text(encoding="utf-8")
            )
            self.assertEqual(failure["status"], "rolled_back")
            self.assertFalse(
                (
                    fixture["common_root"]
                    / "locks"
                    / f"{fixture['release'].name}.lock"
                ).exists()
            )

    def test_changed_current_ledger_is_rejected_before_write(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            fixture = create_ready_fixture(Path(tmp))
            with fixture["no_path"].open(
                "a", encoding="utf-8"
            ) as stream:
                stream.write("\n")
            with self.frozen_inventory_patch(fixture):
                with self.assertRaisesRegex(
                    apply.DecisionApplicationError,
                    "current no-path ledger fingerprint changed",
                ):
                    self.load_plan(fixture)
            self.assertFalse(
                (
                    fixture["release"]
                    / "03_review"
                    / "decision_transactions"
                ).exists()
            )


if __name__ == "__main__":
    unittest.main()
