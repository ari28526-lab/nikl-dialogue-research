from __future__ import annotations

import csv
import json
import shutil
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PYTHON_DIR = PROJECT_ROOT / "scripts" / "python"
if str(PYTHON_DIR) not in sys.path:
    sys.path.insert(0, str(PYTHON_DIR))

import validate_common_pron_researcher_review_xlsx as validator  # noqa: E402


TEMPLATE = (
    PROJECT_ROOT
    / "outputs"
    / "common_pron_r2_review_20260729"
    / "common_pron_r2_researcher_review_20260729_v5.xlsx"
)
TEMPLATE_MANIFEST = TEMPLATE.with_suffix(".manifest.json")


def template_inventory() -> set[str]:
    workbook = load_workbook(TEMPLATE, data_only=False, read_only=False)
    sheet = workbook["발음검토"]
    phones: set[str] = set()
    for row in range(2, 29):
        for column in ("G", "I", "K"):
            phones.update(
                validator.phone_tokens(sheet[f"{column}{row}"].value)
            )
    workbook.close()
    return phones | {"sil", "spn"}


def fill_all_recommended(path: Path) -> None:
    workbook = load_workbook(path, data_only=False, read_only=False)
    sheet = workbook["발음검토"]
    for row in range(2, 29):
        sheet[f"R{row}"] = "approve_recommended"
        if sheet[f"L{row}"].value != "accept_model_candidate":
            sheet[f"U{row}"] = "원음과 제시 근거를 확인하고 권고안을 승인함"
    workbook.save(path)
    workbook.close()


class ValidateCommonPronResearcherReviewWorkbookTests(
    unittest.TestCase
):
    @classmethod
    def setUpClass(cls) -> None:
        if not TEMPLATE.is_file() or not TEMPLATE_MANIFEST.is_file():
            raise unittest.SkipTest("tracked v5 review template is missing")
        cls.template_manifest = json.loads(
            TEMPLATE_MANIFEST.read_text(encoding="utf-8")
        )
        cls.inventory = template_inventory()

    def test_clean_template_copy_is_valid_but_incomplete(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            filled = Path(tmp) / "filled.xlsx"
            shutil.copy2(TEMPLATE, filled)

            result = validator.validate_filled_workbook(
                template_path=TEMPLATE,
                filled_path=filled,
                template_manifest=self.template_manifest,
                inventory=self.inventory,
            )

            self.assertEqual(
                result["decision_report"]["status"],
                "incomplete_pending",
            )
            self.assertFalse(
                result["decision_report"]["ready_for_apply"]
            )
            self.assertEqual(
                result["decision_report"]["pending_decisions"], 27
            )
            self.assertGreater(
                result["immutable_contract"]["cells_compared"], 500
            )

    def test_all_recommended_decisions_normalize_with_two_corrections(
        self,
    ) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            filled = Path(tmp) / "filled.xlsx"
            shutil.copy2(TEMPLATE, filled)
            fill_all_recommended(filled)

            result = validator.validate_filled_workbook(
                template_path=TEMPLATE,
                filled_path=filled,
                template_manifest=self.template_manifest,
                inventory=self.inventory,
            )

            self.assertTrue(
                result["decision_report"]["ready_for_apply"]
            )
            self.assertEqual(len(result["decisions"]), 27)
            self.assertEqual(len(result["corrections"]), 2)
            by_token = {
                row["token"]: row for row in result["decisions"]
            }
            self.assertEqual(
                by_token["읊고"]["approved_pron_phones_mfa"],
                "ɨ p̚ k͈ o",
            )
            self.assertEqual(
                by_token["읊고"]["approved_phone_provenance"],
                "researcher_workbook_manual_same_inventory",
            )
            correction_tokens = {
                row["token"] for row in result["corrections"]
            }
            self.assertEqual(
                correction_tokens, {"외곬수적인", "천구백칤비육"}
            )

    def test_immutable_cell_change_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            filled = Path(tmp) / "filled.xlsx"
            shutil.copy2(TEMPLATE, filled)
            workbook = load_workbook(
                filled, data_only=False, read_only=False
            )
            workbook["발음검토"]["C2"] = "변경"
            workbook.save(filled)
            workbook.close()

            with self.assertRaisesRegex(
                validator.WorkbookIntegrityError,
                r"immutable cell changed: 발음검토!C2",
            ):
                validator.validate_filled_workbook(
                    template_path=TEMPLATE,
                    filled_path=filled,
                    template_manifest=self.template_manifest,
                    inventory=self.inventory,
                )

    def test_custom_phone_outside_inventory_is_not_ready(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            filled = Path(tmp) / "filled.xlsx"
            shutil.copy2(TEMPLATE, filled)
            workbook = load_workbook(
                filled, data_only=False, read_only=False
            )
            sheet = workbook["발음검토"]
            sheet["R2"] = "approve_custom"
            sheet["S2"] = "NOT_A_PHONE"
            sheet["U2"] = "검증용"
            workbook.save(filled)
            workbook.close()

            result = validator.validate_filled_workbook(
                template_path=TEMPLATE,
                filled_path=filled,
                template_manifest=self.template_manifest,
                inventory=self.inventory,
            )

            self.assertEqual(
                result["decision_report"]["status"],
                "invalid_decisions",
            )
            self.assertFalse(
                result["decision_report"]["ready_for_apply"]
            )
            self.assertIn(
                "outside acoustic inventory",
                result["decision_report"]["errors"][0],
            )

    def test_clean_template_cannot_be_used_as_filled_path(self) -> None:
        with self.assertRaisesRegex(
            validator.WorkbookIntegrityError, "Save As copy"
        ):
            validator.validate_filled_workbook(
                template_path=TEMPLATE,
                filled_path=TEMPLATE,
                template_manifest=self.template_manifest,
                inventory=self.inventory,
            )

    def test_alignment_symbols_are_not_lexical_inventory(self) -> None:
        self.assertEqual(
            validator.lexical_phone_inventory(
                {"ɨ", "p̚", "sil", "spn"}
            ),
            {"ɨ", "p̚"},
        )

    def test_ready_workbook_writes_normalized_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            filled = root / "filled.xlsx"
            shutil.copy2(TEMPLATE, filled)
            fill_all_recommended(filled)
            output_manifest = root / "validation.manifest.json"
            inventory_contract = validator.phone_inventory_contract(
                self.inventory
            )
            with patch.object(
                validator,
                "load_frozen_inventory",
                return_value=(
                    self.inventory,
                    inventory_contract,
                    TEMPLATE,
                ),
            ):
                manifest = validator.build_validation_outputs(
                    template_path=TEMPLATE,
                    filled_path=filled,
                    template_manifest_path=TEMPLATE_MANIFEST,
                    output_manifest_path=output_manifest,
                )

            self.assertTrue(manifest["ready_for_apply"])
            decisions_path = root / "validation.decisions.csv"
            corrections_path = root / "validation.corrections.csv"
            with decisions_path.open(
                "r", encoding="utf-8-sig", newline=""
            ) as stream:
                decisions = list(csv.DictReader(stream))
            with corrections_path.open(
                "r", encoding="utf-8-sig", newline=""
            ) as stream:
                corrections = list(csv.DictReader(stream))
            self.assertEqual(len(decisions), 27)
            self.assertEqual(len(corrections), 2)
            self.assertEqual(
                json.loads(output_manifest.read_text(encoding="utf-8"))[
                    "status"
                ],
                "ready_for_apply",
            )


if __name__ == "__main__":
    unittest.main()
