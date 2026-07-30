import csv
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts" / "python"))

from create_mfa_r2_review_workbook import create_workbook  # noqa: E402
from package_mfa_r2_pilot_review import REVIEW_FIELDS  # noqa: E402


class ReviewWorkbookTests(unittest.TestCase):
    def test_creates_two_sheet_validated_review_workbook(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "REVIEW.csv"
            output = root / "REVIEW.xlsx"
            row = {field: "" for field in REVIEW_FIELDS}
            row.update(
                {
                    "review_order": 1,
                    "year": 2020,
                    "utt_id": "SDRW2000000001.1.1.1",
                    "linkage_status": "미검토",
                    "tier_structure_status": "미검토",
                    "boundary_status": "미검토",
                    "csv_searchability_status": "미검토",
                    "overall_infrastructure_decision": "미검토",
                    "wav_file": "2020__x.wav",
                    "textgrid_file": "2020__x.TextGrid",
                    "lab_file": "2020__x.lab",
                    "csv_file": "2020__x.csv",
                }
            )
            with source.open(
                "w", encoding="utf-8-sig", newline=""
            ) as stream:
                writer = csv.DictWriter(
                    stream, fieldnames=REVIEW_FIELDS
                )
                writer.writeheader()
                writer.writerow(row)

            manifest = create_workbook(source, output)

            workbook = load_workbook(output)
            try:
                self.assertEqual(workbook.sheetnames, ["검토입력", "안내"])
                self.assertEqual(workbook["검토입력"].max_row, 2)
                self.assertEqual(
                    workbook["검토입력"]["A2"].value, 1
                )
                self.assertEqual(
                    workbook["검토입력"]["L2"].hyperlink.target,
                    "2020__x.wav",
                )
            finally:
                workbook.close()
            self.assertEqual(manifest["status"], "success")
            self.assertEqual(manifest["rows"], 1)
            self.assertTrue(
                (
                    root / "REVIEW_XLSX_TEMPLATE_MANIFEST.json"
                ).is_file()
            )


if __name__ == "__main__":
    unittest.main()
